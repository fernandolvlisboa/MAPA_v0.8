#!/usr/bin/env python
"""
Corrige o desalinhamento das colunas de ano no Template GT.

## O problema

As fórmulas de ``BP_GT``/``DRE_GT`` somam a coluna de valores da aba oculta
**uma posição à frente** do que o cabeçalho declara::

    BP_GT!D7 = "2021"  ->  SUMIFS(_dados_padronizados!D:D, ...)
    mas _dados_padronizados!D1 = "2022"

Efeitos: a coluna ``C`` (2021) nunca é lida, a coluna ``H`` referenciada por
2025 não existe (sempre zero), e todo valor aparece deslocado um ano.

``docs/TEMPLATE_GT_BP.md`` §4.1 define ``C=2021 … G=2025``, portanto o
cabeçalho está certo e as **fórmulas** é que precisam andar uma coluna à
esquerda: ``D:D -> C:C``, ``E:E -> D:D``, …, ``H:H -> G:G``.

A coluna de critério ``$A:$A`` (códigos) nunca é tocada.

## Uso

    python scripts/fix_template_year_alignment.py --dry-run   # mostra o que mudaria
    python scripts/fix_template_year_alignment.py             # aplica (faz .bak antes)
    python scripts/fix_template_year_alignment.py --verify    # só confere o alinhamento

## Como voltar atrás

Ver ``docs/MIGRACAO_TEMPLATE.md``. Resumo: o script grava
``<template>.bak`` antes de escrever; restaurar é copiar o .bak por cima. O
template também está versionado no git (``git checkout -- templates/``).
"""

from __future__ import annotations

import argparse
import re
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook

TEMPLATE = Path("templates/Template_GT_BP_Padrao_v3.xlsx")
DATA_SHEET = "_dados_padronizados"
SHEETS = ("BP_GT", "DRE_GT")

#: Deslocamento de uma coluna à esquerda nas referências de valor.
SHIFT = {"D": "C", "E": "D", "F": "E", "G": "F", "H": "G"}

#: Só referências de VALOR (sem "$"). O critério é sempre "$A:$A" e fica intacto.
_REF_RE = re.compile(rf"{DATA_SHEET}!([D-H]):\1(?![\w$])")


def _shift_formula(formula: str) -> str:
    return _REF_RE.sub(lambda m: f"{DATA_SHEET}!{SHIFT[m.group(1)]}:{SHIFT[m.group(1)]}", formula)


def apply_fix(template: Path, dry_run: bool = False) -> int:
    wb = load_workbook(template)
    alteradas = 0
    for sheet in SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                novo = _shift_formula(cell.value)
                if novo != cell.value:
                    alteradas += 1
                    if dry_run and alteradas <= 3:
                        print(f"  {sheet}!{cell.coordinate}")
                        print(f"    antes : {cell.value[:88]}")
                        print(f"    depois: {novo[:88]}")
                    if not dry_run:
                        cell.value = novo
    if dry_run:
        print(f"\n[dry-run] {alteradas} fórmula(s) seriam alteradas. Nada foi escrito.")
        return alteradas

    backup = template.with_suffix(template.suffix + ".bak")
    if not backup.exists():
        shutil.copy(template, backup)
        print(f"backup: {backup}")
    wb.save(template)
    print(f"aplicado: {alteradas} fórmula(s) corrigidas em {template}")
    return alteradas


def verify(template: Path) -> bool:
    """Confere que cada coluna de ano lê a coluna de mesmo ano na aba de dados."""
    wb = load_workbook(template)
    data_headers = {}
    ws_data = wb[DATA_SHEET]
    for col in range(1, ws_data.max_column + 1):
        header = ws_data.cell(1, col).value
        if header and str(header).strip().isdigit():
            data_headers[ws_data.cell(1, col).column_letter] = str(header).strip()

    ok = True
    for sheet in SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        anos = {}
        for col in range(4, 12):
            v = ws.cell(7, col).value
            if v and str(v).strip().isdigit():
                anos[ws.cell(7, col).column_letter] = str(v).strip()

        for letra, ano in anos.items():
            col_idx = ws[f"{letra}7"].column
            ref = None
            for row in range(8, ws.max_row + 1):
                f = ws.cell(row, col_idx).value
                if isinstance(f, str) and DATA_SHEET in f:
                    m = re.search(rf"{DATA_SHEET}!([A-Z]+):", f)
                    if m:
                        ref = m.group(1)
                        break
            if ref is None:
                continue
            ano_lido = data_headers.get(ref, "(coluna sem cabeçalho)")
            status = "OK " if ano_lido == ano else "ERRO"
            if ano_lido != ano:
                ok = False
            print(f"  [{status}] {sheet}!{letra} (rótulo {ano}) -> {DATA_SHEET}!{ref} = {ano_lido}")
    print("\nALINHAMENTO CORRETO" if ok else "\nDESALINHADO — rode sem --verify para corrigir")
    return ok


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__.split("\n")[1])
    p.add_argument("--template", default=str(TEMPLATE))
    p.add_argument("--dry-run", action="store_true", help="mostra o que mudaria")
    p.add_argument("--verify", action="store_true", help="só confere o alinhamento")
    args = p.parse_args()

    template = Path(args.template)
    if not template.exists():
        print(f"template não encontrado: {template}", file=sys.stderr)
        return 2

    if args.verify:
        return 0 if verify(template) else 1
    apply_fix(template, dry_run=args.dry_run)
    if not args.dry_run:
        print("\nverificando...")
        return 0 if verify(template) else 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
