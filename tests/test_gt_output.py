"""Testes da projeção e da geração no Template GT."""

import re
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.bp.output import TemplateProjector
from src.bp.output.build_gt_output import (
    ANOS_TEMPLATE,
    FonteBalancete,
    build_gt_output,
)

TEMPLATE = Path("templates/Template_GT_BP_Padrao_v3.xlsx")
BALANCETE = Path("data/samples/Balancete Real Life.xlsx")

pytestmark = pytest.mark.skipif(not TEMPLATE.exists(), reason="Template GT ausente")


@pytest.fixture(scope="module")
def projector():
    return TemplateProjector()


# ---------------------------------------------------------------- projeção


def test_prefixos_lidos_do_template(projector):
    """Os prefixos vêm do .xlsx, nunca hardcoded."""
    assert len(projector.prefixes) >= 80
    assert "1.01.01" in projector.prefixes
    assert "3.02.01.01" in projector.prefixes


def test_projeta_codigo_direto(projector):
    r = projector.project("1.01.01.02.01")
    assert r.ok and r.metodo == "direto"
    assert r.codigo_template == "1.01.01.02.01"


def test_projeta_conta_enriquecida(projector):
    """Contas x.90.* não existem na ECF — sem projeção virariam zero."""
    r = projector.project("1.90.01")  # Clientes (enriquecida)
    assert r.ok and r.metodo == "mapa_explicito"
    assert r.codigo_template == "1.01.02.02"
    assert projector.is_captured(r.codigo_template)


def test_normaliza_bloco_paralelo_3_11(projector):
    """3.11.* (Presumido) vira 3.01.* (Real), que é o que o template usa."""
    r = projector.project("3.11.01.07.01.02")
    assert r.ok and r.metodo.startswith("normalizado_3_11")
    assert r.codigo_template.startswith("3.01.")


def test_sobe_hierarquia_quando_nao_ha_linha_propria(projector):
    r = projector.project("1.01.03.07.04")  # Outros Estoques
    assert r.ok
    assert projector.is_captured(r.codigo_template)


def test_recusa_agrupador_generico(projector):
    """Subtotal do cliente ('ATIVO CIRCULANTE') não vira linha — dupla contagem."""
    r = projector.project("1.01")
    assert not r.ok and r.metodo == "generico_demais"


def test_recusa_codigo_invalido(projector):
    assert not projector.project("1.01 - Caixa").ok
    assert not projector.project("").ok
    assert not projector.project(None).ok


def test_sinais_derivados_do_template(projector):
    """Sinal vem do rótulo da coluna B: '(-)' negativa, resto positiva."""
    assert projector.sign_for("1.01.01") == 1  # Ativo
    assert projector.sign_for("2.01.01.03") == 1  # Passivo (check exige +)
    assert projector.sign_for("3.01.01.01.01.04") == 1  # Receita bruta
    assert projector.sign_for("3.01.01.03") == -1  # (-) Custos
    assert projector.sign_for("3.01.01.07.01.02") == -1  # (-) Despesas pessoal
    assert projector.sign_for("3.02.01.01") == -1  # (-) IRPJ/CSLL


# ---------------------------------------------------------------- build


@pytest.fixture(scope="module")
def built(tmp_path_factory):
    if not BALANCETE.exists():
        pytest.skip("balancete de exemplo ausente")
    out = tmp_path_factory.mktemp("gt") / "cliente_2024.xlsx"
    result = build_gt_output(
        BALANCETE, output_path=out, ano_base=2024,
        nome_cliente="Cliente Teste", data_base="2024-12-31",
    )
    return result, out


def test_gera_arquivo_com_abas_esperadas(built):
    _, out = built
    wb = load_workbook(out)
    for aba in ("Sumário", "BP_GT", "DRE_GT", "Contas Tratadas",
                "Contas Não Identificadas", "_dados_padronizados"):
        assert aba in wb.sheetnames, f"aba {aba} ausente"


def test_template_original_nao_e_modificado(built):
    """O template é copiado, nunca sobrescrito."""
    assert TEMPLATE.exists()
    wb = load_workbook(TEMPLATE)
    assert wb["_dados_padronizados"].max_row == 1  # continua só com header


def test_formulas_preservadas(built):
    """Abrir sem data_only mantém as fórmulas SUMIFS vivas."""
    _, out = built
    wb = load_workbook(out)
    formula = wb["BP_GT"]["D9"].value
    assert isinstance(formula, str) and formula.startswith("=IFERROR(SUMIFS(")


def test_nome_cliente_escrito(built):
    _, out = built
    wb = load_workbook(out)
    assert wb["BP_GT"]["B4"].value == "Cliente Teste"
    assert wb["DRE_GT"]["B4"].value == "Cliente Teste"


def test_dados_padronizados_oculta_e_povoada(built):
    result, out = built
    wb = load_workbook(out)
    ws = wb["_dados_padronizados"]
    assert ws.sheet_state == "hidden"
    assert ws.max_row - 1 == result.linhas_escritas > 0


def test_codigos_escritos_no_formato_ecf(built):
    _, out = built
    ws = load_workbook(out)["_dados_padronizados"]
    rx = re.compile(r"^\d+(\.\d+)*$")
    for r in range(2, ws.max_row + 1):
        cod = ws.cell(r, 1).value
        assert cod and rx.fullmatch(str(cod)), f"código inválido: {cod!r}"


def test_todo_valor_escrito_e_capturado_por_alguma_linha(built):
    """Regressão do bug central: valor escrito fora dos prefixos vira ZERO."""
    _, out = built
    proj = TemplateProjector()
    ws = load_workbook(out)["_dados_padronizados"]
    orfas = [
        ws.cell(r, 1).value
        for r in range(2, ws.max_row + 1)
        if ws.cell(r, 1).value and not proj.is_captured(str(ws.cell(r, 1).value))
    ]
    assert not orfas, f"{len(orfas)} linha(s) órfã(s) — não somadas por ninguém: {orfas[:5]}"


def test_anos_do_template_conferem():
    """Anos de fábrica do template — não são um limite, só o default."""
    assert ANOS_TEMPLATE == (2021, 2022, 2023, 2024, 2025)


# ------------------------------------------------- anos flexíveis / multi-arquivo


def _slots_e_headers(path):
    """Devolve (ano -> coluna lida pelo BP_GT, ano -> coluna declarada nos dados)."""
    from openpyxl.utils import column_index_from_string

    wb = load_workbook(path)
    ws = wb["BP_GT"]
    ref_re = re.compile(r"_dados_padronizados!([A-Z]+):")
    lidos = {}
    for col in range(1, ws.max_column + 1):
        rot = ws.cell(7, col).value
        if not (rot and str(rot).strip().isdigit()):
            continue
        for row in range(8, ws.max_row + 1):
            f = ws.cell(row, col).value
            if isinstance(f, str) and (m := ref_re.search(f)):
                lidos[str(rot)] = column_index_from_string(m.group(1))
                break
    wd = wb["_dados_padronizados"]
    declarados = {
        str(wd.cell(1, c).value).strip(): c
        for c in range(1, wd.max_column + 1)
        if wd.cell(1, c).value and str(wd.cell(1, c).value).strip().isdigit()
    }
    return lidos, declarados


def test_template_alinhado_ano_a_ano():
    """Regressão do bug: cada coluna de ano deve ler a coluna do MESMO ano."""
    lidos, declarados = _slots_e_headers(TEMPLATE)
    assert lidos, "nenhuma coluna de ano encontrada"
    for ano, col in declarados.items():
        assert lidos.get(ano) == col, (
            f"{ano}: BP_GT lê coluna {lidos.get(ano)}, dados declaram {col}"
        )


def test_ano_fora_do_padrao_do_template(tmp_path):
    """O template não fica preso a 2021-2025."""
    out = tmp_path / "b2019.xlsx"
    r = build_gt_output(BALANCETE, out, ano_base=2019)
    assert r.anos == (2019,)
    wb = load_workbook(out)
    assert wb["BP_GT"].cell(7, 4).value == "2019"
    assert wb["DRE_GT"].cell(7, 4).value == "2019"
    # slots não usados ficam vazios — não exibir exercício sem dado
    assert wb["BP_GT"].cell(7, 5).value is None


def test_serie_historica_multi_arquivo(tmp_path):
    """Um arquivo por exercício — como os balancetes existem no mundo real."""
    out = tmp_path / "serie.xlsx"
    r = build_gt_output(
        [FonteBalancete(BALANCETE, 2018), FonteBalancete(BALANCETE, 2019)],
        out, nome_cliente="Série",
    )
    assert r.anos == (2018, 2019)
    assert set(r.por_ano) == {2018, 2019}
    lidos, declarados = _slots_e_headers(out)
    assert set(declarados) == {"2018", "2019"}
    for ano, col in declarados.items():
        assert lidos[ano] == col


def test_cada_ano_escreve_na_sua_coluna(tmp_path):
    """Valores de 2018 não podem vazar para a coluna de 2019."""
    out = tmp_path / "duplo.xlsx"
    build_gt_output(
        [FonteBalancete(BALANCETE, 2018), FonteBalancete(BALANCETE, 2019)], out
    )
    wd = load_workbook(out)["_dados_padronizados"]
    _, declarados = _slots_e_headers(out)
    c18, c19 = declarados["2018"], declarados["2019"]
    # cada linha preenche exatamente uma das duas colunas de ano
    for r in range(2, wd.max_row + 1):
        if not wd.cell(r, 1).value:
            continue
        preenchidas = [c for c in (c18, c19) if wd.cell(r, c).value is not None]
        assert len(preenchidas) == 1, f"linha {r} preenche {len(preenchidas)} anos"


def test_recusa_anos_repetidos(tmp_path):
    with pytest.raises(ValueError, match="mais de um arquivo"):
        build_gt_output(
            [FonteBalancete(BALANCETE, 2024), FonteBalancete(BALANCETE, 2024)],
            tmp_path / "x.xlsx",
        )


def test_recusa_mais_anos_que_o_template_comporta(tmp_path):
    fontes = [FonteBalancete(BALANCETE, 2018 + i) for i in range(6)]
    with pytest.raises(ValueError, match="comporta"):
        build_gt_output(fontes, tmp_path / "x.xlsx")


def test_exige_ano_base_para_caminho_unico(tmp_path):
    with pytest.raises(ValueError, match="ano_base"):
        build_gt_output(BALANCETE, tmp_path / "x.xlsx")


def test_arquivo_inexistente_falha_cedo(tmp_path):
    with pytest.raises(FileNotFoundError):
        build_gt_output("nao_existe.xlsx", tmp_path / "x.xlsx", ano_base=2024)


def test_nao_polui_o_cache_compartilhado(tmp_path):
    """
    Gerar a entrega de um cliente não pode escrever no cache do projeto.

    Regressão: o matcher persiste toda decisão automática, então um nome
    próprio de empresa casado por acaso viraria entrada permanente com score
    1.0 (visto de fato: "tutors consultoria de valores mobiliarios ltda" ->
    conta de hedge), envenenando as execuções seguintes.
    """
    compartilhado = Path("data/match_cache.json")
    antes = compartilhado.read_bytes() if compartilhado.exists() else None

    build_gt_output(BALANCETE, tmp_path / "saida.xlsx", ano_base=2024)

    depois = compartilhado.read_bytes() if compartilhado.exists() else None
    assert antes == depois, "o build alterou data/match_cache.json"


def test_cache_explicito_e_respeitado(tmp_path):
    """Quem quiser reaproveitar decisões entre execuções passa o caminho."""
    cache = tmp_path / "meu_cache.json"
    build_gt_output(BALANCETE, tmp_path / "s.xlsx", ano_base=2024, cache_path=cache)
    assert cache.exists() and cache.stat().st_size > 2


def test_resiste_a_template_desalinhado(tmp_path):
    """
    O código segue as FÓRMULAS, não os rótulos.

    Se o template voltar ao estado desalinhado (rollback) ou for reeditado no
    Excel deslocando colunas, a escrita deve acompanhar — sem órfãs e sem
    cabeçalho de ano sobrando numa coluna que ninguém lê.
    """
    shift = {"C": "D", "D": "E", "E": "F", "F": "G", "G": "H"}
    ref_re = re.compile(r"_dados_padronizados!([C-G]):\1(?![\w$])")

    torto = tmp_path / "template_torto.xlsx"
    wb = load_workbook(TEMPLATE)
    for sheet in ("BP_GT", "DRE_GT"):
        for row in wb[sheet].iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = ref_re.sub(
                        lambda m: f"_dados_padronizados!{shift[m.group(1)]}:{shift[m.group(1)]}",
                        cell.value,
                    )
    wb.save(torto)

    out = tmp_path / "saida_torta.xlsx"
    build_gt_output(BALANCETE, out, ano_base=2024, template_path=torto)

    proj = TemplateProjector(template_path=torto)
    wd = load_workbook(out)["_dados_padronizados"]
    # nenhuma linha órfã
    orfas = [
        wd.cell(r, 1).value
        for r in range(2, wd.max_row + 1)
        if wd.cell(r, 1).value and not proj.is_captured(str(wd.cell(r, 1).value))
    ]
    assert not orfas
    # exatamente um cabeçalho de ano, e é o que pedimos
    anos_no_header = [
        str(wd.cell(1, c).value).strip()
        for c in range(1, wd.max_column + 1)
        if wd.cell(1, c).value and str(wd.cell(1, c).value).strip().isdigit()
    ]
    assert anos_no_header == ["2024"], f"cabeçalhos de ano inesperados: {anos_no_header}"
