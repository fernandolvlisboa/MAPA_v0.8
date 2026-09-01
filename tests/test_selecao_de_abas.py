"""
Escolha de aba: pasta de trabalho tem vários exercícios, e quem decide é o analista.

O problema
----------

Balancete de cliente nem sempre é um arquivo com um período. Vem também como
pasta de trabalho::

    SmartRio Balancetes (2020 2026).xlsx    ->  Balancetes 2020 … Balancetes 2026
    Mascara Balancete Core Jun-26.xlsx      ->  Balancete Dez-2024 … mensal Jun-2026
    06.2026 - Mascara PCH.xlsx              ->  20 abas, incluindo "Balancete"

O leitor devolvia a **primeira** aba que passasse no portão — e num arquivo
cuja aba 0 é "Output Modelo (BP)" isso rende zero contas, com nove abas de
balancete ao lado.

Três degraus de decisão
-----------------------

1. **Escolha do analista** (``aba=``) manda, e nada a sobrepõe.
2. **Nome inequívoco** — uma aba chamada exatamente "Balancete" é declaração
   do próprio arquivo. Vale mais que contagem: numa pasta real "Balancete" e
   "Balancete (2)" rendem 2.275 e 1.869 contas, e "a maior" escolheria errado.
3. **Varredura por resultado**, só quando a leitura normal foi pobre.

O teto de cinco exercícios é o do template, e já existia em
``service.MAX_EXERCICIOS``; o que faltava era o mesmo arquivo poder ocupar
mais de um deles.

Referência: ``REVISAO_QUALIDADE.md`` §17.
"""

from __future__ import annotations

import pytest
from conftest import CORPUS_DIR

from src.bp.app import service
from src.bp.parsers.abas import AbaCandidata, listar_abas, periodo_do_nome
from src.bp.parsers.dispatcher import ParseyCaller

MULTI_ABA = "SmartRio Balancetes (2020 2026).xlsx"
COM_ABA_BALANCETE = "06.2026 - Mascara PCH - Balanco.vCore5.xlsx"


def _corpus(nome: str):
    if not CORPUS_DIR.exists():
        pytest.skip(f"corpus ausente: {CORPUS_DIR}")
    caminho = CORPUS_DIR / nome
    if not caminho.exists():
        pytest.skip(f"arquivo ausente neste workspace: {nome}")
    return caminho


# ============================================================================
# 1. Período deduzido do nome da aba
# ============================================================================


@pytest.mark.parametrize(
    ("nome", "esperado"),
    [
        ("Balancetes 2023", (2023, None)),
        ("Balancete Dez-2025", (2025, 12)),
        ("Balancete mensal Jan-2026", (2026, 1)),
        ("Balancete mensal Abr-2026", (2026, 4)),
        ("Output Modelo (BP)", (None, None)),
        ("Plano de contas", (None, None)),
        ("", (None, None)),
    ],
)
def test_periodo_do_nome(nome, esperado):
    assert periodo_do_nome(nome) == esperado


def test_rotulo_do_periodo():
    assert AbaCandidata("x", 10, 2023, None).periodo == "2023"
    assert AbaCandidata("x", 10, 2026, 6).periodo == "06/2026"
    assert AbaCandidata("x", 10, None).periodo == "—"


# ============================================================================
# 2. Descoberta: só pergunta quando há o que perguntar
# ============================================================================


def test_arquivo_de_aba_unica_nao_gera_pergunta():
    """Lista vazia é o gatilho da interface: nada é perguntado."""
    caminho = _corpus("IBH 18_Balancete_06.2026.xlsx")
    assert listar_abas(caminho) == []


def test_formato_sem_abas_nao_gera_pergunta(tmp_path):
    csv = tmp_path / "b.csv"
    csv.write_text("codigo;descricao;saldo\n1.1;Caixa;100\n")
    assert listar_abas(csv) == []


@pytest.mark.integration
def test_descoberta_acha_os_exercicios_de_uma_serie_historica():
    """
    A contagem de contas é **medida**, não estimada — é ela que distingue um
    balancete de uma aba de resumo com nome parecido.
    """
    abas = listar_abas(_corpus(MULTI_ABA))
    assert len(abas) >= 5, f"só {len(abas)} abas de balancete encontradas"

    anos = [a.ano for a in abas if a.ano]
    assert len(anos) == len(abas), "toda aba desta série deveria ter ano no nome"
    assert len(set(anos)) == len(anos), "anos repetidos entre abas"
    assert all(a.contas >= 20 for a in abas), "aba sem contas suficientes na lista"


# ============================================================================
# 3. A escolha explícita manda
# ============================================================================


@pytest.mark.integration
def test_aba_explicita_vence_a_escolha_automatica():
    caminho = _corpus(MULTI_ABA)
    contas_2021 = ParseyCaller(caminho, aba="Balancetes 2021").parse()
    contas_2025 = ParseyCaller(caminho, aba="Balancetes 2025").parse()
    assert contas_2021 and contas_2025
    assert len(contas_2021) != len(contas_2025), (
        "as duas abas renderam o mesmo — a escolha de aba não está sendo usada"
    )


@pytest.mark.integration
def test_aba_chamada_balancete_vence_a_contagem():
    """
    O caso que motivou a regra: "Balancete" (2.275 contas) e "Balancete (2)"
    (1.869) coexistem, e só o nome diz qual é o balancete de verdade.
    """
    caminho = _corpus(COM_ABA_BALANCETE)
    automatico = ParseyCaller(caminho).parse()
    nomeada = ParseyCaller(caminho, aba="Balancete").parse()
    assert automatico, "o arquivo não rendeu conta nenhuma"
    assert len(automatico) == len(nomeada), (
        "a escolha automática não foi para a aba chamada 'Balancete'"
    )


@pytest.mark.parametrize(
    ("nome", "esperado"),
    [
        ("Balancete", 3),
        ("balancete", 3),
        ("Balancete (2)", 2),
        ("Balancete mensal Jun-2026", 2),
        ("Balancetes 2023", 2),
        ("Resumo do balancete", 1),
        ("Plano de contas", 0),
    ],
)
def test_prioridade_do_nome_da_aba(nome, esperado):
    assert ParseyCaller._prioridade_do_nome(nome) == esperado


# ============================================================================
# 4. Uma aba por exercício, com o teto do template
# ============================================================================


def test_entrada_deduz_o_ano_da_aba():
    entrada = service.Entrada("x.xlsx", aba="Balancetes 2023")
    assert entrada.ano == 2023
    assert "Balancetes 2023" in entrada.nome


def test_o_mesmo_arquivo_pode_ocupar_varios_exercicios():
    """É o ponto da mudança: a série histórica cabe num arquivo só."""
    entradas = [
        service.Entrada("s.xlsx", aba=f"Balancetes {ano}")
        for ano in (2022, 2023, 2024, 2025, 2026)
    ]
    assert len(entradas) == service.MAX_EXERCICIOS
    anos = sorted(e.ano for e in entradas)
    assert anos == [2022, 2023, 2024, 2025, 2026]


def test_a_mesma_aba_duas_vezes_e_recusada():
    entradas = [
        service.Entrada("s.xlsx", ano=2023, aba="Balancetes 2023"),
        service.Entrada("s.xlsx", ano=2024, aba="Balancetes 2023"),
    ]
    problemas = service.validar(entradas, "Cliente")
    assert any("mesma aba" in p for p in problemas)


def test_o_teto_de_exercicios_continua_valendo():
    entradas = [
        service.Entrada("s.xlsx", ano=2020 + i, aba=f"Balancetes {2020 + i}")
        for i in range(service.MAX_EXERCICIOS + 1)
    ]
    problemas = service.validar(entradas, "Cliente")
    assert any(str(service.MAX_EXERCICIOS) in p for p in problemas), (
        f"seis exercícios aceitos com teto de {service.MAX_EXERCICIOS}"
    )


# ============================================================================
# 5. A série histórica chega inteira à entrega
# ============================================================================


@pytest.mark.integration
def test_serie_historica_de_um_arquivo_so_chega_a_entrega(tmp_path):
    from openpyxl import load_workbook

    from src.bp.output.build_gt_output import FonteBalancete, build_gt_output

    caminho = _corpus(MULTI_ABA)
    anos = (2022, 2023, 2024, 2025, 2026)
    fontes = [
        FonteBalancete(caminho, ano, aba=f"Balancetes {ano}") for ano in anos
    ]
    resultado = build_gt_output(
        fontes, tmp_path / "serie.xlsx", nome_cliente="Teste",
        cache_path=tmp_path / "c.json",
    )

    assert resultado.anos == anos
    assert len(resultado.por_ano) == len(anos)
    lidas = {ano: v[0] for ano, v in resultado.por_ano.items()}
    assert all(n > 0 for n in lidas.values()), f"exercício vazio: {lidas}"
    assert len(set(lidas.values())) > 1, (
        "todos os exercícios renderam o mesmo número de contas — "
        "provavelmente a mesma aba foi lida cinco vezes"
    )
    assert resultado.captura_integra

    # Uma cópia do original por exercício, cada uma da sua aba.
    abas = load_workbook(resultado.output_path).sheetnames
    for ano in anos:
        assert f"Original {ano}" in abas


# ============================================================================
# 6. Arquivo que não é balancete puro: o programa diz, e pergunta
# ============================================================================
#
# Dois arquivos reais chegaram como "balancete" e não são: a empresa já havia
# feito a consolidação, numa aba "Consolidado" (uma linha por conta do BP, uma
# coluna por empresa) e numa "Output Modelo (BP)" (De-Para em inglês, períodos
# em colunas). O programa lia alguma aba, tirava centenas de contas e
# entregava — sem conseguir conferir NADA contra a origem, porque origem
# hierárquica não havia.
#
# Dizer "não parece um balancete puro; em qual aba está o balanço?" é mais
# honesto e mais útil que escolher sozinho: o trabalho já está feito pelo
# cliente, o que falta é o template.

from src.bp.parsers.abas import BALANCETE, DEMONSTRATIVO, diagnosticar  # noqa: E402

JA_CONSOLIDADO = (
    "06.2026 - Mascara PCH - Balanco.vCore5.xlsx",
    "Mascara Balancete Core Jun-26 2026 Sent to GT.xlsx",
)


@pytest.mark.integration
@pytest.mark.parametrize("nome", JA_CONSOLIDADO)
def test_arquivo_ja_consolidado_e_reconhecido_e_explicado(nome):
    diagnostico = diagnosticar(_corpus(nome))

    assert not diagnostico.e_balancete_puro
    assert diagnostico.precisa_perguntar, (
        "o programa deveria perguntar em qual aba está o balanço"
    )
    assert diagnostico.motivo, "recusou sem dizer por quê"
    assert "hierárquico" in diagnostico.motivo


@pytest.mark.integration
@pytest.mark.parametrize("nome", JA_CONSOLIDADO)
def test_a_pergunta_lista_ate_as_abas_pequenas(nome):
    """
    A resposta pode ser uma aba de 36 linhas — "Consolidado (jun26)" — que o
    filtro normal de balancete descartaria. Perguntar sem oferecer a resposta
    certa seria pior que não perguntar.
    """
    diagnostico = diagnosticar(_corpus(nome))
    nomes = {a.nome for a in diagnostico.abas}
    esperada = "Consolidado (jun26)" if "PCH" in nome else "Output Modelo (BP)"
    assert esperada in nomes, (
        f"a aba onde o cliente pôs o balanço ({esperada}) não foi oferecida. "
        f"Ofertadas: {sorted(nomes)}"
    )


@pytest.mark.integration
def test_balancete_de_verdade_nao_dispara_a_pergunta():
    """Não-vacuidade: a pergunta não pode aparecer para todo mundo."""
    diagnostico = diagnosticar(_corpus("IBH 18_Balancete_06.2026.xlsx"))
    assert diagnostico.e_balancete_puro
    assert not diagnostico.precisa_perguntar
    assert not diagnostico.motivo


@pytest.mark.integration
def test_serie_historica_e_balancete_puro_e_pergunta_exercicios():
    """
    O outro caminho da mesma tela: há balancete, o que falta é escolher quais
    exercícios — não onde está o balanço.
    """
    diagnostico = diagnosticar(_corpus(MULTI_ABA))
    assert diagnostico.e_balancete_puro
    assert not diagnostico.precisa_perguntar
    assert any(a.tipo == BALANCETE for a in diagnostico.abas)


def test_classificacao_da_aba():
    from src.bp.parsers.abas import OUTRA, AbaCandidata

    assert AbaCandidata("x", 500, 2024, tem_hierarquia=True).tipo == BALANCETE
    assert AbaCandidata("x", 500, 2024).tipo == DEMONSTRATIVO
    assert AbaCandidata("x", 3, None).tipo == OUTRA


def test_aba_unica_ilegivel_diz_o_motivo(tmp_path):
    """Sem abas para listar, o veredito vem do próprio arquivo — medido, não
    presumido. Dizer "não é balancete" por falta de abas seria um 'não' que
    ninguém mediu."""
    quebrado = tmp_path / "x.xlsx"
    quebrado.write_bytes(b"isto nao e uma planilha")
    diagnostico = diagnosticar(quebrado)
    assert not diagnostico.e_balancete_puro
    assert diagnostico.motivo
    assert not diagnostico.precisa_perguntar, "não há aba para oferecer"
