"""
Receita não pode virar despesa — o Plano C que faltava, um nível abaixo.

O defeito
---------

Numa entrega real, a receita de serviços de R$ 4.937.529,00 apareceu como::

    3.01.01.03.01.03   Servicos prestados - mercado interno   -4.937,53

``3.01.01.03`` é a linha **"(-) Custos dos produtos, mercadorias e serviços
vendidos"**. A receita entrou como custo negativo: erro duplo — ela some da
DRE **e** o custo infla pelo mesmo valor, de modo que o resultado do exercício
erra por duas vezes o valor da conta.

Por que nada pegou
------------------

O Plano C restringe por classe (ATIVO/PASSIVO/RESULTADO). Origem e destino
eram **ambos RESULTADO**, então não havia o que restringir. Pior: o cache do
projeto tinha gravado ``servicos prestados -> (-) Custo dos Serviços
Prestados`` com score 1.0, e a consulta ao cache é o passo 1 de ``match()``.

A correção
----------

``utils/natureza`` lê RECEITA/DESPESA da **árvore**, dos dois lados:

- no balancete, "Servicos prestados - mercado interno" nada declara, mas pende
  de ``4 RECEITAS``;
- no plano referencial, "Serviços Prestados por Terceiros" nada declara, mas
  pende de ``3.90.02 Despesas Administrativas e Gerais``.

O mesmo balancete traz as duas contas, com nomes quase idênticos e sinais
opostos. É esse par que o teste tem de separar.

Referência: ``REVISAO_QUALIDADE.md`` §16.
"""

from __future__ import annotations

import pytest
from conftest import CORPUS_DIR

from src.bp.utils.natureza import (
    DESPESA,
    RECEITA,
    mapear_natureza,
    natureza_de_texto,
    totais_por_natureza,
)

# ============================================================================
# 1. Leitura da natureza pelo texto
# ============================================================================


@pytest.mark.parametrize(
    ("descricao", "esperado"),
    [
        ("(-) Custo dos Serviços Prestados", DESPESA),
        ("(-) Outras Receitas Operacionais", DESPESA),  # o "(-)" domina
        ("(+) Receitas financeiras", RECEITA),
        ("Receita da Prestação de Serviços no Mercado Interno", RECEITA),
        ("DESPESAS OPERACIONAIS", DESPESA),
        ("CUSTOS E DESPESAS", DESPESA),
        ("Custo das mercadorias vendidas", DESPESA),  # "custo" vence "venda"
        ("RECEITAS", RECEITA),
        ("Servicos prestados - mercado interno", None),  # não declara nada
        ("ALUGUEIS", None),
        ("", None),
        (None, None),
    ],
)
def test_natureza_de_texto(descricao, esperado):
    assert natureza_de_texto(descricao) == esperado


def test_nao_declarar_e_resposta_legitima():
    """
    ``None`` não é falha: é o caso comum, e chutar seria pior.

    Um chute errado inverte o sinal de uma conta na DRE. Quem responde por uma
    folha muda é a árvore, não o nome dela.
    """
    assert natureza_de_texto("SERVICOS PRESTADOS") is None
    assert natureza_de_texto("Banco do Brasil") is None


# ============================================================================
# 2. Leitura pela árvore — o par homônimo
# ============================================================================


def test_o_ramo_decide_o_que_a_folha_nao_diz():
    contas = [
        {"codigo": "3", "descricao": "CUSTOS E DESPESAS"},
        {"codigo": "3.2", "descricao": "DESPESAS OPERACIONAIS"},
        {"codigo": "3.2.2.01.004", "descricao": "Servicos prestados por terceiros"},
        {"codigo": "4", "descricao": "RECEITAS"},
        {"codigo": "4.1", "descricao": "RECEITAS OPERACIONAIS"},
        {"codigo": "4.1.1.01.002", "descricao": "Servicos prestados - mercado interno"},
    ]
    natureza = mapear_natureza(contas)
    assert natureza["3.2.2.01.004"] == DESPESA
    assert natureza["4.1.1.01.002"] == RECEITA, (
        "a receita foi classificada como despesa — é o defeito do §16 voltando"
    )


def test_ancestral_mais_proximo_vence_o_mais_alto():
    """
    A regra do "ancestral mais alto" quebra feio, e por isso não é usada.

    A raiz ``3`` do Plano Referencial da RFB tem descrição "Redução do IPI na
    **Venda** de Bens de Informática..." — declararia RECEITA e classificaria
    as 451 contas de resultado do plano como receita, custos inclusive.
    """
    contas = [
        {"codigo": "3", "descricao": "Redução do IPI na Venda de Bens de Informática"},
        {"codigo": "3.90.02", "descricao": "Despesas Administrativas e Gerais"},
        {"codigo": "3.90.02.04", "descricao": "Serviços Prestados por Terceiros"},
    ]
    assert mapear_natureza(contas)["3.90.02.04"] == DESPESA


def test_contas_fora_do_resultado_nao_entram_no_mapa():
    """Para Ativo e Passivo a natureza não existe; devolvê-la convidaria uso errado."""
    contas = [
        {"codigo": "1.1.1", "descricao": "DISPONIVEL"},
        {"codigo": "2.1.1", "descricao": "FORNECEDORES"},
        {"codigo": "3.1", "descricao": "DESPESAS"},
    ]
    natureza = mapear_natureza(contas)
    assert set(natureza) == {"3.1"}


def test_totais_somam_o_totalizador_nao_a_arvore_toda():
    """Somar todas as contas contaria pai e filhos — o ramo entra duas vezes."""
    contas = [
        {"codigo": "4", "descricao": "RECEITAS", "saldo": 1000.0},
        {"codigo": "4.1", "descricao": "RECEITAS OPERACIONAIS", "saldo": 1000.0},
        {"codigo": "4.1.1", "descricao": "Servicos", "saldo": 1000.0},
        {"codigo": "3", "descricao": "DESPESAS", "saldo": 400.0},
    ]
    totais = totais_por_natureza(contas, mapear_natureza(contas))
    assert totais[RECEITA] == pytest.approx(1000.0)
    assert totais[DESPESA] == pytest.approx(400.0)


# ============================================================================
# 3. O matcher passa a separar o par
# ============================================================================


@pytest.fixture(scope="module")
def matcher(tmp_path_factory):
    from src.bp.output.build_gt_output import _build_matcher

    return _build_matcher(None, tmp_path_factory.mktemp("cache") / "c.json")


def test_receita_de_servicos_vai_para_linha_de_receita(matcher):
    """
    O caso exato reportado. O cache do projeto tem
    ``servicos prestados -> (-) Custo dos Serviços Prestados`` com score 1.0;
    a natureza tem de vencer o cache.
    """
    resultado = matcher.match(
        "Servicos prestados - mercado interno",
        codigo_origem="4.1.1.01.002",
        natureza_resultado=RECEITA,
    )
    assert resultado.decision is not None
    destino = matcher.natureza_referencial.get(resultado.decision.codigo)
    assert destino == RECEITA, (
        f"receita foi para {resultado.decision.codigo} "
        f"({resultado.decision.descricao}), que é {destino}"
    )


def test_despesa_de_servicos_continua_indo_para_despesa(matcher):
    """A trava não pode ser um bloqueio geral: o par oposto tem de continuar certo."""
    resultado = matcher.match(
        "Servicos prestados por terceiros",
        codigo_origem="3.2.2.01.004",
        natureza_resultado=DESPESA,
    )
    assert resultado.decision is not None
    assert matcher.natureza_referencial.get(resultado.decision.codigo) == DESPESA


def test_natureza_entra_na_chave_do_cache():
    """
    Sem isso o cache continua atropelando a restrição: a consulta a ele é o
    passo 1 de ``match()``, antes de qualquer filtro.
    """
    from src.bp.matchers.conta_matcher import ContaMatcher

    receita = ContaMatcher._chave_cache("servicos prestados", "RESULTADO", RECEITA)
    despesa = ContaMatcher._chave_cache("servicos prestados", "RESULTADO", DESPESA)
    assert receita != despesa
    # Sem classe, a chave antiga é preservada (decisões já gravadas continuam válidas).
    assert ContaMatcher._chave_cache("caixa", None) == "caixa"


def test_referencial_e_classificado_pela_propria_arvore(matcher):
    """
    Não dá para ler só da descrição: "Serviços Prestados por Terceiros" nada
    declara, mas pende de "3.90.02 Despesas Administrativas e Gerais". Era por
    essa fresta que a receita escapava.
    """
    assert matcher.natureza_referencial.get("3.90.02.04") == DESPESA
    assert matcher.natureza_referencial.get("3.01.01.03.01.03") == DESPESA
    assert matcher.natureza_referencial.get("3.01.01.01.01.06") == RECEITA
    # Não-vacuidade: as duas naturezas têm de estar representadas de verdade.
    valores = set(matcher.natureza_referencial.values())
    assert valores == {RECEITA, DESPESA}
    assert len(matcher.natureza_referencial) > 400


# ============================================================================
# 4. A trava final, sobre o destino
# ============================================================================


def test_trava_recusa_receita_indo_para_conta_de_despesa(matcher):
    """
    Não-vacuidade da trava: ela tem de *disparar*.

    Sem isto, o teste acima também passaria com a trava desligada — que é o
    estado em que o defeito chegou ao cliente.
    """
    from src.bp.output.build_gt_output import _resolver
    from src.bp.output.template_map import ProjectionResult

    class _Decisao:
        codigo = "3.01.01.03.01.03"  # (-) Custo dos Serviços Prestados
        descricao = "(-) Custo dos Serviços Prestados"
        score = 1.0

    class _MatcherFalso:
        natureza_referencial = matcher.natureza_referencial

        def match(self, descricao, codigo_origem="", natureza_resultado=None, prazo=None):
            return type("R", (), {"decision": _Decisao(), "needs_review": False})()

    class _ProjectorFalso:
        def project(self, codigo):
            return ProjectionResult(codigo, "3.01.01.03", "direto")

    resolucao = _resolver(
        {"codigo": "4.1.1.01.002", "descricao": "Servicos prestados - mercado interno"},
        _MatcherFalso(),
        _ProjectorFalso(),
        {"4.1.1.01.002": RECEITA},
    )
    assert resolucao.codigo_template is None, "receita aceita em conta de custo"
    assert "natureza incompatível" in resolucao.motivo
    assert RECEITA in resolucao.motivo and DESPESA in resolucao.motivo


# ============================================================================
# 5. O efeito na entrega, no balancete que expôs o defeito
# ============================================================================


@pytest.mark.integration
def test_a_receita_chega_a_entrega_como_receita(tmp_path):
    """
    Ponta a ponta: a linha escrita tem de cair numa linha de receita da DRE,
    com valor positivo.
    """
    from openpyxl import load_workbook

    from src.bp.output.build_gt_output import build_gt_output
    from src.bp.output.template_map import TemplateProjector

    caminho = CORPUS_DIR / "Balancete_Trindade_052025.xlsx"
    if not caminho.exists():
        pytest.skip(f"corpus ausente: {caminho}")

    resultado = build_gt_output(
        caminho, tmp_path / "s.xlsx", ano_base=2025, cache_path=tmp_path / "c.json"
    )
    ws = load_workbook(resultado.output_path)["_dados_padronizados"]
    linhas = [
        (str(ws.cell(row=r, column=1).value or ""), ws.cell(row=r, column=2).value,
         ws.cell(row=r, column=3).value)
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=1).value
    ]
    receita = [x for x in linhas if "ervicos prestados" in str(x[1] or "")]
    assert receita, "a conta de receita sumiu da entrega"

    projector = TemplateProjector()
    for codigo, descricao, valor in receita:
        capturas = projector.linhas_que_capturam(codigo)
        assert capturas, f"{codigo} não é somado por nenhuma linha do template"
        rotulo = capturas[0].rotulo
        assert not rotulo.lstrip().startswith("(-)"), (
            f"receita '{descricao}' foi parar em '{rotulo}' — linha de dedução"
        )
        assert (valor or 0) > 0, f"receita entrou com valor {valor}"
