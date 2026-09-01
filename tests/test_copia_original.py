"""
A entrega carrega dentro dela o balancete que a gerou.

Por que este arquivo existe
---------------------------

Sem a cópia, responder *"de onde saiu este número?"* depende de reencontrar,
meses depois, o balancete numa pasta de rede. Com ela, a resposta está no
próprio arquivo entregue — e o SHA-256 prova **qual** arquivo foi usado, o
que "parece o mesmo balancete" nunca prova.

O que estes testes travam
-------------------------

1. A aba existe e não está vazia (a armadilha clássica: aba criada, zero
   linhas, teste verde).
2. O conteúdo é o do arquivo — conferido contra linhas lidas direto do
   original, não contra uma constante escrita à mão aqui.
3. O SHA-256 na planilha é o do arquivo em disco.
4. Quando o pipeline lê **outro** arquivo no lugar do pedido (o ``.xlsx``
   irmão de um ``.xls``), a aba diz isso. Rastreio que aponta para o arquivo
   errado é pior que rastreio nenhum.
"""

from __future__ import annotations

import hashlib

import pytest
from conftest import require_corpus_file
from openpyxl import load_workbook

from src.bp.output.build_gt_output import build_gt_output
from src.bp.output.origem import MAX_COLUNAS, MAX_LINHAS, ler_origem, nome_da_aba


def _sha_em_disco(caminho) -> str:
    return hashlib.sha256(caminho.read_bytes()).hexdigest()


def _celulas(ws) -> list[tuple]:
    return list(ws.iter_rows(values_only=True))


# --------------------------------------------------------------------------
# Unidade: leitura da origem
# --------------------------------------------------------------------------


def test_sha256_confere_com_o_arquivo_em_disco(balancete_xlsx):
    origem = ler_origem(balancete_xlsx)
    assert origem.sha256 == _sha_em_disco(balancete_xlsx)
    assert origem.tamanho_bytes == balancete_xlsx.stat().st_size


def test_transcricao_preserva_o_preambulo_que_o_parser_descarta(balancete_csv):
    """
    O parser detecta cabeçalho e joga fora o preâmbulo; a cópia não pode.

    É justamente no preâmbulo que estão empresa, período e data de emissão —
    o que identifica o balancete.
    """
    origem = ler_origem(balancete_csv)
    linhas_arquivo = balancete_csv.read_text(encoding="latin-1").split("\n")
    assert len(origem.linhas) == len(linhas_arquivo)
    assert origem.linhas[0][0] == linhas_arquivo[0].rstrip("\r")


def test_transcricao_de_planilha_bate_linha_a_linha_com_o_original(balancete_xlsx):
    """Confere contra o arquivo lido aqui — não contra número escrito à mão."""
    import pandas as pd

    origem = ler_origem(balancete_xlsx)
    df = pd.read_excel(balancete_xlsx, header=None, dtype=object)
    assert len(origem.linhas) == len(df)
    primeira_do_arquivo = [
        None if v != v else v for v in df.iloc[0].tolist()  # NaN -> None
    ]
    assert origem.linhas[0][: len(primeira_do_arquivo)] == primeira_do_arquivo


def test_arquivo_ilegivel_vira_erro_declarado_e_nao_excecao(tmp_path):
    """A entrega não pode quebrar porque a aba de rastreio falhou — mas
    também não pode mentir que está tudo bem."""
    quebrado = tmp_path / "corrompido.xlsx"
    quebrado.write_bytes(b"isto nao e uma planilha")
    origem = ler_origem(quebrado)
    assert origem.erro, "arquivo ilegível precisa declarar o motivo"
    assert not origem.legivel
    assert origem.sha256 == _sha_em_disco(quebrado), (
        "mesmo sem transcrever, o hash identifica o arquivo"
    )


def test_nome_de_aba_respeita_o_limite_do_excel_e_nao_colide():
    usados: set[str] = set()
    nomes = [nome_da_aba(ano, usados) for ano in (2021, 2022, 2023)]
    assert len(set(nomes)) == 3
    assert all(len(n) <= 31 for n in nomes)
    assert all(not set(n) & set("[]:*?/\\") for n in nomes)
    # Repetir o mesmo ano não pode sobrescrever a aba anterior.
    assert nome_da_aba(2021, usados) not in nomes


def test_truncamento_e_declarado_e_nunca_silencioso(monkeypatch, tmp_path):
    import src.bp.output.origem as mod

    monkeypatch.setattr(mod, "MAX_LINHAS", 3)
    arquivo = tmp_path / "grande.csv"
    arquivo.write_text("\n".join(f"linha {i}" for i in range(20)))
    origem = mod.ler_origem(arquivo)
    assert origem.truncado
    assert len(origem.linhas) == 3


def test_tetos_sao_compativeis_com_o_excel():
    """Guarda contra alguém elevar o teto acima do que o formato aceita."""
    assert MAX_LINHAS <= 1_000_000
    assert MAX_COLUNAS <= 16_384


# --------------------------------------------------------------------------
# Integração: a aba dentro da entrega
# --------------------------------------------------------------------------


@pytest.fixture(scope="module")
def entrega(tmp_path_factory, request):
    origem = require_corpus_file("Balancete Real Life.xlsx")
    saida = tmp_path_factory.mktemp("entrega") / "saida.xlsx"
    resultado = build_gt_output(origem, saida, ano_base=2024)
    return resultado, load_workbook(saida), origem


def test_a_entrega_tem_a_aba_de_copia(entrega):
    _, wb, _ = entrega
    assert "Balancete Original" in wb.sheetnames


def test_a_aba_de_copia_nao_esta_vazia(entrega):
    """
    Guarda de não-vacuidade.

    Uma aba criada com só o cabeçalho passaria em qualquer teste de
    existência e não rastrearia nada. O corpo tem que ter mais linhas que o
    cabeçalho de proveniência.
    """
    _, wb, origem_path = entrega
    linhas = _celulas(wb["Balancete Original"])
    assert len(linhas) > 20, "aba de rastreio praticamente vazia"

    texto = "\n".join(
        str(c) for linha in linhas for c in linha if c is not None
    )
    assert "CONTEÚDO DO ARQUIVO" in texto
    assert origem_path.name in texto


def test_o_hash_na_planilha_identifica_o_arquivo_de_origem(entrega):
    _, wb, origem_path = entrega
    linhas = _celulas(wb["Balancete Original"])
    rotulos = {str(linha[0]): linha[1] for linha in linhas if linha and linha[0]}
    assert rotulos["SHA-256:"] == _sha_em_disco(origem_path)


def test_o_sumario_indexa_as_origens(entrega):
    resultado, wb, origem_path = entrega
    texto = "\n".join(
        str(c)
        for linha in _celulas(wb["Sumário"])
        for c in linha
        if c is not None
    )
    assert "ORIGEM DOS DADOS" in texto
    assert _sha_em_disco(origem_path) in texto
    assert resultado.origens and resultado.origens[0].legivel


def test_a_copia_reproduz_valores_do_balancete(entrega):
    """
    Amostra real: um valor numérico do original tem que aparecer na aba.

    Sem isto, a aba poderia conter só as descrições e o teste de "não está
    vazia" continuaria verde.
    """
    resultado, wb, _ = entrega
    origem = resultado.origens[0]
    numeros_origem = {
        round(float(c), 2)
        for linha in origem.linhas
        for c in linha
        if isinstance(c, (int, float)) and not isinstance(c, bool)
    }
    assert numeros_origem, "a transcrição não trouxe número nenhum"

    numeros_aba = {
        round(float(c), 2)
        for linha in _celulas(wb["Balancete Original"])
        for c in linha
        if isinstance(c, (int, float)) and not isinstance(c, bool)
    }
    faltando = numeros_origem - numeros_aba
    assert not faltando, f"valores transcritos não chegaram à aba: {sorted(faltando)[:5]}"


def test_a_copia_fica_antes_das_abas_tecnicas(entrega):
    """Aba de consulta ocasional não pode ficar depois de uma aba oculta."""
    _, wb, _ = entrega
    nomes = wb.sheetnames
    assert nomes.index("Balancete Original") < nomes.index("_dados_padronizados")
    assert nomes.index("Contas Não Identificadas") < nomes.index("Balancete Original")


@pytest.mark.integration
def test_copia_aponta_o_arquivo_realmente_lido_quando_ha_irmao_xlsx():
    """
    ``XlsParser.read()`` prefere o ``.xlsx`` irmão de mesmo nome, calado.

    Se a cópia transcrevesse o ``.xls`` pedido enquanto o pipeline leu o
    ``.xlsx``, ela mostraria um arquivo que não gerou número nenhum. A aba
    tem que declarar de onde o conteúdo veio.
    """
    xls = require_corpus_file("Balancete SPEZZIA TUBOS 01012024-31122024.xls")
    if not xls.with_suffix(".xlsx").exists():
        pytest.skip("este balancete não tem .xlsx irmão neste workspace")

    origem = ler_origem(xls)
    assert origem.lido_de == xls.with_suffix(".xlsx")
    assert "ATENÇÃO" in origem.procedencia
    assert origem.sha256 == _sha_em_disco(xls), (
        "o hash identifica o arquivo PEDIDO; a procedência explica a troca"
    )
