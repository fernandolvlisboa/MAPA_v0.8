"""
O teste do core: **o Ativo da entrega é o Ativo do balancete?**

A pergunta que faltava
----------------------

    "Qual é o ativo desses balancetes? Se o ativo desses balancetes soma dois
    milhões de reais, o ativo do resultado que você vai encontrar deve somar
    dois milhões de reais. O ativo é uma linha que quase sempre aparece no
    balancete já como um totalizador, isso é algo fácil de ser achado."

Havia 446 testes e nenhum respondia isso. Todos mediam **proxy**: quantas
contas casaram, se a árvore da origem fecha, se cada linha escrita é capturada
por uma linha do template. Proxy fica verde com o número final errado — e
ficou: um balancete real entregou Ativo de 2.683.506,57 contra 2.361.053,53
declarados na origem, R$ 322.453,04 a mais, com a suíte inteira verde.

O ponto cego era estrutural: **escrever em ``_dados_padronizados`` não põe
número na entrega**. Quem soma são as fórmulas do template, e nada em Python
as executava. Entre o último teste e o número que o cliente lê havia uma camada
inteira sem cobertura. ``validators/entrega.py`` executa essa camada.

A identidade
------------

Não basta ``entrega == origem``: uma conta pode legitimamente não ter destino
no template (acontece num percentual alto dos clientes — ver §10) e o valor
dela fica de fora **por decisão consciente e reportada**. A identidade certa é::

    entrega + não coberto == origem

O resíduo dessa conta é o que não tem explicação — e é sempre defeito.

O discriminador
---------------

Um balancete cuja **origem** já não fecha (a soma dos filhos não bate com o
pai) não pode produzir entrega consistente: lixo entra, lixo sai. Nesses casos
o que se exige é que o sistema **diga** que a origem está quebrada, não que os
totais batam.

O corte é por *propriedade do dado* (``rollup_integro``), nunca por nome de
arquivo — a regra do §10. Medido no corpus: dos 19 balancetes com totalizador,
15 têm origem íntegra e nos 15 a identidade bate com resíduo 0,00; os 4
restantes têm origem inconsistente por defeito conhecido de parser
(``test_dispatcher_roteamento.py::test_txt_preserva_o_sinal_das_contas_redutoras``).

Referência: ``REVISAO_QUALIDADE.md`` §15.
"""

from __future__ import annotations

import os
import random
import tempfile
from pathlib import Path

import pytest
from conftest import CORPUS_DIR
from openpyxl import load_workbook

from src.bp.output.build_gt_output import (
    _orientacao_por_classe,
    _valor_para_o_template,
    build_gt_output,
)
from src.bp.output.template_map import TemplateProjector
from src.bp.validators.entrega import (
    TOLERANCIA,
    Conferencia,
    avaliar_demonstrativo,
    conferir_totais,
)

TEMPLATE = Path("templates/Template_GT_BP_Padrao_v3.xlsx")
EXTENSOES = {".xls", ".xlsx", ".csv", ".txt"}

#: Controle escolhido por forma (emissores, profundidades e convenções
#: diferentes), não por conveniência.
CONTROLE = (
    "Balancete 072022 122022 - RBM.xls",
    "Balancete SPEZZIA TUBOS 01012024-31122024.xls",
    "202404_2024 - Balancete.xls",
    "Balancete_Trindade_052025.xlsx",
    "VIVAE - Balancete Acumulado 12.2023 - Emitido em 03.05.2024.xls",
)
TAMANHO_AMOSTRA = 3

#: Mínimo de arquivos do controle que precisam chegar à conferência de verdade.
#: Sem este piso, o teste ficaria verde com tudo caindo em `skip`.
PISO_CONFERIVEIS = 3


def _residuos(resultado, escala: float = 1000.0) -> dict[str, float]:
    """``entrega - origem + |não coberto|`` por classe. Zero é o esperado."""
    saida: dict[str, float] = {}
    for conferencia in resultado.entrega.conferencias:
        classe = "ATIVO" if conferencia.nome.startswith("ATIVO") else "PASSIVO"
        nao_coberto = resultado.nao_coberto_por_classe.get(classe, 0.0) / escala
        saida[classe] = conferencia.diferenca + abs(nao_coberto)
    return saida


def _gerar(caminho: Path, ano: int = 2024):
    if not TEMPLATE.exists():
        pytest.skip(f"Template GT ausente: {TEMPLATE}")
    tmp = Path(tempfile.mkdtemp())
    return build_gt_output(
        caminho, tmp / "saida.xlsx", ano_base=ano, cache_path=tmp / "cache.json"
    )


def _conferir(caminho: Path) -> str:
    """
    Roda a conferência e devolve o veredito, para o chamador contabilizar.

    - ``"conferido"`` — origem íntegra e a identidade bate;
    - ``"origem-inconsistente"`` — a origem não fecha; exige-se o aviso;
    - ``"sem-totalizador"`` — não há o que conferir.
    """
    resultado = _gerar(caminho)

    if not resultado.entrega.conferivel:
        assert resultado.entrega.motivo_nao_conferido, (
            f"{caminho.name}: conferência indisponível sem dizer por quê — "
            "silêncio aqui vira 'passou' na leitura de quem revisa"
        )
        return "sem-totalizador"

    integra = bool(resultado.hierarquia and resultado.hierarquia.rollup_integro)
    if not integra:
        assert any("BALANCETE DE ORIGEM não fecha" in a for a in resultado.avisos), (
            f"{caminho.name}: a origem não fecha e a entrega não avisa. "
            "Entregar sem dizer que o insumo está quebrado é o pior caso."
        )
        return "origem-inconsistente"

    residuos = _residuos(resultado)
    assert residuos, f"{caminho.name}: nenhuma classe conferida"
    piores = {c: r for c, r in residuos.items() if abs(r) > TOLERANCIA}
    assert not piores, (
        f"{caminho.name}: o total da entrega não é o total do balancete.\n"
        + "\n".join(f"  {c}" for c in resultado.entrega.conferencias)
        + f"\n  resíduo por classe (entrega - origem + |não coberto|): {piores}\n"
        "  Resíduo diferente de zero é valor inventado ou perdido — não é "
        "conta sem destino, essas já entram no 'não coberto'."
    )
    return "conferido"


# ============================================================================
# 1. O teste geral — controle fixo
# ============================================================================


@pytest.mark.integration
def test_o_ativo_da_entrega_e_o_ativo_do_balancete():
    """
    O core, sobre o controle inteiro de uma vez.

    Roda em bloco (e não parametrizado) para poder exigir o **piso**: pelo
    menos ``PISO_CONFERIVEIS`` arquivos têm de chegar à conferência de verdade.
    Sem o piso, o teste ficaria verde num workspace onde todo arquivo caísse em
    "sem totalizador" — verde que não valida nada, que é o que esta suíte
    existe para impedir.
    """
    if not CORPUS_DIR.exists():
        pytest.skip(f"corpus ausente: {CORPUS_DIR}")

    ausentes = [n for n in CONTROLE if not (CORPUS_DIR / n).exists()]
    assert not ausentes, (
        f"controle ausente do corpus: {ausentes}. Nome errado aqui faz o "
        "arquivo sumir do controle em silêncio — e cair na amostra, onde a "
        "cobertura é aleatória. Foi o que aconteceu com o VIVAE."
    )

    veredito: dict[str, list[str]] = {}
    for nome in CONTROLE:
        veredito.setdefault(_conferir(CORPUS_DIR / nome), []).append(nome)

    conferidos = veredito.get("conferido", [])
    print(f"\n[totais] conferidos: {len(conferidos)} -> {conferidos}")
    for chave in ("origem-inconsistente", "sem-totalizador"):
        if veredito.get(chave):
            print(f"[totais] {chave}: {veredito[chave]}")

    assert len(conferidos) >= PISO_CONFERIVEIS, (
        f"só {len(conferidos)} balancete(s) do controle chegaram à conferência "
        f"de totais (piso {PISO_CONFERIVEIS}). O teste estaria passando por "
        f"vacuidade. Vereditos: { {k: len(v) for k, v in veredito.items()} }"
    )


# ============================================================================
# 2. O teste geral — amostra aleatória
# ============================================================================


@pytest.fixture(scope="module")
def amostra() -> list[Path]:
    """Sorteia fora do controle; a semente vai no log (``BP_SEED`` repete)."""
    if not CORPUS_DIR.exists():
        pytest.skip(f"corpus ausente: {CORPUS_DIR}")
    candidatos = [
        p
        for p in sorted(CORPUS_DIR.iterdir())
        if p.suffix.lower() in EXTENSOES and p.name not in CONTROLE
    ]
    if not candidatos:
        pytest.skip("corpus sem arquivos além do controle")
    bruta = os.environ.get("BP_SEED")
    semente = int(bruta) if bruta and bruta.isdigit() else random.randrange(1_000_000)
    print(f"\n[totais] semente da amostra = {semente}  (BP_SEED={semente} repete)")
    return random.Random(semente).sample(
        candidatos, min(TAMANHO_AMOSTRA, len(candidatos))
    )


@pytest.mark.integration
@pytest.mark.parametrize("indice", range(TAMANHO_AMOSTRA))
def test_totais_da_entrega_num_balancete_sorteado(amostra, indice):
    """
    O que ninguém testa é onde o defeito mora.

    Calibrar contra um arquivo produz correção que funciona nele e quebra o
    modelo nos outros (§10). O sorteio é o que impede isso.
    """
    if indice >= len(amostra):
        pytest.skip("amostra menor que o índice")
    caminho = amostra[indice]
    print(f"[totais] {caminho.name}: {_conferir(caminho)}")


# ============================================================================
# 3. O defeito exato, travado — o abs() que inflava o Ativo
# ============================================================================


def test_conta_redutora_do_balanco_mantem_o_sinal_da_origem():
    """
    Não-vacuidade da correção.

    Era ``abs(saldo) * sign_for(linha)``. Como nenhuma linha do BP_GT é "(-)",
    toda conta redutora do Balanço entrava **positiva**: uma depreciação
    acumulada de -155.617,00 virava +155.617,00 e inflava o Ativo em duas
    vezes o valor dela.
    """
    projector = TemplateProjector()
    valor = _valor_para_o_template(
        bruto=-155_617.00,
        escala=1000.0,
        codigo_origem="1.2.3.02.002",
        codigo_template="1.02.03",
        projector=projector,
        orientacao={"ATIVO": 1},
    )
    assert valor == pytest.approx(-155.617), (
        f"depreciação acumulada entrou como {valor}; o sinal da origem foi perdido"
    )


def test_conta_da_dre_usa_o_sinal_da_linha_do_template():
    """
    A DRE segue outra regra, e tem de continuar seguindo.

    Sob natureza implícita a origem traz receita e despesa **ambas positivas**;
    quem decide o sinal é o rótulo da linha ("(-) Despesas com pessoal"),
    porque as fórmulas da DRE somam.
    """
    projector = TemplateProjector()
    despesa = _valor_para_o_template(
        bruto=50_000.0,
        escala=1000.0,
        codigo_origem="3.1.1.01",
        codigo_template="3.01.01.07.01.01",
        projector=projector,
        orientacao={},
    )
    assert despesa < 0, "despesa tem de entrar negativa na DRE"
    assert abs(despesa) == pytest.approx(50.0)


def test_orientacao_por_classe_poe_os_dois_lados_positivos():
    """
    Balancete que traz o Passivo negativo (para as três classes somarem zero)
    tem de sair com os dois lados positivos — o check do template subtrai.
    """
    class _Hierarquia:
        tem_hierarquia = True
        totais_por_classe = {"ATIVO": 2_000_000.0, "PASSIVO": -2_000_000.0}

    orientacao = _orientacao_por_classe(_Hierarquia(), {})
    assert orientacao["ATIVO"] == 1
    assert orientacao["PASSIVO"] == -1
    # Passivo negativo é a assinatura da origem COM SINAL, e ela vale para a
    # DRE também: um crédito dentro do ramo de despesas precisa manter o sinal.
    assert orientacao["RESULTADO"] == -1

    projector = TemplateProjector()
    # Dentro do Passivo orientado, o sinal RELATIVO da conta é preservado.
    normal = _valor_para_o_template(
        -800_000.0, 1000.0, "2.1.1", "2.01.01.03", projector, orientacao
    )
    redutora = _valor_para_o_template(
        +30_000.0, 1000.0, "2.1.9", "2.01.01.15", projector, orientacao
    )
    assert normal == pytest.approx(800.0)
    assert redutora == pytest.approx(-30.0)


def test_orientacao_mantem_positivo_o_balancete_de_natureza_implicita():
    class _Hierarquia:
        tem_hierarquia = True
        totais_por_classe = {"ATIVO": 2_361_053.53, "PASSIVO": 891_480.90}

    orientacao = _orientacao_por_classe(_Hierarquia(), {})
    assert orientacao == {"ATIVO": 1, "PASSIVO": 1}, (
        "sob natureza implícita não há orientação de RESULTADO: quem decide o "
        "sinal da DRE é o rótulo da linha do template"
    )


def test_sem_arvore_o_sinal_da_origem_e_preservado_intacto():
    """
    Sem totalizador declarado não há evidência de convenção — e inferir errado
    inverte o arquivo inteiro.

    Foi o que um recorte de balancete de cinco contas provocou: a soma da
    classe deu negativa por acaso e a orientação virou todo o Ativo. Sem
    árvore, o sinal da origem passa intacto e, se a convenção for outra, quem
    denuncia é a conferência de totais — muito melhor que inverter calado.
    """
    class _SemArvore:
        tem_hierarquia = False
        totais_por_classe = {"ATIVO": -530.88}

    assert _orientacao_por_classe(_SemArvore(), {}) == {}
    assert _orientacao_por_classe(None, {}) == {}


# ============================================================================
# 4. O avaliador de fórmulas — ele é a régua, então tem de ser conferido
# ============================================================================


@pytest.mark.integration
def test_avaliador_reproduz_o_curinga_do_sumifs():
    """
    A régua precisa medir o que o Excel mede.

    Se o avaliador ignorasse o ``*``, ele nunca veria a dupla contagem que o
    curinga permite — e a conferência de totais seria decorativa.
    """
    caminho = CORPUS_DIR / "Balancete_Trindade_052025.xlsx"
    if not caminho.exists():
        pytest.skip(f"corpus ausente: {caminho}")
    resultado = _gerar(caminho, ano=2025)
    wb = load_workbook(resultado.output_path)

    linhas = avaliar_demonstrativo(wb, "BP_GT", "D")
    assert linhas, "avaliador não produziu linha nenhuma"

    # "Caixa e equivalentes de caixa" (1.01.01) tem de ser a soma EXATA das
    # linhas de _dados_padronizados que começam com 1.01.01 — nem mais (dupla
    # contagem) nem menos (valor perdido).
    dados = wb["_dados_padronizados"]
    esperado = 0.0
    for linha in range(2, dados.max_row + 1):
        codigo = dados.cell(row=linha, column=1).value
        valor = dados.cell(row=linha, column=3).value
        if codigo and str(codigo).strip().startswith("1.01.01"):
            esperado += float(valor or 0.0)
    assert linhas["Caixa e equivalentes de caixa"] == pytest.approx(esperado)

    # E os totais têm de ser a soma das suas parcelas, não números soltos.
    assert linhas["ATIVO TOTAL"] == pytest.approx(
        linhas["Total do Ativo Circulante"] + linhas["Total do Ativo Não Circulante"]
    )


def test_avaliador_recusa_formula_que_nao_modela(tmp_path):
    """
    Devolver zero calado para uma fórmula desconhecida fabricaria um
    "confere". O avaliador tem de falhar alto.
    """
    from openpyxl import Workbook

    wb = Workbook()
    dados = wb.active
    dados.title = "_dados_padronizados"
    dados["A2"], dados["C2"] = "1.01.01", 10.0
    bp = wb.create_sheet("BP_GT")
    bp["B9"] = "Linha esquisita"
    bp["D9"] = "=VLOOKUP(A1,B:C,2,FALSE)"

    with pytest.raises(ValueError, match="não modelada"):
        avaliar_demonstrativo(wb, "BP_GT", "D")


def test_conferencia_sem_totalizador_nao_finge_sucesso():
    relatorio = conferir_totais(load_workbook(TEMPLATE), {})
    assert not relatorio.conferivel
    assert not relatorio.confere, "sem dado para conferir, 'confere' tem de ser falso"
    assert relatorio.motivo_nao_conferido


def test_conferencia_declara_a_diferenca_em_numeros():
    conferencia = Conferencia(nome="ATIVO TOTAL", origem=2361.05, entrega=2683.51)
    assert not conferencia.confere
    assert conferencia.diferenca == pytest.approx(322.46)
    texto = str(conferencia)
    assert "2,361.05" in texto and "2,683.51" in texto and "NÃO" in texto


# ============================================================================
# 5. Balancete aberto: o resultado do período tem de chegar ao PL
# ============================================================================
#
# Balancete de verificação mensal quase sempre vem ABERTO: as contas de
# resultado ainda têm saldo e o lucro não foi transferido para o PL. Nesse
# estado, Ativo != Passivo + PL por construção — a diferença *é* o resultado.
# Medido no corpus: 6 dos 15 balancetes conferíveis são abertos.


@pytest.mark.integration
def test_balancete_aberto_fecha_com_o_resultado_no_pl(tmp_path):
    """
    O caso que o revisor mostrou: Ativo 2.361,05 contra Passivo + PL 891,48,
    diferença 1.469,57 — exatamente ``Receitas - Despesas``. Sem levar o
    resultado ao PL, a entrega mostra "Check: NOK" para um balancete correto.
    """
    caminho = CORPUS_DIR / "Balancete_Trindade_052025.xlsx"
    if not caminho.exists():
        pytest.skip(f"corpus ausente: {caminho}")

    resultado = _gerar(caminho, ano=2025)
    assert resultado.resultado_transferido != 0, (
        "nada foi transferido — o Balanço não vai fechar"
    )

    linhas = avaliar_demonstrativo(load_workbook(resultado.output_path), "BP_GT", "D")
    diferenca = linhas["ATIVO TOTAL"] - linhas["PASSIVO + PATRIMÔNIO LÍQUIDO"]
    assert abs(diferenca) <= TOLERANCIA, (
        f"o Balanço não fecha por {diferenca:,.2f} mesmo com o resultado no PL"
    )


def test_nao_se_pluga_diferenca_que_a_dre_nao_confirma():
    """
    Não-vacuidade, e é a parte que mais importa.

    ``Ativo - Passivo`` diferente de zero **não** prova balancete aberto: também
    aparece quando a extração perdeu uma conta. Plugar cegamente fabricaria um
    balanço fechado em cima de um erro — escondendo justamente o que precisa
    aparecer. A transferência só acontece quando a diferença do Balanço e o
    resultado da DRE, dois caminhos independentes, concordam.
    """
    from src.bp.output.build_gt_output import (
        BuildResult,
        _transferir_resultado_do_periodo,
    )

    class _Hierarquia:
        tem_hierarquia = True
        totais_por_classe = {"ATIVO": 2_000_000.0, "PASSIVO": 500_000.0}

    # A DRE diz 100.000; o Balanço pede 1.500.000. Não coincidem: é erro.
    contas = [
        {"codigo": "4", "descricao": "RECEITAS", "saldo": 300_000.0},
        {"codigo": "3", "descricao": "DESPESAS", "saldo": 200_000.0},
    ]
    resultado = BuildResult(output_path=None)  # type: ignore[arg-type]
    resultado.hierarquia = _Hierarquia()
    dados: list = []
    _transferir_resultado_do_periodo(
        contas, resultado, dados, [], 1000.0, {"ATIVO": 1, "PASSIVO": 1}
    )

    assert dados == [], "plugou uma diferença que a DRE não confirma"
    assert resultado.resultado_transferido == 0
    assert any("NÃO coincidem" in a for a in resultado.avisos)


def test_transfere_quando_os_dois_caminhos_concordam():
    from src.bp.output.build_gt_output import (
        BuildResult,
        _transferir_resultado_do_periodo,
    )

    class _Hierarquia:
        tem_hierarquia = True
        totais_por_classe = {"ATIVO": 2_000_000.0, "PASSIVO": 500_000.0}

    contas = [
        {"codigo": "4", "descricao": "RECEITAS", "saldo": 2_000_000.0},
        {"codigo": "3", "descricao": "DESPESAS", "saldo": 500_000.0},
    ]
    resultado = BuildResult(output_path=None)  # type: ignore[arg-type]
    resultado.hierarquia = _Hierarquia()
    dados: list = []
    _transferir_resultado_do_periodo(
        contas, resultado, dados, [], 1000.0, {"ATIVO": 1, "PASSIVO": 1}
    )

    assert len(dados) == 1
    assert dados[0]["codigo_padronizado"] == "2.03.04.01"
    assert dados[0]["valor"] == pytest.approx(1500.0)
    assert resultado.resultado_transferido == pytest.approx(1500.0)


def test_balancete_ja_encerrado_nao_recebe_lancamento():
    from src.bp.output.build_gt_output import (
        BuildResult,
        _transferir_resultado_do_periodo,
    )

    class _Hierarquia:
        tem_hierarquia = True
        totais_por_classe = {"ATIVO": 900_000.0, "PASSIVO": 900_000.0}

    resultado = BuildResult(output_path=None)  # type: ignore[arg-type]
    resultado.hierarquia = _Hierarquia()
    dados: list = []
    _transferir_resultado_do_periodo(
        [{"codigo": "4", "descricao": "RECEITAS", "saldo": 10.0}],
        resultado, dados, [], 1000.0, {"ATIVO": 1, "PASSIVO": 1},
    )
    assert dados == []
    assert resultado.resultado_transferido == 0


# ============================================================================
# 6. A DRE — o mesmo teste do core, do outro lado
# ============================================================================
#
# O Balanço pode fechar com a DRE inteira errada: uma receita que entra como
# custo não move o Ativo. Foi o que aconteceu — R$ 4,9 milhões de receita de
# serviços entraram como custo negativo e o ATIVO TOTAL continuou correto.


def _resultado_e_nao_coberto(resultado, escala: float = 1000.0):
    """A referência da origem e o que ficou sem destino, na moeda da entrega."""
    nao_coberto = sum(resultado.nao_coberto_por_natureza.values()) / escala
    origem_com_sinal = (
        resultado.hierarquia is not None
        and resultado.hierarquia.totais_por_classe.get("PASSIVO", 0.0) < 0
    )
    return resultado.resultado_da_origem, (
        -nao_coberto if origem_com_sinal else nao_coberto
    )


@pytest.mark.integration
def test_o_lucro_liquido_da_entrega_e_o_resultado_do_balancete():
    """
    A identidade, igual à do Balanço::

        entrega + não coberto == resultado da origem

    Medido no controle: os cinco batem ao centavo.
    """
    if not CORPUS_DIR.exists():
        pytest.skip(f"corpus ausente: {CORPUS_DIR}")

    ausentes = [n for n in CONTROLE if not (CORPUS_DIR / n).exists()]
    assert not ausentes, f"controle ausente do corpus: {ausentes}"

    conferidos: list[str] = []
    divergentes: list[str] = []
    for nome in CONTROLE:
        resultado = _gerar(CORPUS_DIR / nome)
        if not (resultado.hierarquia and resultado.hierarquia.rollup_integro):
            continue
        if not resultado.dre.conferivel:
            continue
        conferidos.append(nome)
        if not resultado.dre.confere:
            divergentes.append(f"{nome}: {resultado.dre.conferencias[0]}")

    assert len(conferidos) >= PISO_CONFERIVEIS, (
        f"só {len(conferidos)} balancete(s) chegaram à conferência da DRE — "
        "o teste estaria passando por vacuidade"
    )
    assert not divergentes, "a DRE não bate com a origem:\n" + "\n".join(divergentes)
    print(f"\n[dre] conferidos: {conferidos}")


def test_referencia_da_dre_muda_com_a_convencao_da_origem():
    """
    Não-vacuidade da régua: usar a fórmula errada produz um "não bate" que é da
    comparação, não do dado.

    - **com sinal** (receita credora negativa): ``-(total da classe RESULTADO)``
      — não passa pelo mapa de naturezas, e é isso que a torna robusta. Num
      balancete real, os ramos "3 RESULTADO LÍQUIDO ANTES DO IRPJ" e
      "4 IMPOSTOS SOBRE O LUCRO" não declaram natureza nenhuma; uma referência
      baseada no mapa deixava 86,73 mil de IRPJ/CSLL de fora.
    - **natureza implícita** (ambas positivas): ``|receitas| - |despesas|``,
      porque aí o total da classe soma as duas com o mesmo sinal.
    """
    from src.bp.utils.natureza import resultado_do_periodo

    com_sinal = [
        {"codigo": "3", "descricao": "RECEITAS", "saldo": -1000.0},
        {"codigo": "4", "descricao": "DESPESAS", "saldo": 400.0},
    ]
    assert resultado_do_periodo(com_sinal, -600.0, True) == pytest.approx(600.0)

    implicita = [
        {"codigo": "4", "descricao": "RECEITAS", "saldo": 1000.0},
        {"codigo": "3", "descricao": "CUSTOS E DESPESAS", "saldo": 400.0},
    ]
    assert resultado_do_periodo(implicita, 1400.0, False) == pytest.approx(600.0)


def test_um_discriminador_so_para_a_convencao_do_arquivo():
    """
    Não-vacuidade do que quase deu errado: ter DOIS sinais para a mesma
    pergunta.

    A primeira versão deduzia a convenção da DRE pelo sinal das naturezas.
    Num balancete com LUCRO, o ramo de resultado classificado DESPESA soma
    negativo (o lucro está lá dentro), os dois totais ficam com o mesmo sinal,
    a dedução conclui "natureza implícita" e a referência sai invertida — foi
    acusado 17,27 milhões de erro numa entrega correta.

    O Passivo não compensa receita com despesa, e por isso responde certo
    mesmo em ano de lucro.
    """
    from src.bp.output.build_gt_output import _origem_com_sinal
    from src.bp.utils.natureza import resultado_do_periodo

    class _ComLucro:
        # Assinatura real de um balancete de cliente que teve lucro.
        totais_por_classe = {
            "ATIVO": 282_701_208.15,
            "PASSIVO": -274_027_823.96,
            "RESULTADO": -8_673_384.19,
        }

    assert _origem_com_sinal(_ComLucro())
    # Os dois ramos somam NEGATIVO — o sinal das naturezas diria "implícita".
    contas = [
        {"codigo": "3", "descricao": "RECEITAS", "saldo": -37_615.17},
        {"codigo": "4", "descricao": "DESPESAS", "saldo": -8_635_769.02},
    ]
    assert resultado_do_periodo(contas, -8_673_384.19, True) == pytest.approx(
        8_673_384.19
    ), "o lucro saiu com o sinal trocado"

    class _NaturezaImplicita:
        totais_por_classe = {"ATIVO": 2_361_053.53, "PASSIVO": 891_480.90}

    assert not _origem_com_sinal(_NaturezaImplicita())


def test_credito_dentro_do_ramo_de_despesa_mantem_o_sinal():
    """
    O defeito que o VIVAE escondia, e o último a cair.

    "CREDITO DE PIS E COFINS" tem saldo -191.565,72 dentro do ramo de despesas
    financeiras: é um **crédito**, que reduz a despesa. Classificado DESPESA
    pelo ramo e passado por ``abs() * (-1)``, entregava -191.565,72 quando
    devia entregar +191.565,72 — a DRE errava por duas vezes o valor da conta,
    383.131,44, com o Balanço fechando normalmente.
    """
    from src.bp.output.build_gt_output import _RESULTADO, _valor_para_o_template
    from src.bp.output.template_map import TemplateProjector

    projector = TemplateProjector()
    credito = _valor_para_o_template(
        bruto=-191_565.72,
        escala=1000.0,
        codigo_origem="5.5.1.004.0001",
        codigo_template="3.01.01.09.01.08",
        projector=projector,
        orientacao={_RESULTADO: -1},
    )
    assert credito == pytest.approx(191.56572), (
        f"crédito entrou como {credito}; o sinal da origem foi apagado"
    )

    # E a despesa normal do mesmo ramo continua negativa.
    despesa = _valor_para_o_template(
        bruto=104_700.0, escala=1000.0, codigo_origem="5.5.1.002",
        codigo_template="3.01.01.09.01.08", projector=projector,
        orientacao={_RESULTADO: -1},
    )
    assert despesa == pytest.approx(-104.7)


def test_origem_de_natureza_implicita_continua_usando_o_rotulo_da_linha():
    """A trava não pode virar regra geral: sem sinal utilizável, quem decide
    é o rótulo do template, porque as fórmulas da DRE somam."""
    from src.bp.output.build_gt_output import _valor_para_o_template
    from src.bp.output.template_map import TemplateProjector

    despesa = _valor_para_o_template(
        bruto=50_000.0, escala=1000.0, codigo_origem="3.1.1.01",
        codigo_template="3.01.01.07.01.01",
        projector=TemplateProjector(), orientacao={},
    )
    assert despesa == pytest.approx(-50.0)
