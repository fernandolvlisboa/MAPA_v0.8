"""
O vão entre "o dado foi escrito" e "o dado aparece na entrega".

O problema
----------

Escrever uma linha em ``_dados_padronizados`` não põe número nenhum em BP_GT
ou DRE_GT. Quem soma são as fórmulas do template::

    =IFERROR(SUMIFS(_dados_padronizados!C:C, _dados_padronizados!$A:$A, $C9&"*"), 0)

Entre as duas coisas havia um vão que nenhum teste media, e ele erra dos dois
lados:

- **nenhuma linha do template captura o código** → o valor é escrito e não
  aparece em lugar nenhum. Pior que sumir: as contagens do Sumário classificam
  a conta como "tratada", então o relatório afirma 100% de match enquanto o
  dinheiro evaporou.
- **duas linhas capturam** → o mesmo valor entra duas vezes. É o efeito que o
  curinga ``*`` permite: ``1.01*`` também casa ``1.01.02``, de modo que uma
  linha "pai" e uma linha "filha" do template somariam o mesmo dado — a dupla
  contagem que o revisor apontou ao ver o asterisco na fórmula.

Hoje o template não tem prefixos aninhados. Isso é uma **propriedade dele**,
não uma garantia do sistema: uma edição no Excel que acrescente uma linha
totalizadora ao lado das analíticas quebra a entrega em silêncio. Estes testes
transformam a propriedade em conferência.

Metodologia
-----------

Conforme ``REVISAO_QUALIDADE.md`` §10, a trava sobre o corpus roda no
**controle fixo** (escolhido por forma) *mais* uma **amostra aleatória** com
semente impressa. Calibrar num balancete só é o que produz correção que
funciona em um arquivo e quebra o modelo nos outros.

Referência: ``REVISAO_QUALIDADE.md`` §13.
"""

from __future__ import annotations

import random
import re

import pytest
from conftest import CORPUS_DIR, corpus_disponivel
from openpyxl import load_workbook

from src.bp.output.build_gt_output import BuildResult, _conferir_captura, build_gt_output
from src.bp.output.template_map import LinhaTemplate, TemplateProjector

#: Mesmo controle de ``test_corpus_regressao.py`` — escolhidos por forma.
CONTROLE = (
    "Balancete 072022 122022 - RBM.xls",
    "Balancete SPEZZIA TUBOS 01012024-31122024.xls",
    "202404_2024 - Balancete.xls",
    "Balancete Real Life.xlsx",
    "Balancete_Trindade_052025.xlsx",
)
TAMANHO_AMOSTRA = 3
EXTENSOES = {".xls", ".xlsx", ".csv", ".txt"}

#: Extrai o critério de cada SUMIFS: ou um literal ``"1.01.01*"``, ou a forma
#: ``$C9&"*"`` que se resolve pela coluna C daquela linha.
_CRITERIO_RE = re.compile(
    r'_dados_padronizados!\$A:\$A\s*,\s*(?:"([^"]*)"|\$?C(\d+)\s*&\s*"([^"]*)")'
)


# ============================================================================
# 1. O modelo que o código tem do template bate com o template de verdade
# ============================================================================


def test_prefixos_lidos_batem_com_os_criterios_reais_das_formulas():
    """
    ``TemplateProjector`` modela o template lendo a **coluna C**; o Excel soma
    pelo **critério dentro da fórmula**. Se os dois divergirem, o código
    projeta para um alvo que não existe e o valor some.

    É a trava que pega uma edição de fórmula sem edição da coluna C — e
    vice-versa.
    """
    projector = TemplateProjector()
    wb = load_workbook(projector.template_path)

    das_formulas: set[str] = set()
    for aba in ("BP_GT", "DRE_GT"):
        if aba not in wb.sheetnames:
            continue
        ws = wb[aba]
        for r in range(1, ws.max_row + 1):
            formula = ws.cell(row=r, column=4).value
            if not (isinstance(formula, str) and "SUMIFS" in formula.upper()):
                continue
            for literal, ref, sufixo in _CRITERIO_RE.findall(formula):
                if literal:
                    das_formulas.add(literal.rstrip("*"))
                else:
                    base = str(ws.cell(row=int(ref), column=3).value or "").strip()
                    das_formulas.add(f"{base}{sufixo}".rstrip("*"))

    assert das_formulas, "nenhum critério SUMIFS lido — o probe está quebrado"
    assert das_formulas == set(projector.prefixes), (
        "o modelo do template (coluna C) divergiu dos critérios das fórmulas.\n"
        f"só nas fórmulas: {sorted(das_formulas - set(projector.prefixes))}\n"
        f"só na coluna C : {sorted(set(projector.prefixes) - das_formulas)}"
    )


def test_prefixos_do_template_nao_sao_aninhados():
    """
    A propriedade que torna o curinga ``*`` seguro.

    Enquanto nenhum prefixo do template for prefixo de outro, nenhum código
    cai em duas linhas — e o asterisco só faz o que deve: agregar os códigos
    referenciais mais profundos na linha da entrega.
    """
    aninhados = TemplateProjector().prefixos_aninhados()
    assert not aninhados, (
        "prefixos aninhados no template: todo valor sob o mais longo é somado "
        f"em DUAS linhas. Pares: {aninhados[:5]}"
    )


def test_todo_prefixo_pertence_a_exatamente_uma_linha():
    projector = TemplateProjector()
    assert projector.linhas, "template sem linhas de captura"
    for prefixo in projector.prefixes:
        donas = [linha for linha in projector.linhas if prefixo in linha.prefixos]
        assert len(donas) == 1, f"{prefixo!r} declarado em {len(donas)} linhas"


# ============================================================================
# 2. O detector detecta — testado contra template sintético
# ============================================================================


def _projector_sintetico(*linhas: tuple[str, ...]) -> TemplateProjector:
    """Um projector com linhas de captura fabricadas, para exercitar o detector.

    Sem isto, os testes do detector só provariam que ele não dispara — o que
    um detector permanentemente desligado também prova.
    """
    projector = TemplateProjector.__new__(TemplateProjector)
    projector.linhas = tuple(
        LinhaTemplate("BP_GT", i, f"linha {i}", prefixos)
        for i, prefixos in enumerate(linhas, start=1)
    )
    projector.prefixes = tuple(
        sorted({p for ps in linhas for p in ps}, key=lambda c: (-len(c), c))
    )
    return projector


def test_detecta_valor_que_nao_chega_a_entrega():
    projector = _projector_sintetico(("1.01",), ("2.01",))
    result = BuildResult(output_path=None)  # type: ignore[arg-type]
    dados = [
        {"codigo_padronizado": "1.01.05", "valor": 100.0},
        {"codigo_padronizado": "9.99.99", "valor": 42.5},
    ]
    _conferir_captura(dados, projector, result)

    assert result.linhas_sem_captura == [("9.99.99", 42.5)]
    assert result.valor_sem_captura == 42.5
    assert not result.captura_integra
    assert any("VALOR PERDIDO NA ENTREGA" in a for a in result.avisos)


def test_detecta_dupla_contagem_por_prefixo_aninhado():
    """O cenário exato que o asterisco permite: linha pai ao lado da filha."""
    projector = _projector_sintetico(("1.01",), ("1.01.05",))
    result = BuildResult(output_path=None)  # type: ignore[arg-type]
    _conferir_captura(
        [{"codigo_padronizado": "1.01.05.02", "valor": 80.0}], projector, result
    )

    assert result.linhas_capturadas_duas_vezes == [("1.01.05.02", 80.0, 2)]
    assert result.valor_contado_duas_vezes == 80.0
    assert not result.captura_integra
    assert any("DUPLA CONTAGEM" in a for a in result.avisos)
    assert any("TEMPLATE INCONSISTENTE" in a for a in result.avisos), (
        "a causa (prefixos aninhados) tem de ser nomeada, não só o efeito"
    )


def test_captura_integra_quando_cada_codigo_cai_em_uma_linha_so():
    projector = _projector_sintetico(("1.01",), ("2.01",))
    result = BuildResult(output_path=None)  # type: ignore[arg-type]
    _conferir_captura(
        [
            {"codigo_padronizado": "1.01.05", "valor": 100.0},
            {"codigo_padronizado": "2.01.01.03", "valor": -100.0},
        ],
        projector,
        result,
    )
    assert result.captura_integra
    assert not result.avisos


# ============================================================================
# 3. A trava sobre balancetes reais — controle + amostra aleatória
# ============================================================================


def _arquivos_do_corpus() -> list:
    if not corpus_disponivel():
        return []
    return sorted(p for p in CORPUS_DIR.iterdir() if p.suffix.lower() in EXTENSOES)


def _conferir_arquivo(caminho, tmp_path) -> BuildResult:
    resultado = build_gt_output(caminho, tmp_path / "saida.xlsx", ano_base=2024)
    if resultado.linhas_escritas == 0:
        pytest.skip(f"{caminho.name}: nenhuma linha escrita, nada a conferir")
    assert resultado.captura_integra, (
        f"{caminho.name}: {len(resultado.linhas_sem_captura)} linha(s) somando "
        f"{resultado.valor_sem_captura:,.2f} não aparecem na entrega; "
        f"{len(resultado.linhas_capturadas_duas_vezes)} contadas duas vezes "
        f"(+{resultado.valor_contado_duas_vezes:,.2f})"
    )
    return resultado


@pytest.mark.integration
@pytest.mark.parametrize("nome", CONTROLE)
def test_controle_toda_linha_escrita_aparece_uma_vez_na_entrega(nome, tmp_path):
    caminho = CORPUS_DIR / nome
    if not corpus_disponivel():
        pytest.skip(f"corpus ausente: {CORPUS_DIR}")
    if not caminho.exists():
        pytest.fail(f"controle ausente do corpus: {caminho}")
    resultado = _conferir_arquivo(caminho, tmp_path)
    print(f"[captura] {nome}: {resultado.linhas_escritas} linhas, todas capturadas 1x")


@pytest.fixture(scope="module")
def amostra_captura():
    """Sorteia fora do controle. A semente vai no log — teste aleatório sem
    semente impressa é teste que não se consegue depurar."""
    import os

    candidatos = [p for p in _arquivos_do_corpus() if p.name not in CONTROLE]
    if not candidatos:
        pytest.skip("corpus sem arquivos além do controle")
    bruta = os.environ.get("BP_SEED")
    semente = int(bruta) if bruta and bruta.isdigit() else random.randrange(1_000_000)
    print(f"\n[captura] semente da amostra = {semente}  (BP_SEED={semente} para repetir)")
    return random.Random(semente).sample(candidatos, min(TAMANHO_AMOSTRA, len(candidatos)))


@pytest.mark.integration
@pytest.mark.parametrize("indice", range(TAMANHO_AMOSTRA))
def test_amostra_toda_linha_escrita_aparece_uma_vez_na_entrega(
    amostra_captura, indice, tmp_path
):
    if indice >= len(amostra_captura):
        pytest.skip("amostra menor que o índice")
    caminho = amostra_captura[indice]
    resultado = _conferir_arquivo(caminho, tmp_path)
    print(f"[captura] {caminho.name}: {resultado.linhas_escritas} linhas, todas 1x")
