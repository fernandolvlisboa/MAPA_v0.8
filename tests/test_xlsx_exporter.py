"""
Testes do exportador .xlsx.

HISTÓRICO — por que este arquivo foi reescrito
----------------------------------------------
A versão anterior apontava para
``auxil/BP_teste/VIVAE - Balancete Acumulado 12.2023 ....xls``, que **não
existe** no repositório. O pipeline engole o erro de leitura
(``ParseyCaller.read()`` → ``except Exception: return None``) e o exporter
gera uma planilha bem-formada com **zero contas**. Resultado: os 5 testes
passavam sem exercitar nada — ``Accounts``, ``Unmatched`` e ``Validation``
tinham 0 linhas, e toda asserção era ou ``0 == 0`` ou um laço vazio.

Pior: o cabeçalho vazio mascarava um **índice de coluna errado**.
``test_unmatched_consistency`` lia ``row[8]`` como ``needs_review``, mas
``row[8]`` é ``saldo_atual`` — ``needs_review`` é ``row[14]``. Sobre dados
reais o teste falharia. Ele só era verde porque não havia dados.

Regra desta suíte: **nenhuma asserção sobre coleção pode passar vazia.**
Todo teste que varre linhas afirma antes que há linhas para varrer.
"""

from __future__ import annotations

import openpyxl
import pytest

from src.bp.exporters.xlsx_exporter import export_balance_sheet_to_xlsx

# Índices de coluna da aba Accounts (Contrato V2, 19 colunas).
# Ancorados por nome em test_accounts_layout_contrato_v2 — se o contrato
# mudar, aquele teste falha primeiro e aponta o que renumerar aqui.
COL_ACCOUNTS = {
    "nivel": 0,
    "codigo_original": 1,
    "codigo_alocado": 2,
    "descricao_original": 3,
    "descricao_plano_contas": 4,
    "saldo_anterior": 5,
    "credito": 6,
    "debito": 7,
    "saldo_atual": 8,
    "parent_id": 9,
    "is_analytical": 10,
    "match_codigo": 11,
    "match_descricao": 12,
    "match_score": 13,
    "needs_review": 14,
    "ignored": 15,
    "saldo_somado": 16,
    "rollup_diff": 17,
    "rollup_ok": 18,
}

COL_VALIDATION = {
    "codigo": 0,
    "descricao": 1,
    "saldo": 2,
    "saldo_calculado": 3,
    "diff": 4,
    "rel_diff_%": 5,
    "ok": 6,
}

ABAS_ESPERADAS = {
    "Summary",
    "Accounts",
    "Hierarchy",
    "Unmatched",
    "Variations",
    "Synonyms",
    "Validation",
}


@pytest.fixture(scope="module")
def workbook(tmp_path_factory, balancete_xls):
    """
    Exporta o balancete real UMA vez por módulo e devolve o workbook.

    Escopo de módulo porque o export leva ~4s (parse + matching contra o plano
    de contas inteiro); repetir por teste é o que fazia esta suíte custar caro
    sem ganho.
    """
    destino = tmp_path_factory.mktemp("export") / "export_test.xlsx"
    caminho = export_balance_sheet_to_xlsx(
        input_path=balancete_xls, output_path=destino
    )
    assert caminho.exists(), "Arquivo XLSX não foi criado"
    return openpyxl.load_workbook(caminho, data_only=True)


def _linhas(ws) -> list:
    """Linhas de dados (sem cabeçalho)."""
    return list(ws.iter_rows(min_row=2))


# ============================================================================
# Guarda de não-vacuidade — o teste que existe para os outros não mentirem
# ============================================================================


def test_export_produz_contas_reais(workbook):
    """
    Guarda contra a regressão que motivou esta reescrita.

    Se este teste falhar, TODOS os outros deste módulo viraram vacuosos:
    eles varrem linhas, e sem linhas eles passam sem validar nada.
    """
    accounts = _linhas(workbook["Accounts"])
    assert len(accounts) > 100, (
        f"Export produziu {len(accounts)} contas. Um número baixo indica que o "
        f"arquivo de entrada não foi lido (o pipeline engole o erro e gera "
        f"planilha vazia) — não que o balancete é pequeno."
    )


# ============================================================================
# Estrutura e contrato
# ============================================================================


def test_exporter_creates_workbook(workbook):
    faltando = ABAS_ESPERADAS - set(workbook.sheetnames)
    assert not faltando, f"Sheets faltando: {faltando}"


def test_accounts_layout_contrato_v2(workbook):
    """
    Ancora os nomes das 19 colunas do Contrato V2 nas posições que os demais
    testes assumem. É este teste que impede um índice errado de passar
    despercebido (foi assim que ``needs_review`` virou ``saldo_atual``).
    """
    ws = workbook["Accounts"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert len(headers) == 19, f"Contrato V2 tem 19 colunas, achei {len(headers)}"
    for nome, idx in COL_ACCOUNTS.items():
        assert headers[idx] == nome, (
            f"Coluna {idx} deveria ser {nome!r}, é {headers[idx]!r}. "
            f"Atualize COL_ACCOUNTS neste arquivo."
        )


def test_validation_layout(workbook):
    ws = workbook["Validation"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for nome, idx in COL_VALIDATION.items():
        assert headers[idx] == nome, (
            f"Coluna {idx} de Validation deveria ser {nome!r}, é {headers[idx]!r}"
        )


# ============================================================================
# Métricas e consistência entre abas
# ============================================================================


def test_summary_metrics(workbook):
    ws = workbook["Summary"]
    metrics = {
        row[0].value: row[1].value for row in ws.iter_rows(min_row=2) if row[0].value
    }
    for key in [
        "Total Accounts",
        "Matched",
        "Needs Review",
        "Ignored",
        "Rollup Discrepancies",
    ]:
        assert key in metrics, f"Métrica {key} ausente"
        assert isinstance(metrics[key], (int, float)), f"Métrica {key} não numérica"

    assert metrics["Total Accounts"] == len(_linhas(workbook["Accounts"])), (
        "Summary.Total Accounts não bate com o número de linhas em Accounts"
    )


def test_unmatched_consistency(workbook):
    """
    Contas com ``needs_review`` verdadeiro devem ser exatamente as listadas em
    Unmatched.

    A versão anterior lia ``row[8]`` (``saldo_atual``) em vez de ``row[14]``
    (``needs_review``) e só passava porque a planilha estava vazia.
    """
    accounts = _linhas(workbook["Accounts"])
    assert accounts, "sem contas — o teste seria vacuoso"

    needs_review = sum(
        1 for row in accounts if bool(row[COL_ACCOUNTS["needs_review"]].value)
    )
    unmatched = len(_linhas(workbook["Unmatched"]))
    assert needs_review == unmatched, (
        f"needs_review={needs_review} vs linhas em Unmatched={unmatched}"
    )


def test_validation_diff_columns(workbook):
    """Toda linha de Validation tem código e colunas numéricas coerentes."""
    linhas = _linhas(workbook["Validation"])
    assert linhas, "Validation vazia — o teste seria vacuoso"

    for row in linhas:
        codigo = row[COL_VALIDATION["codigo"]].value
        saldo = row[COL_VALIDATION["saldo"]].value
        saldo_calc = row[COL_VALIDATION["saldo_calculado"]].value
        diff = row[COL_VALIDATION["diff"]].value
        rel = row[COL_VALIDATION["rel_diff_%"]].value

        assert codigo, "Código vazio em Validation"
        for nome, valor in (
            ("saldo", saldo),
            ("saldo_calculado", saldo_calc),
            ("diff", diff),
            ("rel_diff_%", rel),
        ):
            assert isinstance(valor, (int, float)), (
                f"{nome} não numérico na conta {codigo}: {valor!r}"
            )
        # `diff` é definido como saldo - saldo_calculado; se divergir, o
        # relatório de discrepância está mentindo sobre a própria conta.
        assert abs((saldo - saldo_calc) - diff) < 0.01, (
            f"diff inconsistente na conta {codigo}: "
            f"{saldo} - {saldo_calc} != {diff}"
        )


def test_discrepancy_summary_matches_validation(workbook):
    """Summary.Rollup Discrepancies == linhas com ok=False em Validation."""
    summary_metrics = {
        row[0].value: row[1].value
        for row in workbook["Summary"].iter_rows(min_row=2)
        if row[0].value
    }
    discrep_metric = summary_metrics.get("Rollup Discrepancies")
    assert discrep_metric is not None, "Métrica Rollup Discrepancies não encontrada"

    linhas = _linhas(workbook["Validation"])
    assert linhas, "Validation vazia — o teste seria vacuoso"
    failures = sum(1 for row in linhas if row[COL_VALIDATION["ok"]].value is False)

    assert failures == discrep_metric, (
        f"Discrepâncias mismatch: Summary={discrep_metric} vs Validation={failures}"
    )


def test_accounts_e_hierarchy_cobrem_o_mesmo_conjunto(workbook):
    """Hierarchy é uma view de Accounts — não pode perder nem inventar conta."""
    accounts = _linhas(workbook["Accounts"])
    hierarchy = _linhas(workbook["Hierarchy"])
    assert accounts and hierarchy, "abas vazias — o teste seria vacuoso"
    assert len(accounts) == len(hierarchy), (
        f"Accounts tem {len(accounts)} linhas e Hierarchy {len(hierarchy)}"
    )


def test_original_tab_handles_pd_na(tmp_path, balancete_xls):
    """pd.NA no DataFrame original não pode estourar a escrita do Excel."""
    import pandas as pd

    out = tmp_path / "orig_na_export.xlsx"
    original_df = pd.DataFrame({"colA": [1, pd.NA], "colB": [pd.NA, "x"]})
    export_balance_sheet_to_xlsx(balancete_xls, out, original_data=original_df)
    assert out.exists()
    wb = openpyxl.load_workbook(out, data_only=True)
    assert "Original" in wb.sheetnames
    ws = wb["Original"]
    assert ws.max_row == 3
    assert ws["A3"].value is None or ws["B2"].value is None
