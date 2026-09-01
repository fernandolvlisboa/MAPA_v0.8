from __future__ import annotations

from pathlib import Path
import pytest
from openpyxl import load_workbook
import pandas as pd

from src.bp.exporters.xlsx_exporter import export_balance_sheet_to_xlsx


@pytest.mark.integration
def test_export_spezzia_full(tmp_path: Path):
    input_path = Path(
        "data/samples/Balancete SPEZZIA TUBOS 01012024-31122024.xls"
    )
    if not input_path.exists():
        pytest.skip("SPEZZIA training file not present in workspace")

    out = tmp_path / "spezzia_export.xlsx"
    export_balance_sheet_to_xlsx(input_path, out)
    assert out.exists()

    wb = load_workbook(out, data_only=True)
    expected_tabs = [
        "Summary",
        "Accounts",
        "Hierarchy",
        "Unmatched",
        "Variations",
        "Synonyms",
        "Validation",
        "Original",
    ]
    for name in expected_tabs:
        assert name in wb.sheetnames

    # Accounts: 19 columns per Contract V2
    ws_acc = wb["Accounts"]
    headers = [c.value for c in next(ws_acc.iter_rows(min_row=1, max_row=1))]
    assert len(headers) == 19

    # Original: has at least header + one data row
    ws_orig = wb["Original"]
    assert ws_orig.max_row >= 2


def test_export_original_sheet_handles_pd_na(tmp_path: Path):
    # Use a stable training file for parsing; inject pd.NA DataFrame into Original tab
    input_path = Path("data/samples/Balancete Real Life.xlsx")
    if not input_path.exists():
        pytest.skip("Training file not present in workspace")

    out = tmp_path / "orig_na_export.xlsx"
    original_df = pd.DataFrame({"colA": [1, pd.NA], "colB": [pd.NA, "x"]})

    export_balance_sheet_to_xlsx(input_path, out, original_data=original_df)
    assert out.exists()

    wb = load_workbook(out, data_only=True)
    assert "Original" in wb.sheetnames
    ws = wb["Original"]
    # Two data rows + header
    assert ws.max_row == 3
    # pd.NA should be written as empty cells
    assert ws["A3"].value is None or ws["B2"].value is None
