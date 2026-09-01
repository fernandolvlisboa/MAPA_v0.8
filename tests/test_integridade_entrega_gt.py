"""
Integridade do arquivo de entrega ao cliente (Template GT).

``build_gt_output`` é o **último elo** da cadeia: o .xlsx que sai daqui é o que
o cliente recebe. Ele consome ``ParseyCaller(input_path).parse()``, então
herda tudo o que o dispatcher entrega — para o bem e para o mal.

Antes da unificação da conversão numérica, 3 de 5 contas chegavam com valor
errado, **todas com ``score = 1.0``** na aba de auditoria e nenhuma em "Contas
Não Identificadas": para o analista que conferia a entrega, o processamento
tinha sido perfeito.

===================  ==============  ==============  ==============
Conta                Valor real      Antes           Agora
===================  ==============  ==============  ==============
CAIXA GERAL          ``1.234,56``    1234.56         1234.56
BANCOS               ``1234.56``     **123456**      1234.56
CLIENTES             ``(5.000,00)``  **0**           5000 (sinal §4)
FORNECEDORES         ``3.000,00 C``  **0**           3000
ESTOQUES             ``2.000,00``    2000            2000
===================  ==============  ==============  ==============

Este arquivo é a trava dessa correção no ponto em que ela mais importa.
Referência: ``REVISAO_QUALIDADE.md`` §7.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.bp.output.build_gt_output import build_gt_output

pytestmark = [pytest.mark.contrato, pytest.mark.integration]

TEMPLATE = Path("templates/Template_GT_BP_Padrao_v3.xlsx")

# Linha 1 de "Contas Tratadas" é um aviso de uso interno; o cabeçalho é a 2.
_LINHA_CABECALHO_TRATADAS = 2
_COL_DESCRICAO = 2
_COL_VALOR = 6
_COL_SCORE = 7


@pytest.fixture(scope="module")
def entrega(tmp_path_factory):
    """
    Gera uma entrega a partir de um balancete sintético com os quatro formatos
    numéricos que aparecem em balancete brasileiro real.

    ``escala=1.0`` para que os números na entrega sejam diretamente comparáveis
    com a entrada (o default divide por 1000).
    """
    if not TEMPLATE.exists():
        pytest.skip(f"Template GT ausente: {TEMPLATE}")

    tmp = tmp_path_factory.mktemp("gt")
    origem = tmp / "balancete.xlsx"
    pd.DataFrame(
        [
            ("1.1.01.01", "CAIXA GERAL", "1.234,56"),  # BR canônico
            ("1.1.01.02", "BANCOS CONTA MOVIMENTO", "1234.56"),  # decimal c/ ponto
            ("1.1.02.01", "CLIENTES", "(5.000,00)"),  # negativo entre parênteses
            ("2.1.01.01", "FORNECEDORES", "3.000,00 C"),  # sufixo credor
            ("1.1.03.01", "ESTOQUES", "2.000,00"),  # BR canônico
        ],
        columns=["Codigo", "Descricao", "Saldo"],
    ).to_excel(origem, index=False)

    saida = tmp / "entrega.xlsx"
    resultado = build_gt_output(
        origem,
        saida,
        ano_base=2024,
        cache_path=tmp / "cache.json",  # nunca o cache compartilhado
        escala=1.0,
    )
    wb = load_workbook(saida)
    ws = wb["Contas Tratadas"]
    valores = {
        row[_COL_DESCRICAO]: {"valor": row[_COL_VALOR], "score": row[_COL_SCORE]}
        for row in ws.iter_rows(
            min_row=_LINHA_CABECALHO_TRATADAS + 1, values_only=True
        )
        if row[_COL_DESCRICAO]
    }
    return resultado, valores


# ============================================================================
# 1. O que a entrega acerta — trava de não-regressão
# ============================================================================


def test_entrega_processa_todas_as_contas(entrega):
    """Guarda de não-vacuidade: sem contas, todo teste abaixo é vazio."""
    resultado, valores = entrega
    assert resultado.contas_lidas == 5
    assert len(valores) == 5, f"contas ausentes do de-para: {sorted(valores)}"


@pytest.mark.parametrize(
    "conta,esperado", [("CAIXA GERAL", 1234.56), ("ESTOQUES", 2000.0)]
)
def test_formato_br_canonico_chega_correto(entrega, conta, esperado):
    _, valores = entrega
    assert valores[conta]["valor"] == pytest.approx(esperado)


# ============================================================================
# 2. Os formatos que antes chegavam errados ao cliente
# ============================================================================


def test_nenhum_valor_entregue_e_zero_indevido(entrega):
    """
    Nenhuma das 5 contas pode chegar zerada: todas têm saldo na origem.

    Era o sintoma mais perigoso — a conta zerada é plausível, some do balanço
    sem entrar na fila de revisão, e sai com score 1.0 na auditoria.
    """
    _, valores = entrega
    zeradas = [c for c, v in valores.items() if v["valor"] == 0]
    assert not zeradas, f"contas com saldo na origem chegaram zeradas: {zeradas}"


def test_decimal_com_ponto_nao_infla_na_entrega(entrega):
    _, valores = entrega
    assert valores["BANCOS CONTA MOVIMENTO"]["valor"] == pytest.approx(1234.56)


def test_negativo_entre_parenteses_nao_zera_na_entrega(entrega):
    _, valores = entrega
    assert valores["CLIENTES"]["valor"] != 0


def test_sufixo_dc_nao_zera_na_entrega(entrega):
    _, valores = entrega
    assert valores["FORNECEDORES"]["valor"] != 0


# ============================================================================
# 3. A rede que existe — e para onde ela aponta
# ============================================================================


def test_desequilibrio_e_acusado(entrega):
    """
    O balanço desequilibrado é acusado — e com a reconciliação, não com um
    "Ativo != Passivo" que não diz nada ao analista.
    """
    resultado, _ = entrega
    assert resultado.avisos, "nenhum aviso emitido apesar do balanço desequilibrado"
    assert any("Não fecha por" in a for a in resultado.avisos)


def test_aviso_nomeia_saldo_ilegivel_como_causa(tmp_path):
    """
    Quando há saldo ilegível, o aviso precisa dizer isso — antes ele culpava
    a convenção de sinais e mandava o analista revisar uma aba vazia.
    """
    if not TEMPLATE.exists():
        pytest.skip(f"Template GT ausente: {TEMPLATE}")

    origem = tmp_path / "com_lixo.xlsx"
    pd.DataFrame(
        [
            ("1.1.01.01", "CAIXA GERAL", "1.234,56"),
            ("1.1.01.02", "BANCOS CONTA MOVIMENTO", "ilegível"),
            ("2.1.01.01", "FORNECEDORES", "3.000,00 C"),
        ],
        columns=["Codigo", "Descricao", "Saldo"],
    ).to_excel(origem, index=False)

    resultado = build_gt_output(
        origem,
        tmp_path / "saida.xlsx",
        ano_base=2024,
        cache_path=tmp_path / "c.json",
        escala=1.0,
    )
    assert resultado.saldos_ilegiveis == 1
    texto = " ".join(resultado.avisos).lower()
    assert "ilegível" in texto, f"aviso não nomeia a causa: {resultado.avisos}"
    assert "convenção de sinais" not in texto, (
        "o aviso continua culpando a convenção de sinais quando a causa é o "
        "saldo ilegível"
    )


def test_entrega_limpa_nao_reporta_saldo_ilegivel(entrega):
    """Trava o contrário: sem saldo ilegível, o aviso não aparece."""
    resultado, _ = entrega
    assert resultado.saldos_ilegiveis == 0
    assert not any("ilegível" in a for a in resultado.avisos)


def test_parser_vazio_e_acusado(tmp_path):
    """
    Quando o parser não extrai nada, ``_validar`` acusa em vez de entregar uma
    planilha vazia em silêncio. É a última defesa do analista e precisa
    continuar valendo.
    """
    if not TEMPLATE.exists():
        pytest.skip(f"Template GT ausente: {TEMPLATE}")

    vazio = tmp_path / "sem_contas.xlsx"
    pd.DataFrame({"ColunaA": [1, 2], "ColunaB": ["x", "y"]}).to_excel(
        vazio, index=False
    )
    resultado = build_gt_output(
        vazio, tmp_path / "saida.xlsx", ano_base=2024, cache_path=tmp_path / "c.json"
    )
    assert resultado.linhas_escritas == 0
    assert any("Nenhuma linha escrita" in a for a in resultado.avisos)


# ============================================================================
# 4. O sinal do saldo é descartado
# ============================================================================


@pytest.fixture(scope="module")
def entrega_negativa(tmp_path_factory):
    """Balancete com um saldo negativo legível (float, sem ambiguidade)."""
    if not TEMPLATE.exists():
        pytest.skip(f"Template GT ausente: {TEMPLATE}")

    tmp = tmp_path_factory.mktemp("gt_neg")
    origem = tmp / "neg.xlsx"
    pd.DataFrame(
        [("1.1.01.02", "BANCOS CONTA MOVIMENTO", -5000.0)],
        columns=["Codigo", "Descricao", "Saldo"],
    ).to_excel(origem, index=False)

    saida = tmp / "neg_out.xlsx"
    build_gt_output(
        origem, saida, ano_base=2024, cache_path=tmp / "c.json", escala=1.0
    )
    ws = load_workbook(saida)["Contas Tratadas"]
    linhas = [
        row
        for row in ws.iter_rows(min_row=_LINHA_CABECALHO_TRATADAS + 1, values_only=True)
        if row[_COL_DESCRICAO]
    ]
    assert linhas, "de-para vazio — o teste seria vacuoso"
    return linhas[0][_COL_VALOR]


def test_saldo_credor_em_conta_de_ativo_chega_negativo(entrega_negativa):
    """
    O sinal da origem sobrevive até o cliente — era o oposto disso.

    Era ``abs(valor) * sign_for(codigo_template)``: o sinal lido era jogado
    fora e rederivado da convenção da *linha do template*. Como nenhuma linha
    do BP_GT é "(-)", um saldo credor numa conta de ativo (cheque especial,
    banco a descoberto) chegava ao cliente como ativo **positivo** — e a
    anomalia virava um número plausível.

    O efeito não era cosmético: no balancete que expôs isso, depreciação e
    amortização acumuladas entraram positivas e inflaram o Ativo em
    R$ 322.453,04 (2.683.506,57 entregues contra 2.361.053,53 na origem).

    Ver REVISAO_QUALIDADE.md §15.
    """
    assert entrega_negativa == pytest.approx(-5000.0), (
        f"saldo credor de -5.000,00 chegou como {entrega_negativa}; o sinal da "
        "origem foi perdido de novo — é o defeito do §15 voltando"
    )


# ============================================================================
# 5. Hierarquia: invariantes sobre VÁRIOS balancetes, não sobre um
# ============================================================================
#
# Nada aqui assere sobre conta específica de balancete específico. Uma conta
# sem destino no template ("PARCELAMENTOS", no RBM) é PARTICULARIDADE
# ESPERADA — vai acontecer num percentual alto dos clientes. Fixar teste nela
# amarra o modelo a um arquivo. O que se exige é a invariante: seja lá o que
# fique de fora, tem de ser reconciliado, e nada pode evaporar.
#
# A regra metodológica está em REVISAO_QUALIDADE.md §10.

CORPUS = Path("data/samples")

#: Balancetes de controle COM hierarquia. Escolhidos por forma (emissores e
#: profundidades diferentes), não por conveniência. A varredura do corpus
#: inteiro está em ``test_corpus_regressao.py``.
CONTROLE_COM_HIERARQUIA = [
    "Balancete 072022 122022 - RBM.xls",
    "Balancete SPEZZIA TUBOS 01012024-31122024.xls",
    "202404_2024 - Balancete.xls",
]


def _entrega(nome: str, ano: int = 2022):
    """Gera a entrega de um balancete do corpus, ou pula."""
    import tempfile

    if not TEMPLATE.exists():
        pytest.skip(f"Template GT ausente: {TEMPLATE}")
    caminho = CORPUS / nome
    if not caminho.exists():
        pytest.skip(f"corpus ausente: {caminho}")

    tmp = Path(tempfile.mkdtemp())
    saida = tmp / "saida.xlsx"
    return build_gt_output(
        caminho, saida, ano_base=ano, cache_path=tmp / "c.json"
    ), saida


@pytest.mark.parametrize("nome", CONTROLE_COM_HIERARQUIA)
def test_origem_e_aritmeticamente_integra(nome):
    """
    Todo agrupador do balancete de origem confere com a soma dos filhos.

    É a checagem mais barata e mais forte sobre a extração: se ela passa, o
    parser leu tudo, leu certo, e não perdeu nem inventou linha.
    """
    from src.bp.parsers.dispatcher import ParseyCaller
    from src.bp.validators.hierarquia import conferir_hierarquia

    caminho = CORPUS / nome
    if not caminho.exists():
        pytest.skip(f"corpus ausente: {caminho}")

    relatorio = conferir_hierarquia(ParseyCaller(caminho).parse())
    assert relatorio.tem_hierarquia, "sem árvore, o teste seria vacuoso"
    assert relatorio.rollup_integro, (
        f"{nome}: {relatorio.pais_divergentes} agrupador(es) não fecham — "
        + "; ".join(str(d) for d in relatorio.divergencias[:3])
    )
    assert relatorio.equacao_fecha, (
        f"{nome}: Ativo + Passivo + Resultado = {relatorio.desequilibrio:,.2f}"
    )


@pytest.mark.parametrize("nome", CONTROLE_COM_HIERARQUIA)
def test_cobertura_completa(nome):
    """
    **A invariante central.** Por classe contábil, o que foi emitido mais o que
    ficou de fora reproduz exatamente o total da origem.

    Quebrar isso significa conta contada duas vezes (agrupador e filhos juntos)
    ou conta perdida (folha sem destino) — os dois erros que faziam o balanço
    não fechar e que coexistiam sem que nenhum teste os pegasse.
    """
    resultado, _ = _entrega(nome)

    assert resultado.contas_lidas > 100, "sem contas, o teste seria vacuoso"
    assert resultado.hierarquia.tem_hierarquia

    for classe, origem in resultado.hierarquia.totais_por_classe.items():
        emitido = resultado.emitido_por_classe.get(classe, 0.0)
        fora = resultado.nao_coberto_por_classe.get(classe, 0.0)
        assert emitido + fora == pytest.approx(origem, abs=0.01), (
            f"{nome} / {classe}: emitido {emitido:,.2f} + fora {fora:,.2f} = "
            f"{emitido + fora:,.2f}, mas a origem tem {origem:,.2f}"
        )
    assert resultado.cobertura_completa


@pytest.mark.parametrize("nome", CONTROLE_COM_HIERARQUIA)
def test_nenhum_codigo_emitido_e_ancestral_de_outro(nome):
    """Dupla contagem: emitir o agrupador E os filhos soma o ramo duas vezes."""
    import tempfile
    from collections import defaultdict

    from src.bp.output.build_gt_output import _build_matcher, _resolver
    from src.bp.output.template_map import TemplateProjector
    from src.bp.parsers.dispatcher import ParseyCaller
    from src.bp.validators.hierarquia import selecionar_para_projecao

    caminho = CORPUS / nome
    if not caminho.exists() or not TEMPLATE.exists():
        pytest.skip("corpus ou template ausente")

    tmp = Path(tempfile.mkdtemp())
    contas = ParseyCaller(caminho).parse()
    matcher = _build_matcher(None, tmp / "c.json")
    projector = TemplateProjector()

    por_codigo = defaultdict(list)
    for r in (_resolver(c, matcher, projector) for c in contas):
        por_codigo[str(r.conta.get("codigo", "")).strip()].append(r)

    selecao = selecionar_para_projecao(
        contas, lambda c: any(x.codigo_template for x in por_codigo.get(c, ()))
    )
    assert selecao.codigos, f"{nome}: seleção vazia — o teste seria vacuoso"

    emitidos = set(selecao.codigos)
    for codigo in emitidos:
        partes = codigo.split(".")
        ancestrais = {".".join(partes[:n]) for n in range(1, len(partes))}
        conflito = ancestrais & emitidos
        assert not conflito, f"{nome}: {codigo} emitido junto com {conflito}"


#: Piso de cobertura de VALOR, medido sobre o corpus.
#:
#: Contar contas não serve como métrica: um código emitido pode cobrir várias
#: homônimas, e uma folha absorvida pelo agrupador não é conta perdida. O que
#: interessa é quanto do dinheiro da origem chega à entrega.
#:
#: Medido em 7 balancetes: quatro cobrem 100%, dois passam de 99%, e o RBM
#: cobre 88,6% — **o RBM é o pior caso do corpus, não o representativo**. O
#: piso fica abaixo dele com folga: o objetivo é pegar regressão, não
#: cristalizar o outlier.
PISO_COBERTURA_DE_VALOR = 0.85


@pytest.mark.parametrize("nome", CONTROLE_COM_HIERARQUIA)
def test_cobertura_de_valor_acima_do_piso(nome):
    """
    Quanto do valor da origem chega à entrega.

    Contas com nome próprio do cliente não casam com plano de contas nenhum —
    e não precisam: são absorvidas pelo agrupador, e o valor delas viaja no
    total dele. O que este teste impede é o valor sumir.
    """
    resultado, _ = _entrega(nome)
    cobertura = resultado.cobertura_de_valor
    assert cobertura >= PISO_COBERTURA_DE_VALOR, (
        f"{nome}: só {cobertura:.2%} do valor chegou à entrega "
        f"({abs(resultado.valor_nao_coberto):,.2f} sem destino em "
        f"{resultado.contas_nao_identificadas} conta(s))"
    )


def test_cobertura_de_valor_e_coerente_com_a_reconciliacao():
    """
    Sem hierarquia não há o que cobrir — a métrica não pode fingir 0%.
    E cobertura 100% implica reconciliação que fecha.
    """
    from src.bp.output.build_gt_output import BuildResult

    vazio = BuildResult(output_path=Path("x"))
    assert vazio.cobertura_de_valor == 1.0
    assert vazio.reconciliacao.fecha


# ============================================================================
# 6. Reconciliação: se não fecha, tem de estar explicado
# ============================================================================


@pytest.mark.parametrize("nome", CONTROLE_COM_HIERARQUIA)
def test_reconciliacao_explica_a_diferenca(nome):
    """
    Dizer "não fecha" não serve. Ou fecha, ou a diferença **é exatamente** a
    soma de N contas nomeadas — aí o analista sabe que não há nada escondido.

    Note que o teste **não exige que o balancete não feche**: se fechar, ótimo.
    O que não se admite é diferença sem explicação. Uma conta sem destino no
    template é particularidade esperada; resíduo inexplicado é defeito.
    """
    resultado, _ = _entrega(nome)
    reconc = resultado.reconciliacao

    assert reconc.residuo == pytest.approx(0.0, abs=0.01), (
        f"{nome}: sobrou {reconc.residuo:,.2f} sem explicação — há conta "
        f"contada duas vezes ou perdida"
    )
    assert reconc.explicada

    if reconc.fecha:
        assert not reconc.contas, "fecha mas há contas sem destino?"
        return

    assert reconc.contas, f"{nome}: não fecha e não lista o que faltou"
    assert reconc.soma_sem_destino == pytest.approx(
        sum(c.valor for c in reconc.contas), abs=0.01
    )
    for conta in reconc.contas:
        assert conta.codigo and conta.descricao and conta.motivo
        assert conta.valor != 0, "conta zerada não explica diferença nenhuma"

    valores = [abs(c.valor) for c in reconc.contas]
    assert valores == sorted(valores, reverse=True), "maior primeiro"

    texto = reconc.mensagem()
    assert f"{abs(reconc.desequilibrio):,.2f}" in texto
    assert "explicam 100%" in texto


@pytest.mark.parametrize("nome", CONTROLE_COM_HIERARQUIA)
def test_sumario_traz_a_reconciliacao(nome):
    """A prova precisa chegar ao arquivo entregue, não só ao log."""
    resultado, saida = _entrega(nome)
    reconc = resultado.reconciliacao

    ws = load_workbook(saida)["Sumário"]
    celulas = [c for row in ws.iter_rows(values_only=True) for c in row if c is not None]
    texto = " ".join(str(c) for c in celulas)

    if reconc.fecha:
        assert "POR QUE NÃO FECHA" not in texto
        return

    assert "POR QUE NÃO FECHA" in texto
    assert "CONTAS QUE EXPLICAM A DIFERENÇA" in texto
    assert "Resíduo sem explicação" in texto
    for conta in reconc.contas:
        assert conta.codigo in texto, f"{conta.codigo} não foi listada no Sumário"
    assert round(reconc.soma_sem_destino, 2) in celulas


def test_reconciliacao_denuncia_residuo_inexplicado():
    """
    O caso perigoso: a diferença NÃO é explicada pelas contas sem destino.
    Significa conta contada duas vezes ou perdida — a mensagem tem de gritar,
    não tranquilizar.
    """
    from src.bp.output.build_gt_output import ContaSemDestino, Reconciliacao

    reconc = Reconciliacao(
        desequilibrio=1000.0,
        soma_sem_destino=-400.0,
        contas=[ContaSemDestino("2.1", "FORNECEDORES", -400.0, "sem match")],
    )
    assert not reconc.explicada
    assert reconc.residuo == pytest.approx(600.0)
    texto = reconc.mensagem()
    assert "ATENÇÃO" in texto
    assert "600.00" in texto or "600,00" in texto


def test_reconciliacao_quando_fecha():
    from src.bp.output.build_gt_output import Reconciliacao

    reconc = Reconciliacao(desequilibrio=0.0, soma_sem_destino=0.0)
    assert reconc.fecha
    assert "fecha" in reconc.mensagem().lower()


@pytest.mark.integration
def test_balancete_txt_chega_ao_cliente(tmp_path):
    """
    Regressão de ``.txt``: contas que antes eram descartadas em silêncio pelo
    roteamento agora chegam à entrega com saldo real.
    """
    if not TEMPLATE.exists():
        pytest.skip(f"Template GT ausente: {TEMPLATE}")
    txt = CORPUS / "2019-01.TXT"
    if not txt.exists():
        pytest.skip("corpus ausente")

    resultado = build_gt_output(
        txt, tmp_path / "txt.xlsx", ano_base=2019, cache_path=tmp_path / "c.json"
    )
    assert resultado.contas_lidas > 400
    assert resultado.total_ativo != 0, "ativo zerado: o saldo não chegou"
    # Poucas linhas para muitas contas é o esperado: a hierarquia absorve as
    # folhas no agrupador. O que não pode é o VALOR se perder.
    assert resultado.linhas_escritas > 20
    assert resultado.contas_absorvidas > 100


# ============================================================================
# Trava de classe: conta de resultado não entra no Balanço
# ============================================================================
#
# O revisor apontou, sobre uma entrega real: "custos e despesas foram parar no
# balanço, a conta é claramente de resultado". A causa-raiz era a coluna de
# código não ser detectada (ver `test_deteccao_codigo.py`): sem código, a
# classe da origem vira None, o Plano C desliga, e "Aluguel e Condomínio A
# PAGAR" (passivo) casa com "Condomínio" (despesa 3.x) com score 1.0.
#
# A causa foi corrigida. Estes testes travam a CONSEQUÊNCIA, que é o que o
# cliente vê — e travam no ponto final, sobre o código do template, onde a
# conferência não depende de nenhuma etapa anterior ter funcionado.


def _classe(codigo: str) -> str | None:
    from src.bp.utils.codigo import classe_from_codigo

    return classe_from_codigo(codigo)


@pytest.mark.parametrize("nome", CONTROLE_COM_HIERARQUIA)
def test_nenhuma_conta_muda_de_classe_ao_ser_projetada(nome, tmp_path):
    """
    Ativo continua no Ativo, resultado continua na DRE.

    Uma projeção que troca a classe não é "um match ruim": ela corrompe o
    Balanço e a DRE ao mesmo tempo, tirando valor de um e pondo no outro.
    """
    caminho = CORPUS / nome
    if not TEMPLATE.exists():
        pytest.skip(f"Template GT ausente: {TEMPLATE}")
    if not caminho.exists():
        pytest.skip(f"corpus ausente: {caminho}")
    resultado = build_gt_output(
        caminho, tmp_path / "saida.xlsx", ano_base=2024, cache_path=tmp_path / "c.json"
    )
    wb = load_workbook(resultado.output_path)
    ws = wb["Contas Tratadas"]

    linhas = [
        (str(r[1]), str(r[5]))  # codigo_original, codigo_template
        for r in ws.iter_rows(min_row=3, values_only=True)
        if r[1] and r[5]
    ]
    assert linhas, f"{nome}: nenhuma conta tratada — o teste passaria vazio"

    cruzam = [
        (co, ct)
        for co, ct in linhas
        if _classe(co) and _classe(ct) and _classe(co) != _classe(ct)
    ]
    assert not cruzam, (
        f"{nome}: {len(cruzam)} conta(s) mudaram de classe na projeção "
        f"(ex.: {cruzam[:3]}) — resultado indo para o Balanço, ou vice-versa"
    )


def test_a_trava_de_classe_recusa_em_vez_de_classificar_errado():
    """
    Não-vacuidade: a trava tem de *disparar* quando a classe cruza.

    Sem isto, o teste acima também passaria com a trava permanentemente
    desligada — que é exatamente o estado em que o defeito chegou ao cliente.
    """
    from src.bp.output.build_gt_output import _resolver
    from src.bp.output.template_map import ProjectionResult

    class _Decisao:
        codigo = "3.01.01.07.01.04"
        descricao = "Condomínio"
        score = 1.0

    class _MatcherFalso:
        natureza_referencial: dict[str, str] = {}

        def match(self, descricao, codigo_origem="", natureza_resultado=None, prazo=None):
            return type(
                "R", (), {"decision": _Decisao(), "needs_review": False}
            )()

    class _ProjectorFalso:
        def project(self, codigo):
            return ProjectionResult(codigo, "3.01.01.07.01.04", "direto")

    conta = {"codigo": "2.1.2.05.001", "descricao": "ALUGUEL E CONDOMINIO A PAGAR"}
    resolucao = _resolver(conta, _MatcherFalso(), _ProjectorFalso())

    assert resolucao.codigo_template is None, (
        "conta de PASSIVO foi aceita num código de RESULTADO"
    )
    assert "classe incompatível" in resolucao.motivo
    assert "PASSIVO" in resolucao.motivo and "RESULTADO" in resolucao.motivo


def test_a_trava_de_classe_nao_recusa_projecao_legitima():
    """A trava não pode ser um bloqueio geral: o caso normal tem de passar."""
    from src.bp.output.build_gt_output import _resolver
    from src.bp.output.template_map import ProjectionResult

    class _Decisao:
        codigo = "1.01.01.01"
        descricao = "Caixa"
        score = 1.0

    class _MatcherFalso:
        natureza_referencial: dict[str, str] = {}

        def match(self, descricao, codigo_origem="", natureza_resultado=None, prazo=None):
            return type("R", (), {"decision": _Decisao(), "needs_review": False})()

    class _ProjectorFalso:
        def project(self, codigo):
            return ProjectionResult(codigo, "1.01.01.01", "direto")

    resolucao = _resolver(
        {"codigo": "1.1.1.01.001", "descricao": "CAIXA"},
        _MatcherFalso(),
        _ProjectorFalso(),
    )
    assert resolucao.codigo_template == "1.01.01.01"
    assert not resolucao.motivo
