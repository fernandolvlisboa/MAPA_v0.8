"""
TemplateProjector — projeta códigos internos nos códigos que o Template GT consome.

## O problema que isto resolve

O Template GT agrega por ``SUMIFS(... , $C<linha> & "*")``: cada linha de
``BP_GT``/``DRE_GT`` tem na coluna C um código ECF e captura tudo que **começa
com** esse prefixo. São 86 prefixos no total.

Qualquer código escrito em ``_dados_padronizados`` que não comece por um desses
86 prefixos **não é somado por ninguém** — o valor simplesmente não aparece no
relatório, sem erro e sem aviso.

Medido no corpus atual: **75% do peso** dos códigos aprendidos cairia nesse
buraco. Duas famílias:

1. **Contas enriquecidas** (``1.90.*``, ``2.90.*``, ``3.90.*``) — criadas pelos
   Planos E/G para cobrir linhas de balancete ausentes na ECF. Não existem no
   plano oficial, logo nenhum prefixo do template as captura.
2. **Bloco paralelo ``3.11.*``** — a ECF tem a DRE duplicada em ``3.01.*``
   (Lucro Real) e ``3.11.*`` (Presumido/Arbitrado), com descrições idênticas. O
   template usa ``3.01.*``; tudo que casou em ``3.11.*`` se perderia.

## Estratégia de projeção (em ordem)

1. ``nao_projetar`` — agrupadores de topo (``1.01``, ``3.01.01``…) não viram
   linha do template: sinalizam matching genérico demais e vão para revisão.
2. **Mapa explícito** (``data/template_projection.json``) — enriquecidas e
   casos especiais da ECF que o template não lista.
3. **Normalização ``3.11.* -> 3.01.*``** — blocos paralelos.
4. **Direto** — o código já é capturado por algum prefixo.
5. **Subida na hierarquia** — remove o último segmento até achar um prefixo.

Os prefixos são lidos **do próprio arquivo do template** em runtime; nunca
duplicados em código. Editar o ``.xlsx`` no Excel basta.
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from pathlib import Path

CODE_RE = re.compile(r"^\d+(\.\d+)*$")

# Colunas/abas do Template GT de onde os prefixos são lidos.
_TEMPLATE_SHEETS = ("BP_GT", "DRE_GT")
_LABEL_COLUMN = 2  # coluna B — descrição da linha (define o sinal)
_CODE_COLUMN = 3  # coluna C — código ECF consumido pelo SUMIFS


def _repo_root() -> Path:
    return Path(__file__).resolve().parent.parent.parent.parent


@dataclass(frozen=True)
class LinhaTemplate:
    """Uma linha de BP_GT/DRE_GT e os prefixos que o SUMIFS dela captura.

    Existe para poder perguntar *quantas linhas* somam um dado código. Só o
    conjunto de prefixos não responde isso: duas linhas diferentes poderiam
    declarar prefixos aninhados e somar o mesmo valor duas vezes — que é
    exatamente o risco do curinga ``&"*"`` (ver ``REVISAO_QUALIDADE.md`` §13).
    """

    aba: str
    linha: int
    rotulo: str
    prefixos: tuple[str, ...]

    def captura(self, codigo: str) -> bool:
        """Mesma semântica do curinga do Excel: prefixo de string."""
        return any(codigo.startswith(p) for p in self.prefixos)


@dataclass
class ProjectionResult:
    """Resultado da projeção de um código."""

    codigo_original: str
    codigo_template: str | None
    metodo: str
    motivo: str = ""

    @property
    def ok(self) -> bool:
        return self.codigo_template is not None


@dataclass
class ProjectionStats:
    """Contagem por método, para auditoria do build."""

    total: int = 0
    por_metodo: dict[str, int] = field(default_factory=dict)
    nao_projetados: list[str] = field(default_factory=list)

    def registrar(self, r: ProjectionResult) -> None:
        self.total += 1
        self.por_metodo[r.metodo] = self.por_metodo.get(r.metodo, 0) + 1
        if not r.ok:
            self.nao_projetados.append(r.codigo_original)


class TemplateProjector:
    """Projeta códigos internos nos códigos consumidos pelo Template GT."""

    def __init__(
        self,
        template_path: str | Path | None = None,
        projection_path: str | Path | None = None,
    ):
        root = _repo_root()
        self.template_path = Path(template_path) if template_path else (
            root / "templates" / "Template_GT_BP_Padrao_v3.xlsx"
        )
        projection_path = Path(projection_path) if projection_path else (
            root / "data" / "template_projection.json"
        )

        self.prefixes, self.signs, self.linhas = self._load_prefixes_and_signs(
            self.template_path
        )
        cfg = self._load_projection(projection_path)
        self.mapa: dict[str, str] = cfg.get("mapa", {})
        self.nao_projetar: frozenset[str] = frozenset(
            cfg.get("nao_projetar", {}).get("codigos", [])
        )
        self.stats = ProjectionStats()

    # ------------------------------------------------------------------
    @staticmethod
    def _load_prefixes_and_signs(
        template_path: Path,
    ) -> tuple[tuple[str, ...], dict[str, int], tuple[LinhaTemplate, ...]]:
        """
        Lê os prefixos (coluna C) e o sinal esperado (coluna B) do template.

        O sinal vem do rótulo da própria linha — ``(-) Despesas com pessoal``
        é negativa, ``(+) Receitas financeiras`` positiva —, de modo que
        editar o template no Excel mantém código e convenção em sincronia.

        Linhas sem prefixo (todo o Balanço e as receitas brutas) ficam
        positivas: o check ``Ativo = Passivo + PL`` do template exige que
        ambos os lados sejam positivos.
        """
        if not template_path.exists():
            raise FileNotFoundError(
                f"Template GT não encontrado: {template_path}. "
                "Coloque o .xlsx em templates/ (ver docs/TEMPLATE_GT_BP.md)."
            )
        from openpyxl import load_workbook

        wb = load_workbook(template_path, read_only=True)
        found: set[str] = set()
        signs: dict[str, int] = {}
        linhas: list[LinhaTemplate] = []
        try:
            for sheet in _TEMPLATE_SHEETS:
                if sheet not in wb.sheetnames:
                    continue
                ws = wb[sheet]
                for numero, row in enumerate(
                    ws.iter_rows(
                        min_col=_LABEL_COLUMN, max_col=_CODE_COLUMN, values_only=True
                    ),
                    start=1,
                ):
                    label, codigo = row[0], row[-1]
                    if not codigo or "Código" in str(codigo):
                        continue
                    label = str(label or "")
                    sign = -1 if label.lstrip().startswith("(-)") else 1
                    # Uma linha pode declarar vários códigos unidos por "|".
                    da_linha: list[str] = []
                    for part in str(codigo).split("|"):
                        part = part.strip()
                        if CODE_RE.fullmatch(part):
                            found.add(part)
                            signs[part] = sign
                            da_linha.append(part)
                    if da_linha:
                        linhas.append(
                            LinhaTemplate(sheet, numero, label.strip(), tuple(da_linha))
                        )
        finally:
            wb.close()
        # Mais longos primeiro: casamento por prefixo deve preferir o específico.
        return (
            tuple(sorted(found, key=lambda c: (-len(c), c))),
            signs,
            tuple(linhas),
        )

    def sign_for(self, template_code: str) -> int:
        """
        Sinal (+1/-1) que o template espera para um código já projetado.

        Deduções, custos, despesas, depreciação e IRPJ/CSLL entram negativos
        porque as fórmulas do DRE somam (``EBITDA = D22+SUM(D25:D30)``).
        Ativo e Passivo entram positivos porque o check do Balanço subtrai
        (``ROUND(D26-D52,2)=0``).
        """
        if template_code in self.signs:
            return self.signs[template_code]
        for p in self.prefixes:  # já ordenados do mais específico ao mais curto
            if template_code.startswith(p):
                return self.signs.get(p, 1)
        return 1

    @staticmethod
    def _load_projection(path: Path) -> dict:
        if not path.exists():
            return {}
        with open(path, encoding="utf-8") as f:
            return json.load(f)

    # ------------------------------------------------------------------
    def is_captured(self, codigo: str) -> bool:
        """O template soma este código em alguma linha?"""
        return any(codigo.startswith(p) for p in self.prefixes)

    def linhas_que_capturam(self, codigo: str) -> tuple[LinhaTemplate, ...]:
        """
        Quais linhas do template somam este código.

        O contrato da entrega é **exatamente uma**:

        - **zero** → o valor é escrito em ``_dados_padronizados`` e não aparece
          em lugar nenhum. Some sem aviso: as contagens do Sumário o tratam
          como "conta tratada", e o balanço não fecha por um valor que nenhum
          relatório nomeia.
        - **duas ou mais** → o mesmo valor é somado em duas linhas. É a dupla
          contagem que o curinga ``&"*"`` permite quando os prefixos de duas
          linhas são aninhados.
        """
        return tuple(linha for linha in self.linhas if linha.captura(codigo))

    def prefixos_aninhados(self) -> tuple[tuple[str, str], ...]:
        """
        Pares ``(curto, longo)`` em que um prefixo captura o outro.

        Enquanto isto for vazio, o curinga ``&"*"`` não pode gerar dupla
        contagem: nenhum código cai em duas linhas. É a propriedade que torna
        o template seguro — e que ninguém garantia até aqui. Uma edição no
        Excel que acrescente uma linha "pai" ao lado das filhas quebra isso
        em silêncio, e é por isso que existe teste em cima desta função.
        """
        nus = sorted(set(self.prefixes))
        return tuple(
            (a, b) for a in nus for b in nus if a != b and b.startswith(a)
        )

    def project(self, codigo: str | None) -> ProjectionResult:
        """
        Projeta um código interno no código que o template consome.

        Returns:
            ProjectionResult — ``codigo_template`` é None quando não há projeção
            possível (a conta deve ir para "Contas Não Identificadas").
        """
        if not codigo or not CODE_RE.fullmatch(str(codigo)):
            r = ProjectionResult(str(codigo or ""), None, "invalido", "código fora do formato ECF")
            self.stats.registrar(r)
            return r

        codigo = str(codigo)

        # 1. Agrupador de topo: matching genérico demais para virar linha.
        if codigo in self.nao_projetar:
            r = ProjectionResult(
                codigo, None, "generico_demais",
                "código é agrupador de topo; requer conta analítica",
            )
            self.stats.registrar(r)
            return r

        # 2. Mapa explícito (enriquecidas e casos especiais).
        alvo = self.mapa.get(codigo)
        if alvo:
            r = ProjectionResult(codigo, alvo, "mapa_explicito")
            self.stats.registrar(r)
            return r

        # 3. Normaliza o bloco paralelo 3.11.* -> 3.01.* ANTES das etapas
        #    seguintes: a ECF duplica a DRE em Lucro Real (3.01) e Presumido
        #    (3.11) com descrições idênticas, e o template só usa 3.01. A
        #    normalização precede o casamento direto E a subida hierárquica,
        #    senão um 3.11 analítico sem linha própria nunca acharia ancestral.
        alvo_base, metodo_base = codigo, "direto"
        if codigo.startswith("3.11."):
            alvo_base = "3.01." + codigo[5:]
            metodo_base = "normalizado_3_11"
            # o mapa explícito pode cobrir a forma normalizada
            alvo = self.mapa.get(alvo_base)
            if alvo:
                r = ProjectionResult(codigo, alvo, "normalizado_3_11+mapa")
                self.stats.registrar(r)
                return r

        # 4. Já capturado diretamente (na forma normalizada, quando aplicável).
        if self.is_captured(alvo_base):
            r = ProjectionResult(codigo, alvo_base, metodo_base)
            self.stats.registrar(r)
            return r

        # 5. Sobe na hierarquia até achar um prefixo consumido.
        partes = alvo_base.split(".")
        for corte in range(len(partes) - 1, 0, -1):
            ancestral = ".".join(partes[:corte])
            if ancestral in self.nao_projetar:
                break
            if self.is_captured(ancestral):
                metodo = "ancestral" if metodo_base == "direto" else "normalizado_3_11+ancestral"
                r = ProjectionResult(codigo, ancestral, metodo)
                self.stats.registrar(r)
                return r

        r = ProjectionResult(
            codigo, None, "sem_projecao",
            "nenhuma linha do template captura este código",
        )
        self.stats.registrar(r)
        return r

    def reset_stats(self) -> None:
        self.stats = ProjectionStats()
