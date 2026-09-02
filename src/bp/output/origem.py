"""
Aba "Original" — a cópia do balancete de origem dentro da entrega.

Por que isso existe
-------------------

O arquivo entregue ao cliente é o resultado de uma cadeia de decisões
(parser -> matcher -> projeção -> SUMIFS). Quando alguém pergunta *"de onde
saiu esse número?"*, a resposta precisa caber no próprio arquivo — sem
depender de encontrar, meses depois, o balancete que o originou numa pasta
qualquer da rede.

Esta aba transcreve o conteúdo do arquivo de origem **como ele foi lido**,
linha a linha, mais um cabeçalho de proveniência com o SHA-256 do arquivo.
O hash é o que transforma "parece o mesmo arquivo" em prova: quem tiver o
original pode conferir o digest e saber que é exatamente aquele.

Limite honesto
--------------

XLSX não guarda anexos binários por meio do openpyxl, então isto é uma
**transcrição fiel do conteúdo**, não o arquivo original embutido byte a
byte. Formatação, fórmulas e imagens do balancete de origem não vêm junto —
os valores e o texto, sim. O SHA-256 identifica o arquivo de onde vieram.
"""

from __future__ import annotations

import hashlib
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from ..parsers.conversao import convertido_para_xlsx

#: Teto de linhas transcritas por arquivo. Excel comporta ~1.048.576 linhas no
#: total; balancetes reais têm centenas. O teto existe para que um arquivo
#: patológico não inviabilize a entrega inteira — quando ele é atingido, a aba
#: diz explicitamente que truncou (silêncio aqui seria pior que o truncamento).
MAX_LINHAS = 50_000

#: Teto de colunas. Mesma lógica.
MAX_COLUNAS = 64

#: Extensões lidas como texto puro (uma linha do arquivo = uma linha da aba).
_TEXTUAIS = {".csv", ".txt"}

#: Ordem de tentativa de encoding para os formatos textuais. Balancete
#: brasileiro exportado de sistema legado costuma ser latin-1; UTF-8 vem
#: primeiro porque, quando o arquivo é UTF-8, latin-1 "funciona" mas corrompe
#: os acentos silenciosamente.
_ENCODINGS = ("utf-8-sig", "utf-8", "latin-1")


@dataclass
class Origem:
    """Conteúdo transcrito de um arquivo de origem, com sua proveniência."""

    path: Path
    sha256: str
    tamanho_bytes: int
    modificado_em: str
    linhas: list[list[Any]] = field(default_factory=list)
    truncado: bool = False
    erro: str = ""
    #: Caminho de onde o conteúdo veio de fato. Difere de ``path`` quando o
    #: pipeline lê outro arquivo no lugar do pedido — ver ``_conteudo_de``.
    lido_de: Path | None = None
    #: Como o conteúdo foi obtido, em uma frase. Vai para a aba.
    procedencia: str = ""

    @property
    def legivel(self) -> bool:
        return bool(self.linhas)


def sha256_de(path: Path, *, bloco: int = 1 << 20) -> str:
    """SHA-256 do arquivo, lido em blocos (não carrega o arquivo na memória)."""
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for pedaco in iter(lambda: fh.read(bloco), b""):
            h.update(pedaco)
    return h.hexdigest()


def ler_origem(path: str | Path) -> Origem:
    """
    Transcreve o arquivo de origem.

    Nunca levanta exceção por causa do conteúdo: um arquivo ilegível vira uma
    ``Origem`` com ``erro`` preenchido e sem linhas. A entrega não pode falhar
    porque a aba de rastreio não conseguiu ser montada — mas ela também não
    pode mentir dizendo que está tudo bem, daí o campo ``erro`` aparecer na aba.
    """
    path = Path(path)
    stat = path.stat()
    origem = Origem(
        path=path,
        sha256=sha256_de(path),
        tamanho_bytes=stat.st_size,
        modificado_em=datetime.fromtimestamp(stat.st_mtime).strftime("%d/%m/%Y %H:%M"),
    )

    try:
        linhas, lido_de, procedencia = _conteudo_de(path)
    except Exception as exc:
        origem.erro = f"{type(exc).__name__}: {exc}"
        return origem

    origem.lido_de = lido_de
    origem.procedencia = procedencia
    if len(linhas) > MAX_LINHAS:
        origem.truncado = True
        linhas = linhas[:MAX_LINHAS]
    origem.linhas = [linha[:MAX_COLUNAS] for linha in linhas]
    if not origem.linhas:
        origem.erro = "arquivo lido, mas sem conteúdo transcritível"
    return origem


def _conteudo_de(path: Path) -> tuple[list[list[Any]], Path, str]:
    """
    Transcreve o arquivo **pela mesma rota que o parser usa para lê-lo**.

    Isto não é detalhe: ``XlsParser.read()`` prefere um ``.xlsx`` irmão de
    mesmo nome quando ele existe, sem avisar ninguém. Se a aba de rastreio
    transcrevesse o ``.xls`` e o pipeline tivesse lido o ``.xlsx``, a cópia
    "para facilitar o rastreio" estaria mostrando um arquivo que não gerou
    número nenhum — o pior tipo de rastreio, o que dá confiança errada.

    Devolve ``(linhas, arquivo realmente lido, frase de procedência)``.
    """
    suffix = path.suffix.lower()

    if suffix in _TEXTUAIS:
        return _ler_texto(path), path, "lido como texto, linha a linha"

    if suffix == ".pdf":
        return _ler_pdf(path), path, "texto extraído do PDF, página a página"

    if suffix == ".xls":
        irmao = path.with_suffix(".xlsx")
        if irmao.exists():
            return (
                _ler_tabular(irmao),
                irmao,
                f"ATENÇÃO: o conteúdo veio de {irmao.name}, não do .xls — o "
                f"parser prefere o .xlsx de mesmo nome quando ele existe",
            )
        with convertido_para_xlsx(path) as convertido:
            if convertido is not None:
                return (
                    _ler_tabular(convertido),
                    path,
                    "convertido de .xls para .xlsx pelo LibreOffice",
                )

    return _ler_tabular(path), path, "lido diretamente da planilha"


def _ler_texto(path: Path) -> list[list[Any]]:
    """CSV/TXT: uma linha do arquivo por linha da aba, verbatim.

    Não separa por delimitador de propósito. O objetivo é rastreio: o analista
    precisa ver a linha como ela está no arquivo, inclusive o preâmbulo de
    cabeçalho que o parser descartou.
    """
    for enc in _ENCODINGS:
        try:
            texto = path.read_text(encoding=enc)
        except (UnicodeDecodeError, LookupError):
            continue
        return [[linha.rstrip("\r")] for linha in texto.split("\n")]
    return [[path.read_bytes().decode("latin-1", errors="replace")]]


def _ler_pdf(path: Path) -> list[list[Any]]:
    """PDF: uma linha de texto da página por linha da aba, na ordem de leitura."""
    import pdfplumber

    linhas: list[list[Any]] = []
    with pdfplumber.open(path) as pdf:
        for numero, pagina in enumerate(pdf.pages, start=1):
            linhas.append([f"═══ página {numero} ═══"])
            for linha in (pagina.extract_text() or "").split("\n"):
                linhas.append([linha])
    return linhas


def _ler_tabular(path: Path) -> list[list[Any]]:
    """Planilha: todas as abas, valores como estão, sem inferir cabeçalho.

    ``header=None`` é o ponto: as funções de leitura do pipeline detectam
    cabeçalho e descartam o preâmbulo. Aqui não se descarta nada — quem lê a
    aba precisa ver o arquivo, não a interpretação dele.
    """
    import pandas as pd

    abas = pd.read_excel(path, sheet_name=None, header=None, dtype=object)
    linhas: list[list[Any]] = []
    multiplas = len(abas) > 1
    for nome, df in abas.items():
        if multiplas:
            linhas.append([f"═══ aba do arquivo de origem: {nome} ═══"])
        for registro in df.itertuples(index=False, name=None):
            linhas.append([_celula(v) for v in registro])
    return linhas


def _celula(valor: Any) -> Any:
    """Converte para algo que o openpyxl aceita, preservando números."""
    if valor is None:
        return None
    if isinstance(valor, (int, float, str, bool, datetime)):
        # NaN do pandas é float e vira célula vazia (não a string "nan").
        if isinstance(valor, float) and valor != valor:
            return None
        return valor
    return str(valor)


def nome_da_aba(ano: int | None, usados: set[str]) -> str:
    """
    Nome de aba válido no Excel (<=31 chars, sem ``[]:*?/\\``) e único.

    Com um exercício só, "Balancete Original" já diz tudo. Com série
    histórica, o ano precisa aparecer — senão o analista não sabe qual das
    cópias é a de 2023.
    """
    base = "Balancete Original" if ano is None else f"Original {ano}"
    nome = base[:31]
    sufixo = 2
    while nome in usados:
        cauda = f" ({sufixo})"
        nome = base[: 31 - len(cauda)] + cauda
        sufixo += 1
    usados.add(nome)
    return nome


def escrever_aba_origem(wb, nome: str, origem: Origem) -> None:
    """Cria a aba com o cabeçalho de proveniência e a transcrição."""
    from openpyxl.styles import Font

    if nome in wb.sheetnames:
        del wb[nome]
    ws = wb.create_sheet(nome)

    cabecalho: list[tuple[str, Any]] = [
        ("CÓPIA DO BALANCETE DE ORIGEM", ""),
        (
            "Transcrição fiel do conteúdo lido. Serve para rastrear qualquer "
            "número da entrega até a linha que o originou.",
            "",
        ),
        ("", ""),
        ("Arquivo:", origem.path.name),
        ("Caminho na geração:", str(origem.path)),
        ("Tamanho (bytes):", origem.tamanho_bytes),
        ("Modificado em:", origem.modificado_em),
        ("SHA-256:", origem.sha256),
    ]
    if origem.truncado:
        cabecalho.append(
            ("ATENÇÃO:", f"transcrição truncada nas primeiras {MAX_LINHAS} linhas")
        )
    if origem.erro:
        cabecalho.append(("NÃO FOI POSSÍVEL TRANSCREVER:", origem.erro))

    for i, (rotulo, valor) in enumerate(cabecalho, start=1):
        ws.cell(row=i, column=1, value=rotulo)
        ws.cell(row=i, column=2, value=valor)
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.cell(row=2, column=1).font = Font(italic=True, size=9)
    if origem.erro:
        ws.cell(row=len(cabecalho), column=1).font = Font(bold=True, color="C00000")

    primeira = len(cabecalho) + 2
    ws.cell(row=primeira, column=1, value="CONTEÚDO DO ARQUIVO").font = Font(bold=True)

    for offset, linha in enumerate(origem.linhas, start=primeira + 1):
        for col, valor in enumerate(linha, start=1):
            if valor is not None and valor != "":
                ws.cell(row=offset, column=col, value=valor)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 52
    ws.freeze_panes = ws.cell(row=primeira + 1, column=1).coordinate
