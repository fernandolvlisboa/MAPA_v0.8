"""
Gera o arquivo final no Template GT — a entrega para o cliente.

Arquitetura híbrida (ver ``docs/TEMPLATE_GT_BP.md``):

- **Template** = camada de apresentação. Formatação, fórmulas SUMIFS, blocos de
  análise e identidade visual GT já estão prontos no ``.xlsx``. O código
  **nunca** os recria nem os reformata.
- **Python** = camada de dados. Só alimenta a aba oculta
  ``_dados_padronizados``; as fórmulas do template fazem a agregação sozinhas.

Fluxo::

    balancete do cliente -> parser -> matcher -> projeção p/ código do template
                                                     |
                                       _dados_padronizados (oculta)
                                                     |
                                      SUMIFS do template agrega
                                                     |
                                            BP_GT / DRE_GT

Abas do resultado:

===========================  ==========  ======================================
Aba                          Origem      Público
===========================  ==========  ======================================
``Sumário``                  gerada      **interno** — qualidade do processamento
``BP_GT`` / ``DRE_GT``       template    **cliente** — a entrega
``Contas Tratadas``          gerada      interno — auditoria do de-para
``Contas Não Identificadas`` gerada      interno — fila de revisão do analista
``_instrucoes``              template    documentação embutida
``_dados_padronizados``      alimentada  oculta — insumo do SUMIFS
===========================  ==========  ======================================
"""

from __future__ import annotations

import re
import shutil
from collections import defaultdict
from collections.abc import Callable, Sequence
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from ..matchers import ContaMatcher
from ..parsers.dispatcher import ParseyCaller
from ..utils.codigo import classe_from_codigo
from ..utils.natureza import mapear_natureza, resultado_do_periodo, totais_por_natureza
from ..utils.prazo import mapear_prazo, prazo_do_codigo_referencial
from ..utils.synonyms import is_garbage_description
from ..validators.entrega import RelatorioEntrega, conferir_dre, conferir_totais
from ..validators.hierarquia import conferir_hierarquia, selecionar_para_projecao
from .origem import Origem, escrever_aba_origem, ler_origem, nome_da_aba
from .template_map import TemplateProjector

CODE_RE = re.compile(r"^\d+(\.\d+)*$")
_REF_RE = re.compile(r"_dados_padronizados!([A-Z]+):")

#: Anos que o template traz de fábrica. Não são fixos: ``_aplicar_anos()``
#: reescreve os rótulos conforme os anos realmente fornecidos (ex.: 2018-2021).
ANOS_TEMPLATE = (2021, 2022, 2023, 2024, 2025)

#: Linha dos rótulos de ano em BP_GT/DRE_GT.
_LINHA_ANOS = 7

_ABAS_GERADAS = ("Sumário", "Contas Tratadas", "Contas Não Identificadas")

#: Prefixos das abas de cópia do original. Como o nome carrega o exercício,
#: ``_ordenar_abas`` as posiciona por prefixo, não por nome exato.
_PREFIXOS_ORIGEM = ("Balancete Original", "Original ")
#: Ordem final das abas (esquerda -> direita), conforme docs/TEMPLATE_GT_BP.md.
_ORDEM_ABAS = (
    "Sumário",
    "BP_GT",
    "DRE_GT",
    "Contas Tratadas",
    "Contas Não Identificadas",
    "_instrucoes",
    "_dados_padronizados",
)


@dataclass
class FonteBalancete:
    """Um balancete e o exercício a que ele se refere.

    Um balancete brasileiro cobre **um** período (as colunas típicas são
    ``Saldo Anterior | Débito | Crédito | Saldo Atual`` — movimentação do mesmo
    exercício, não anos distintos). Série histórica se monta com **um período
    por fonte**, e é isso que esta estrutura representa.

    O período pode vir num arquivo próprio **ou numa aba** de uma pasta de
    trabalho: planilha de cliente traz "Balancetes 2020" … "Balancetes 2026"
    lado a lado, e cada aba é um exercício. ``aba`` nomeia qual usar; com
    ``None``, a escolha automática decide (ver ``ParseyCaller``).
    """

    path: str | Path
    ano: int
    escala: float = 1000.0  # divisor para milhares
    #: Aba da planilha, quando o arquivo traz mais de um exercício.
    aba: str | None = None

    def __post_init__(self) -> None:
        self.path = Path(self.path)


@dataclass
class ContaSemDestino:
    """Uma conta cujo valor não chegou à entrega, com o quanto ela pesa."""

    codigo: str
    descricao: str
    valor: float
    motivo: str


@dataclass
class Reconciliacao:
    """
    Prova aritmética de por que a entrega não fecha.

    Não basta dizer "não fecha": é preciso mostrar que a diferença **é
    exatamente** a soma das contas que ficaram sem destino. Quando bate, o
    analista sabe que não há nada escondido — só aquelas N contas — e pode
    decidir com segurança. Quando NÃO bate, há um problema estrutural
    (dupla contagem, conta perdida) que nenhuma revisão manual resolveria.

    A identidade vem de duas coisas já garantidas: a origem fecha
    (Ativo + Passivo + Resultado = 0) e a cobertura é completa (emitido +
    sem destino = origem, por classe). Logo ``emitido == -sem_destino``.
    """

    desequilibrio: float
    """Soma do que foi emitido, nos sinais da origem. Zero = fecha."""

    soma_sem_destino: float
    """Soma das contas que não chegaram à entrega."""

    contas: list[ContaSemDestino] = field(default_factory=list)

    @property
    def residuo(self) -> float:
        """O que sobra depois de explicar a diferença. Deve ser zero."""
        return self.desequilibrio + self.soma_sem_destino

    @property
    def fecha(self) -> bool:
        return abs(self.desequilibrio) <= 0.01

    @property
    def explicada(self) -> bool:
        """A diferença é integralmente explicada pelas contas listadas."""
        return abs(self.residuo) <= 0.01

    def mensagem(self) -> str:
        if self.fecha:
            return "Balanço fecha: nenhuma conta ficou sem destino."
        n = len(self.contas)
        base = (
            f"Não fecha por {abs(self.desequilibrio):,.2f}. "
            f"Há {n} conta(s) sem destino no template somando "
            f"{abs(self.soma_sem_destino):,.2f}"
        )
        if self.explicada:
            return (
                f"{base} — exatamente a diferença. As {n} conta(s) explicam "
                f"100% do desequilíbrio; não há mais nada faltando."
            )
        return (
            f"{base}, mas sobram {abs(self.residuo):,.2f} sem explicação. "
            f"ATENÇÃO: há conta contada duas vezes ou perdida no caminho."
        )


@dataclass
class BuildResult:
    """Resultado da geração, para log e testes."""

    output_path: Path
    anos: tuple[int, ...] = ()
    linhas_escritas: int = 0
    contas_lidas: int = 0
    contas_tratadas: int = 0
    contas_nao_identificadas: int = 0
    #: Contas cujo campo de saldo existia mas não pôde ser convertido. É
    #: diferente de conta zerada, e é a causa que o aviso de desequilíbrio
    #: precisa nomear para não mandar o analista ao lugar errado.
    saldos_ilegiveis: int = 0
    #: Contas com nome próprio que subiram para um agrupador mapeado. Não são
    #: perda: o valor delas está no total do agrupador.
    contas_absorvidas: int = 0
    #: Conferência aritmética do balancete de origem (soma dos filhos vs pai).
    hierarquia: Any = None
    #: Valor (em moeda de origem) das contas sem destino no template. É o
    #: montante EXATO pelo qual o balanço deixa de fechar.
    valor_nao_coberto: float = 0.0
    #: Soma, na moeda e nos sinais DA ORIGEM, do que foi emitido — por classe.
    #: Os totais do lado do template passam por `abs() * sign_for()`, que
    #: descarta o sinal lido, então não servem para conferir nada. Estes servem.
    emitido_por_classe: dict[str, float] = field(default_factory=dict)
    nao_coberto_por_classe: dict[str, float] = field(default_factory=dict)
    #: Mesma abertura, mas por natureza de resultado (RECEITA/DESPESA). A
    #: abertura por classe não serve para conferir a DRE: receita e despesa
    #: são ambas RESULTADO e têm sinais opostos, então a soma delas não diz
    #: nada sobre o lucro.
    emitido_por_natureza: dict[str, float] = field(default_factory=dict)
    nao_coberto_por_natureza: dict[str, float] = field(default_factory=dict)
    #: Contas sem destino, para nomear na reconciliação.
    contas_sem_destino: list[ContaSemDestino] = field(default_factory=list)
    total_ativo: float = 0.0
    total_passivo: float = 0.0
    avisos: list[str] = field(default_factory=list)
    #: Por ano: (contas_lidas, tratadas, nao_identificadas)
    por_ano: dict[int, tuple[int, int, int]] = field(default_factory=dict)
    #: Proveniência de cada balancete de origem transcrito na entrega.
    origens: list[Origem] = field(default_factory=list)
    #: Linhas escritas que NENHUMA linha do template soma. O valor está na aba
    #: de dados e não aparece na entrega — some sem aviso.
    linhas_sem_captura: list[tuple[str, float]] = field(default_factory=list)
    #: Linhas escritas que DUAS OU MAIS linhas do template somam. É a dupla
    #: contagem que o curinga ``&"*"`` permite.
    linhas_capturadas_duas_vezes: list[tuple[str, float, int]] = field(
        default_factory=list
    )
    #: Resultado do período levado ao PL porque o balancete é "aberto" — as
    #: contas de resultado ainda têm saldo e o lucro não foi transferido.
    #: Zero quando o balancete já vem encerrado.
    resultado_transferido: float = 0.0
    #: A conferência que importa: o total entregue bate com o da origem?
    #: Avalia as fórmulas do template — é a única medida que enxerga o número
    #: que o cliente lê. Ver ``validators/entrega.py``.
    entrega: RelatorioEntrega = field(default_factory=RelatorioEntrega)
    #: A mesma conferência do outro lado: o lucro líquido entregue é o
    #: resultado da origem? O Balanço pode fechar com a DRE inteira errada.
    dre: RelatorioEntrega = field(default_factory=RelatorioEntrega)
    #: Resultado do exercício lido na origem, na moeda da entrega.
    resultado_da_origem: float = 0.0

    @property
    def captura_integra(self) -> bool:
        """Todo valor escrito aparece na entrega, e uma vez só."""
        return not self.linhas_sem_captura and not self.linhas_capturadas_duas_vezes

    @property
    def valor_sem_captura(self) -> float:
        return sum(v for _, v in self.linhas_sem_captura)

    @property
    def valor_contado_duas_vezes(self) -> float:
        return sum(v * (n - 1) for _, v, n in self.linhas_capturadas_duas_vezes)

    @property
    def match_rate(self) -> float:
        base = self.contas_tratadas + self.contas_nao_identificadas
        return self.contas_tratadas / base if base else 0.0

    @property
    def reconciliacao(self) -> Reconciliacao:
        """Prova de por que não fecha — e de que nada mais está faltando."""
        return Reconciliacao(
            desequilibrio=sum(self.emitido_por_classe.values()),
            soma_sem_destino=self.valor_nao_coberto,
            contas=sorted(
                self.contas_sem_destino, key=lambda c: abs(c.valor), reverse=True
            ),
        )

    @property
    def cobertura_de_valor(self) -> float:
        """
        Fração do valor da origem que chegou à entrega.

        É a métrica comparável entre balancetes — e a que interessa ao
        negócio. Contar *contas* não serve: um código emitido pode cobrir
        várias contas homônimas, e uma folha absorvida não é conta perdida.
        O que se mede é dinheiro.
        """
        if self.hierarquia is None or not self.hierarquia.tem_hierarquia:
            return 1.0
        origem = sum(abs(v) for v in self.hierarquia.totais_por_classe.values())
        if origem == 0:
            return 1.0
        return 1.0 - abs(self.valor_nao_coberto) / origem

    @property
    def cobertura_completa(self) -> bool:
        """
        Nada evaporou: por classe, emitido + não coberto == total da origem.

        É a invariante mais forte do pipeline e a única que vale sempre, mesmo
        quando o template distorce os sinais. Se ela quebra, há conta sendo
        contada duas vezes ou sumindo — e nenhum total a jusante é confiável.
        """
        if self.hierarquia is None or not self.hierarquia.tem_hierarquia:
            return True  # sem árvore não há o que conferir
        for classe, origem in self.hierarquia.totais_por_classe.items():
            reconstruido = self.emitido_por_classe.get(
                classe, 0.0
            ) + self.nao_coberto_por_classe.get(classe, 0.0)
            if abs(reconstruido - origem) > max(0.01, abs(origem) * 1e-6):
                return False
        return True

    @property
    def balanco_confere(self) -> bool:
        """
        A entrega fecha?

        Exige as três coisas: a origem ser consistente, nada ter evaporado no
        caminho, e todo valor ter destino no template. Antes comparava
        ``total_ativo`` com ``total_passivo`` — dois números já passados por
        ``abs()``, incapazes de detectar um sinal perdido ou uma conta a menos.
        """
        # A conferência ponta a ponta manda: ela lê o número que o cliente lê.
        # As outras são proxy, e proxy já ficou verde com o total errado.
        if self.entrega.conferivel and not self.entrega.confere:
            return False
        if self.dre.conferivel and not self.dre.confere:
            return False
        if not self.cobertura_completa:
            return False
        if abs(self.valor_nao_coberto) > 0.01:
            return False
        if self.hierarquia is not None and self.hierarquia.tem_hierarquia:
            return self.hierarquia.rollup_integro and self.hierarquia.equacao_fecha
        base = max(abs(self.total_ativo), abs(self.total_passivo), 1.0)
        return abs(self.total_ativo - self.total_passivo) / base < 0.01


def build_gt_output(
    fontes: str | Path | FonteBalancete | Sequence[FonteBalancete],
    output_path: str | Path,
    ano_base: int | None = None,
    nome_cliente: str | None = None,
    data_base: str | None = None,
    template_path: str | Path | None = None,
    plano_path: str | Path | None = None,
    cache_path: str | Path | None = None,
    escala: float = 1000.0,
    on_progress: Callable[[str], None] | None = None,
) -> BuildResult:
    """
    Padroniza um ou mais balancetes e entrega no Template GT.

    Aceita três formas de entrada:

    - **um caminho** + ``ano_base`` — o caso simples::

          build_gt_output("balancete.xlsx", "saida.xlsx", ano_base=2024)

    - **uma lista de** :class:`FonteBalancete` — série histórica, um arquivo por
      exercício (é assim que os balancetes existem no mundo real)::

          build_gt_output([
              FonteBalancete("bal_2022.xlsx", 2022),
              FonteBalancete("bal_2023.xlsx", 2023),
              FonteBalancete("bal_2024.xlsx", 2024),
          ], "saida.xlsx")

    Os anos **não** precisam ser 2021-2025: o template comporta cinco
    exercícios quaisquer e os rótulos são reescritos conforme os anos
    fornecidos (2018-2021 funciona igual).

    Args:
        fontes: caminho único, uma ``FonteBalancete`` ou lista delas.
        output_path: destino do arquivo final.
        ano_base: exercício, obrigatório apenas quando ``fontes`` é um caminho.
        nome_cliente: célula B4 de BP_GT/DRE_GT. Default: nome do 1º arquivo.
        data_base: data-base exibida no Sumário. Default: 31/12 do último ano.
        template_path: default ``templates/Template_GT_BP_Padrao_v3.xlsx``.
        plano_path: default ``data/plano_referencial.json``.
        cache_path: cache de matching. Default **efêmero** — a entrega de um
            cliente não escreve no cache compartilhado do projeto (evita
            gravar como permanente um match casual de nome próprio).
        escala: divisor para milhares, aplicado quando ``fontes`` é um caminho.
        on_progress: callback opcional que recebe uma frase do que está
            acontecendo agora. Existe para a interface gráfica poder mostrar
            andamento honesto numa geração que leva dezenas de segundos; quem
            chama de script simplesmente não passa nada.

    Returns:
        BuildResult com contagens por ano, totais e avisos de validação.

    Raises:
        ValueError: nenhuma fonte, anos repetidos, ou mais anos que o template
            comporta.
    """
    from openpyxl import load_workbook

    def _avisar(mensagem: str) -> None:
        if on_progress:
            on_progress(mensagem)

    lista = _normalizar_fontes(fontes, ano_base, escala)
    output_path = Path(output_path)

    nome_cliente = nome_cliente or Path(lista[0].path).stem
    anos = tuple(sorted(f.ano for f in lista))
    data_base = data_base or f"{anos[-1]}-12-31"

    projector = TemplateProjector(template_path=template_path)
    matcher = _build_matcher(plano_path, cache_path)

    # 1. Copiar o template — o original nunca é tocado.
    tpl = Path(template_path) if template_path else projector.template_path
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(tpl, output_path)

    # 2. Abrir a cópia SEM data_only (senão as fórmulas viram valores fixos).
    wb = load_workbook(output_path)

    # 3. Descobre, pelas PRÓPRIAS fórmulas, qual coluna da aba de dados cada
    #    coluna de ano lê. Assim o código não depende de o template estar
    #    alinhado — e continua correto se ele for reeditado.
    slots = _slots_de_ano(wb)
    if len(anos) > len(slots):
        raise ValueError(
            f"O template comporta {len(slots)} exercícios; foram fornecidos {len(anos)}: {anos}"
        )

    # 4. Reescreve os rótulos de ano conforme os anos realmente fornecidos.
    mapa_ano_coluna, mapa_ano_exibicao = _aplicar_anos(wb, anos, slots)

    # 5. Nome do cliente (única escrita permitida em BP_GT/DRE_GT).
    for aba in ("BP_GT", "DRE_GT"):
        if aba in wb.sheetnames:
            wb[aba]["B4"] = nome_cliente

    # 6. Padroniza cada fonte e alimenta a aba oculta consultada pelo SUMIFS.
    result = BuildResult(output_path=output_path, anos=anos)
    hierarquia_do_ultimo: Any = None
    escala_do_ultimo = 1.0
    todas_linhas: list[dict[str, Any]] = []
    todas_tratadas: list[dict[str, Any]] = []
    todas_nao_ident: list[dict[str, Any]] = []

    for indice, fonte in enumerate(sorted(lista, key=lambda f: f.ano), start=1):
        _avisar(
            f"Lendo e classificando o balancete de {fonte.ano}"
            + (f" ({indice} de {len(lista)})" if len(lista) > 1 else "")
        )
        dados, tratadas, nao_ident, parcial = _padronizar(
            Path(fonte.path), matcher, projector, fonte.escala, fonte.aba
        )
        for d in dados:
            d["ano"] = fonte.ano
        for t in (*tratadas, *nao_ident):
            t["ano"] = fonte.ano
        todas_linhas.extend(dados)
        todas_tratadas.extend(tratadas)
        todas_nao_ident.extend(nao_ident)

        result.contas_lidas += parcial.contas_lidas
        result.contas_tratadas += parcial.contas_tratadas
        result.contas_nao_identificadas += parcial.contas_nao_identificadas
        result.saldos_ilegiveis += parcial.saldos_ilegiveis
        result.contas_absorvidas += parcial.contas_absorvidas
        result.resultado_transferido += parcial.resultado_transferido
        result.valor_nao_coberto += parcial.valor_nao_coberto
        for classe, v in parcial.emitido_por_classe.items():
            result.emitido_por_classe[classe] = (
                result.emitido_por_classe.get(classe, 0.0) + v
            )
        result.contas_sem_destino.extend(parcial.contas_sem_destino)
        for classe, v in parcial.nao_coberto_por_classe.items():
            result.nao_coberto_por_classe[classe] = (
                result.nao_coberto_por_classe.get(classe, 0.0) + v
            )
        for nat, v in parcial.emitido_por_natureza.items():
            result.emitido_por_natureza[nat] = (
                result.emitido_por_natureza.get(nat, 0.0) + v
            )
        for nat, v in parcial.nao_coberto_por_natureza.items():
            result.nao_coberto_por_natureza[nat] = (
                result.nao_coberto_por_natureza.get(nat, 0.0) + v
            )
        if result.hierarquia is None:
            result.hierarquia = parcial.hierarquia
        result.por_ano[fonte.ano] = (
            parcial.contas_lidas,
            parcial.contas_tratadas,
            parcial.contas_nao_identificadas,
        )
        # Balanço é conferido no exercício mais recente.
        if fonte.ano == anos[-1]:
            result.total_ativo = parcial.total_ativo
            result.total_passivo = parcial.total_passivo
            hierarquia_do_ultimo = parcial.hierarquia
            escala_do_ultimo = fonte.escala
            result.resultado_da_origem = parcial.resultado_da_origem
            result.nao_coberto_por_natureza = dict(parcial.nao_coberto_por_natureza)
        result.avisos.extend(f"[{fonte.ano}] {a}" for a in parcial.avisos)

    result.linhas_escritas = len(todas_linhas)

    # Conferência da captura: cada linha escrita tem que ser somada por
    # exatamente uma linha do template. Ver `_conferir_captura`.
    _conferir_captura(todas_linhas, projector, result)

    _avisar("Preenchendo o Template GT")
    _escrever_dados_padronizados(wb, todas_linhas, mapa_ano_coluna)

    # A conferência ponta a ponta: avalia as fórmulas do template e compara o
    # ATIVO TOTAL entregue com o "ATIVO" declarado no balancete. É a única
    # medida que enxerga o número que o cliente lê; todas as outras são proxy.
    #  Os totais da origem vêm na moeda do balancete; a entrega está na escala
    #  do template (milhares). Comparar sem converter daria "não bate" sempre.
    totais_origem = {
        classe: total / escala_do_ultimo
        for classe, total in (
            getattr(hierarquia_do_ultimo, "totais_por_classe", {}) or {}
        ).items()
    }
    result.entrega = conferir_totais(
        wb,
        totais_origem,
        coluna=_letra_da_coluna(mapa_ano_exibicao.get(anos[-1], 4)),
        resultado_transferido=result.resultado_transferido,
    )
    if result.entrega.conferivel and not result.entrega.confere:
        for divergente in result.entrega.divergentes:
            result.avisos.append(
                f"TOTAL DA ENTREGA NÃO BATE COM A ORIGEM — {divergente}. "
                f"O número que o cliente lê está errado; não use esta saída."
            )
    elif not result.entrega.conferivel:
        # A entrega saiu SEM ser conferida contra a origem. Sem este aviso, o
        # arquivo tinha cara de perfeito — captura_integra=True, zero alertas —
        # e ninguém sabia que o balanço nunca foi checado. Era o único caso em
        # que o programa mentia por omissão: o balancete Real Life saía com 96
        # linhas e nenhuma palavra de que o total não fora verificado. Ver §26.
        result.avisos.append(
            "O TOTAL DA ENTREGA NÃO FOI CONFERIDO contra a origem "
            f"({result.entrega.motivo_nao_conferido}). Não há garantia de que o "
            "Ativo entregue é o Ativo do balancete — confira à mão antes de usar."
        )

    # E a mesma conferência do outro lado. O Balanço pode fechar com a DRE
    # inteira errada: uma receita que entra como custo não move o Ativo.
    nao_coberto_dre = sum(result.nao_coberto_por_natureza.values()) / escala_do_ultimo
    if _origem_com_sinal(hierarquia_do_ultimo):
        nao_coberto_dre = -nao_coberto_dre
    result.dre = conferir_dre(
        wb,
        result.resultado_da_origem,
        coluna=_letra_da_coluna(mapa_ano_exibicao.get(anos[-1], 4)),
        nao_coberto=nao_coberto_dre,
    )
    if result.dre.conferivel and not result.dre.confere:
        for divergente in result.dre.divergentes:
            result.avisos.append(
                f"A DRE NÃO BATE COM A ORIGEM — {divergente}. O Balanço pode "
                f"estar fechando com a DRE errada; confira antes de usar."
            )
    elif not result.dre.conferivel:
        result.avisos.append(
            "A DRE NÃO FOI CONFERIDA contra a origem "
            f"({result.dre.motivo_nao_conferido}). Confira o resultado à mão."
        )

    # 7. Cópia do balancete de origem — rastreio.
    #    Sem ela, responder "de onde saiu este número?" exige reencontrar, meses
    #    depois, o arquivo que gerou a entrega. Com ela, a resposta está no
    #    próprio arquivo, e o SHA-256 prova qual balancete foi usado.
    #    Vem ANTES do Sumário porque o Sumário indexa as origens pelo hash.
    _avisar("Copiando o balancete de origem para dentro da entrega")
    usados: set[str] = set()
    um_so = len(lista) == 1
    for fonte in sorted(lista, key=lambda f: f.ano):
        origem = ler_origem(fonte.path)
        result.origens.append(origem)
        escrever_aba_origem(wb, nome_da_aba(None if um_so else fonte.ano, usados), origem)
        if origem.erro:
            result.avisos.append(
                f"[{fonte.ano}] cópia do original não pôde ser transcrita "
                f"({origem.erro}) — o rastreio fica só pelo SHA-256."
            )

    # 8. Abas de acompanhamento (uso interno).
    _criar_aba_sumario(wb, nome_cliente, data_base, anos, result)
    _criar_aba_tabular(
        wb,
        "Contas Tratadas",
        ["ano", "codigo_original", "descricao_original", "codigo_padronizado",
         "descricao_padronizada", "codigo_template", "valor", "score"],
        todas_tratadas,
        "Auditoria do de-para. Uso interno — não faz parte da entrega ao cliente.",
    )
    _criar_aba_tabular(
        wb,
        "Contas Não Identificadas",
        ["ano", "codigo_original", "descricao_original", "motivo_no_match", "valor"],
        todas_nao_ident,
        "Fila de revisão do analista. Uso interno — não faz parte da entrega.",
    )

    _ordenar_abas(wb)
    _avisar("Salvando a planilha")
    if "_dados_padronizados" in wb.sheetnames:
        wb["_dados_padronizados"].sheet_state = "hidden"

    wb.save(output_path)
    return result


def _letra_da_coluna(indice: int) -> str:
    """Índice 1-based -> letra ('C', 'D', ...). O template não passa de Z."""
    return chr(ord("A") + indice - 1) if 1 <= indice <= 26 else "D"


def _normalizar_fontes(
    fontes: str | Path | FonteBalancete | Sequence[FonteBalancete],
    ano_base: int | None,
    escala: float,
) -> list[FonteBalancete]:
    """Aceita caminho único, uma fonte ou lista; valida anos."""
    if isinstance(fontes, (str, Path)):
        if ano_base is None:
            raise ValueError("ano_base é obrigatório quando 'fontes' é um caminho")
        lista = [FonteBalancete(fontes, ano_base, escala)]
    elif isinstance(fontes, FonteBalancete):
        lista = [fontes]
    else:
        lista = list(fontes)

    if not lista:
        raise ValueError("nenhuma fonte de balancete fornecida")
    anos = [f.ano for f in lista]
    repetidos = {a for a in anos if anos.count(a) > 1}
    if repetidos:
        raise ValueError(f"há mais de um arquivo para o(s) exercício(s) {sorted(repetidos)}")
    for f in lista:
        if not Path(f.path).exists():
            raise FileNotFoundError(f"balancete não encontrado: {f.path}")
    return lista


def _slots_de_ano(wb) -> list[tuple[int, int]]:
    """
    Descobre os pares (coluna de exibição, coluna de dados) lendo as fórmulas.

    Percorre a linha de rótulos de ``BP_GT`` e, para cada coluna, encontra a
    primeira fórmula que referencia ``_dados_padronizados`` para saber qual
    coluna de valores ela soma. É essa referência — e não o rótulo — que
    determina onde escrever.
    """
    from openpyxl.utils import column_index_from_string

    ws = wb["BP_GT"]
    slots: list[tuple[int, int]] = []
    for col in range(1, ws.max_column + 1):
        rotulo = ws.cell(_LINHA_ANOS, col).value
        if not (rotulo and str(rotulo).strip().isdigit()):
            continue
        for row in range(_LINHA_ANOS + 1, ws.max_row + 1):
            f = ws.cell(row, col).value
            if isinstance(f, str) and (m := _REF_RE.search(f)):
                slots.append((col, column_index_from_string(m.group(1))))
                break
    return slots


def _aplicar_anos(
    wb, anos: tuple[int, ...], slots: list[tuple[int, int]]
) -> tuple[dict[int, int], dict[int, int]]:
    """
    Reescreve os rótulos de ano e devolve dois mapas por exercício.

    ``({ano: coluna em _dados_padronizados}, {ano: coluna em BP_GT/DRE_GT})``.
    O primeiro diz **onde escrever**; o segundo, **onde ler o resultado** — é
    a coluna cujas fórmulas a conferência ponta a ponta avalia. Os dois não
    coincidem (o template exibe a partir de D e lê a partir de C), e confundi-
    los faz a conferência avaliar uma coluna vazia e acusar tudo como errado.

    Os anos são atribuídos aos slots em ordem crescente, da esquerda para a
    direita. Slots sobrando são esvaziados, para não exibir um exercício sem
    dados. Só rótulos são tocados — nenhuma fórmula é alterada.
    """
    ws_dados = wb["_dados_padronizados"]
    mapa: dict[int, int] = {}
    exibicao: dict[int, int] = {}

    # Limpa rótulos de ano remanescentes em colunas que nenhum slot ocupa.
    # Sem isso, um template cujas fórmulas não cubram todas as colunas de ano
    # deixaria um cabeçalho órfão (ex.: "2021" numa coluna que ninguém lê).
    colunas_de_slot = {col_dados for _, col_dados in slots}
    for col in range(1, ws_dados.max_column + 1):
        valor = ws_dados.cell(row=1, column=col).value
        if col not in colunas_de_slot and valor and str(valor).strip().isdigit():
            ws_dados.cell(row=1, column=col).value = None

    for i, (col_exib, col_dados) in enumerate(slots):
        ano = anos[i] if i < len(anos) else None
        rotulo = str(ano) if ano else None
        # Atribuição explícita: ws.cell(r, c, None) do openpyxl NÃO limpa a
        # célula (o valor None é ignorado), e slots sobrando ficariam exibindo
        # exercícios sem dado nenhum.
        for aba in ("BP_GT", "DRE_GT"):
            if aba in wb.sheetnames:
                wb[aba].cell(row=_LINHA_ANOS, column=col_exib).value = rotulo
        ws_dados.cell(row=1, column=col_dados).value = rotulo
        if ano:
            mapa[ano] = col_dados
            exibicao[ano] = col_exib
    return mapa, exibicao


# ---------------------------------------------------------------------------
# Padronização
# ---------------------------------------------------------------------------


def _build_matcher(
    plano_path: str | Path | None, cache_path: str | Path | None
) -> ContaMatcher:
    """
    Monta o matcher da geração.

    ``cache_path=None`` usa um cache **efêmero**: gerar a entrega de um cliente
    não deve escrever no cache compartilhado do projeto. O matcher persiste
    toda decisão automática, então um nome próprio de empresa casado por acaso
    viraria uma entrada permanente com score 1.0 — envenenando as próximas
    execuções. Cache é assunto do treino (que tem o seu, isolado), não da
    entrega. Passe um caminho explícito para reaproveitar decisões entre
    execuções do mesmo trabalho.
    """
    import tempfile

    from ..generators.plano_contas import PlanodeContas

    root = Path(__file__).resolve().parent.parent.parent.parent
    plano_path = Path(plano_path) if plano_path else root / "data" / "plano_referencial.json"
    if cache_path is None:
        cache_path = Path(tempfile.mkdtemp(prefix="bp_gt_cache_")) / "cache.json"
    return ContaMatcher(PlanodeContas(plano_path), cache_path=cache_path)


@dataclass
class _Resolucao:
    """Para onde uma conta do cliente vai no template — ou por que não vai."""

    conta: dict[str, Any]
    codigo_template: str | None = None
    decisao: Any = None
    motivo: str = ""


def _resolver(
    conta: dict[str, Any],
    matcher: ContaMatcher,
    projector: TemplateProjector,
    naturezas: dict[str, str] | None = None,
    prazos: dict[str, str] | None = None,
) -> _Resolucao:
    """Match + projeção de UMA conta. Não decide se ela será emitida."""
    descricao = str(conta.get("descricao", ""))
    if not descricao or is_garbage_description(descricao):
        return _Resolucao(conta, motivo="linha sem descrição útil")

    codigo_conta = str(conta.get("codigo", ""))
    natureza_origem = (naturezas or {}).get(codigo_conta)
    prazo_origem = (prazos or {}).get(codigo_conta)
    r = matcher.match(
        descricao,
        codigo_origem=codigo_conta,
        natureza_resultado=natureza_origem,
        prazo=prazo_origem,
    )
    if not r.decision or r.needs_review:
        return _Resolucao(conta, motivo="sem match confiável no plano referencial")

    proj = projector.project(r.decision.codigo)
    if not proj.ok:
        return _Resolucao(
            conta, decisao=r.decision, motivo=f"casou em {r.decision.codigo} mas {proj.motivo}"
        )

    # Trava de classe no ponto FINAL, sobre o código do template.
    #
    # O Plano C já restringe o match por classe, mas ele age sobre a descrição
    # e depende de o código de origem ter sido reconhecido. Quando a coluna de
    # código não é detectada, a classe da origem vira `None`, a restrição
    # desliga e "Aluguel e Condomínio A PAGAR" (passivo) casa com "Condomínio"
    # (despesa 3.x) com score 1.0 — foi assim que custos e despesas foram parar
    # no Balanço. Esta conferência é a última, e é sobre o destino real:
    # conta de resultado não entra em BP_GT, conta patrimonial não entra na DRE.
    #
    # Recusar custa o valor da conta (ela vai para "Contas Não Identificadas" e
    # aparece na reconciliação). Aceitar custa a corretude do Balanço, em
    # silêncio. Recusar é a escolha certa.
    classe_origem = classe_from_codigo(codigo_conta)
    classe_destino = classe_from_codigo(proj.codigo_template)
    if classe_origem and classe_destino and classe_origem != classe_destino:
        return _Resolucao(
            conta,
            decisao=r.decision,
            motivo=(
                f"classe incompatível: a conta é {classe_origem} na origem e o "
                f"match a levaria para {proj.codigo_template} ({classe_destino})"
            ),
        )

    # Trava de natureza, o mesmo raciocínio um nível abaixo. A trava de classe
    # não vê este erro: receita e despesa são ambas RESULTADO. Sem ela, uma
    # receita de serviços de R$ 4,9 milhões entrou na entrega como custo
    # negativo — a DRE errou por DUAS vezes o valor da conta, e o Balanço
    # deixou de fechar por causa disso. Ver REVISAO_QUALIDADE.md §16.
    if natureza_origem:
        natureza_destino = matcher.natureza_referencial.get(r.decision.codigo)
        if natureza_destino and natureza_destino != natureza_origem:
            return _Resolucao(
                conta,
                decisao=r.decision,
                motivo=(
                    f"natureza incompatível: a conta é {natureza_origem} na "
                    f"origem e o match a levaria para {r.decision.codigo} "
                    f"({natureza_destino})"
                ),
            )

    # Trava de prazo, o terceiro eixo. Classe e natureza não veem este erro:
    # circulante e não circulante são ambos ATIVO. Sem ela, "Aplicação
    # Financeira - CDB" foi para Imobilizado e o Ativo Circulante entregue
    # ficou 28,7 milhões menor que o da origem — com o ATIVO TOTAL correto,
    # porque ele é a soma. Ver REVISAO_QUALIDADE.md §18.9.
    if prazo_origem:
        prazo_destino = prazo_do_codigo_referencial(proj.codigo_template)
        if prazo_destino and prazo_destino != prazo_origem:
            return _Resolucao(
                conta,
                decisao=r.decision,
                motivo=(
                    f"prazo incompatível: a conta é {prazo_origem} na origem e "
                    f"o match a levaria para {proj.codigo_template} "
                    f"({prazo_destino})"
                ),
            )

    return _Resolucao(conta, codigo_template=proj.codigo_template, decisao=r.decision)


def _padronizar(
    input_path: Path,
    matcher: ContaMatcher,
    projector: TemplateProjector,
    escala: float,
    aba: str | None = None,
) -> tuple[list[dict], list[dict], list[dict], BuildResult]:
    """
    Parse + match + projeção, **respeitando a hierarquia do balancete**.

    A versão anterior emitia uma linha por conta que casasse, sem olhar a
    árvore. Isso produzia os dois erros opostos ao mesmo tempo:

    - **dupla contagem** quando o agrupador e os filhos casavam (o total de
      ``BANCOS CONTA MOVIMENTO`` já contém as seis contas bancárias abaixo
      dele; emitir os sete valores soma o ramo duas vezes);
    - **valor perdido** quando uma folha com nome próprio não casava — e não
      casa mesmo: "SICOOB - UNISUDESTE - RBM 62540-0" não existe em plano de
      contas nenhum. O valor dela simplesmente sumia do balanço.

    Era a causa de o balanço não fechar. Agora ``selecionar_para_projecao``
    escolhe **um corte** da árvore: o nível mapeado mais alto de cada ramo.
    A conta com nome próprio é absorvida pelo agrupador, e a identidade do
    rollup — conferida antes, em ``conferir_hierarquia`` — garante que o total
    do agrupador é exatamente a soma dos filhos.
    """
    result = BuildResult(output_path=Path())
    contas = ParseyCaller(input_path, aba=aba).parse()
    result.contas_lidas = len(contas)
    result.hierarquia = conferir_hierarquia(contas)

    naturezas = mapear_natureza(contas)
    prazos = mapear_prazo(contas)
    resolucoes = [_resolver(c, matcher, projector, naturezas, prazos) for c in contas]
    por_codigo: dict[str, list[_Resolucao]] = defaultdict(list)
    for res in resolucoes:
        por_codigo[str(res.conta.get("codigo", "")).strip()].append(res)

    def _mapeado(codigo: str) -> bool:
        return any(r.codigo_template for r in por_codigo.get(codigo, ()))

    if result.hierarquia.tem_hierarquia:
        selecao = selecionar_para_projecao(contas, _mapeado)
        a_emitir, nao_cobertos = selecao.codigos, selecao.nao_cobertos
        result.contas_absorvidas = selecao.total_absorvidos
    else:
        # Balancete sem código hierárquico: não há árvore, cada conta responde
        # por si — é o comportamento antigo, e é o certo aqui.
        a_emitir = sorted(c for c in por_codigo if _mapeado(c))
        nao_cobertos = sorted(c for c in por_codigo if not _mapeado(c))

    orientacao = _orientacao_por_classe(result.hierarquia, por_codigo, contas)

    dados: list[dict[str, Any]] = []
    tratadas: list[dict[str, Any]] = []
    nao_ident: list[dict[str, Any]] = []

    for codigo in a_emitir:
        grupo = por_codigo[codigo]
        resolvida = next(r for r in grupo if r.codigo_template)
        # Soma o grupo inteiro: códigos repetidos são normais em balancete real
        # (no RBM, `2.1.1.01.0010` cobre duas contas distintas) e o nó vale a
        # soma do que está sob ele.
        bruto = 0.0
        for r in grupo:
            lido = r.conta.get("saldo")
            if lido is None and "saldo" in r.conta:
                result.saldos_ilegiveis += 1
            bruto += lido or 0.0

        classe = classe_from_codigo(codigo)
        if classe:
            result.emitido_por_classe[classe] = (
                result.emitido_por_classe.get(classe, 0.0) + bruto
            )
        natureza = naturezas.get(codigo)
        if natureza:
            result.emitido_por_natureza[natureza] = (
                result.emitido_por_natureza.get(natureza, 0.0) + bruto
            )

        codigo_template = resolvida.codigo_template
        valor = _valor_para_o_template(
            bruto, escala, codigo, codigo_template, projector, orientacao
        )

        dados.append({
            "codigo_padronizado": codigo_template,
            "descricao_original": str(resolvida.conta.get("descricao", "")),
            "valor": valor,
        })
        tratadas.append({
            "codigo_original": codigo,
            "descricao_original": str(resolvida.conta.get("descricao", "")),
            "codigo_padronizado": resolvida.decisao.codigo,
            "descricao_padronizada": resolvida.decisao.descricao,
            "codigo_template": codigo_template,
            "valor": valor,
            "score": round(resolvida.decisao.score, 2),
        })

        if codigo_template.startswith("1"):
            result.total_ativo += valor
        elif codigo_template.startswith("2"):
            result.total_passivo += valor

    for codigo in nao_cobertos:
        grupo = por_codigo.get(codigo, [])
        if not grupo:
            continue
        r = grupo[0]
        bruto = sum(x.conta.get("saldo") or 0.0 for x in grupo)
        result.valor_nao_coberto += bruto
        classe = classe_from_codigo(codigo)
        if classe:
            result.nao_coberto_por_classe[classe] = (
                result.nao_coberto_por_classe.get(classe, 0.0) + bruto
            )
        natureza = naturezas.get(codigo)
        if natureza:
            result.nao_coberto_por_natureza[natureza] = (
                result.nao_coberto_por_natureza.get(natureza, 0.0) + bruto
            )
        motivo = r.motivo or "sem match confiável no plano referencial"
        result.contas_sem_destino.append(
            ContaSemDestino(
                codigo=codigo,
                descricao=str(r.conta.get("descricao", "")),
                valor=bruto,
                motivo=motivo,
            )
        )
        nao_ident.append({
            "codigo_original": codigo,
            "descricao_original": str(r.conta.get("descricao", "")),
            "motivo_no_match": motivo,
            "valor": _escalar(bruto, escala),
        })

    if result.hierarquia is not None:
        result.resultado_da_origem = _escalar(
            resultado_do_periodo(
                contas,
                result.hierarquia.totais_por_classe.get("RESULTADO", 0.0),
                _origem_com_sinal(result.hierarquia),
            ),
            escala,
        )

    _transferir_resultado_do_periodo(
        contas, result, dados, tratadas, escala, orientacao
    )

    result.contas_tratadas = len(tratadas)
    result.contas_nao_identificadas = len(nao_ident)
    result.linhas_escritas = len(dados)
    result.avisos = _validar(dados, result)
    return dados, tratadas, nao_ident, result


#: Linha do template que recebe o resultado do período de um balancete aberto.
#: É onde ele mora contabilmente enquanto não há encerramento do exercício.
_CODIGO_LUCROS_ACUMULADOS = "2.03.04.01"


def _transferir_resultado_do_periodo(
    contas: list[dict[str, Any]],
    result: BuildResult,
    dados: list[dict[str, Any]],
    tratadas: list[dict[str, Any]],
    escala: float,
    orientacao: dict[str, int],
) -> None:
    """
    Leva ao PL o resultado do período, quando o balancete está **aberto**.

    O problema
    ----------

    Balancete de verificação mensal quase sempre vem aberto: as contas de
    resultado ainda têm saldo e o lucro do período **não foi transferido** para
    o Patrimônio Líquido. Nesse estado, Ativo ≠ Passivo + PL por construção — a
    diferença *é* o resultado.

    Medido no balancete que expôs isso: Ativo 2.361.053,53, Passivo + PL
    891.480,90, diferença 1.469.572,63 — exatamente
    ``Receitas 4.941.899,84 - Despesas 3.472.327,21``. Sem a transferência, a
    entrega mostra "Check: NOK" para um balancete perfeitamente correto.

    Por que a conferência dupla
    ---------------------------

    A diferença ``Ativo - Passivo`` sozinha **não** prova balancete aberto: ela
    também aparece quando a extração perdeu uma conta. Plugar cegamente seria
    fabricar um balanço fechado em cima de um erro — o pior desfecho possível,
    porque esconde justamente o que precisa aparecer.

    Por isso a transferência só acontece quando os dois caminhos independentes
    concordam: a diferença do Balanço e o resultado da DRE. Quando divergem,
    nada é lançado e fica o aviso — falha de forma segura.
    """
    hierarquia = result.hierarquia
    if hierarquia is None or not getattr(hierarquia, "tem_hierarquia", False):
        return

    totais = hierarquia.totais_por_classe
    ativo = totais.get("ATIVO", 0.0) * orientacao.get("ATIVO", 1)
    passivo = totais.get("PASSIVO", 0.0) * orientacao.get("PASSIVO", 1)
    if not ativo and not passivo:
        return
    falta = ativo - passivo
    if abs(falta) <= TOLERANCIA_ORIGEM:
        return  # balancete já encerrado: não há resultado a transferir

    # O resultado pela DRE, caminho independente. O módulo por natureza é
    # proposital: o sinal de cada ramo é convenção do emissor, a grandeza não.
    por_natureza = totais_por_natureza(contas, mapear_natureza(contas))
    receitas = abs(por_natureza.get("RECEITA", 0.0))
    despesas = abs(por_natureza.get("DESPESA", 0.0))
    if not receitas and not despesas:
        return  # sem contas de resultado não há o que transferir
    resultado = receitas - despesas

    if abs(falta - resultado) > max(TOLERANCIA_ORIGEM, abs(falta) * 1e-6):
        result.avisos.append(
            f"O Balanço não fecha por {falta:,.2f}, e o resultado da DRE é "
            f"{resultado:,.2f} — os dois NÃO coincidem. Não é balancete aberto: "
            f"há conta perdida ou contada duas vezes. Nada foi lançado no PL."
        )
        return

    valor = _escalar(falta, escala)
    result.resultado_transferido = valor
    dados.append({
        "codigo_padronizado": _CODIGO_LUCROS_ACUMULADOS,
        "descricao_original": "Resultado do exercício (balancete aberto)",
        "valor": valor,
    })
    tratadas.append({
        "codigo_original": "(gerado)",
        "descricao_original": "Resultado do exercício (balancete aberto)",
        "codigo_padronizado": _CODIGO_LUCROS_ACUMULADOS,
        "descricao_padronizada": "Lucros/prejuízos acumulados",
        "codigo_template": _CODIGO_LUCROS_ACUMULADOS,
        "valor": valor,
        "score": 1.0,
    })


#: Tolerância na moeda da ORIGEM (reais), não na da entrega.
TOLERANCIA_ORIGEM = 0.01

#: Classes cujo destino é o Balanço. O template não traz nenhuma linha "(-)"
#: em BP_GT: lá o sinal de cada conta é o da ORIGEM, não o da linha.
_CLASSES_DO_BALANCO = ("ATIVO", "PASSIVO")

#: Chave da orientação da DRE dentro do mesmo mapa das classes do Balanço.
_RESULTADO = "RESULTADO"


def _orientacao_por_classe(
    hierarquia: Any,
    por_codigo: dict[str, list[_Resolucao]],
    contas: list[dict[str, Any]] | None = None,
) -> dict[str, int]:
    """
    +1 ou -1 por classe, para pôr Ativo e Passivo positivos como o template pede.

    Balancete brasileiro usa duas convenções. Umas trazem Ativo e Passivo
    **ambos positivos** (natureza implícita, a classe é que diz se é devedora
    ou credora); outras trazem o Passivo **negativo**, para que a soma das três
    classes dê zero. O template exige os dois lados positivos — a conferência
    dele é ``ROUND(D26-D52,2)=0``, uma subtração.

    A orientação é decidida pelo **totalizador da própria classe**, não conta a
    conta. É isso que preserva o sinal *relativo* de dentro dela: depreciação
    acumulada é redutora do Imobilizado e continua negativa depois de orientar
    a classe, exatamente como está na origem.

    Só o totalizador **declarado** na origem orienta — e "declarado" exige
    árvore de verdade (``tem_hierarquia``). Sem árvore, as raízes são folhas
    órfãs e somá-las não é totalizador nenhum: num recorte de balancete a soma
    da classe pode dar negativa por acaso, e a inferência inverteria o arquivo
    inteiro. Foi o que aconteceu com o balancete sintético de cinco contas dos
    testes de formato numérico. Sem evidência, não se mexe no sinal — se a
    convenção for outra, a conferência de totais denuncia, o que é muito
    melhor que inverter em silêncio.
    """
    if hierarquia is None or not getattr(hierarquia, "tem_hierarquia", False):
        return {}  # sem árvore não há totalizador declarado: não se mexe no sinal
    totais = dict(hierarquia.totais_por_classe)
    orientacao = {
        classe: (-1 if total < 0 else 1)
        for classe, total in totais.items()
        if classe in _CLASSES_DO_BALANCO
    }
    if _origem_com_sinal(hierarquia):
        # Origem com sinal: o sinal de cada conta de resultado é informação, e
        # apagá-lo inverte um crédito dentro do ramo de despesas.
        orientacao[_RESULTADO] = -1
    return orientacao


def _origem_com_sinal(hierarquia: Any) -> bool:
    """
    A origem traz o Passivo **negativo** — logo, receita credora negativa também.

    É o discriminador único da convenção do arquivo, e vale para o Balanço e
    para a DRE. Ter dois sinais diferentes para a mesma pergunta foi o defeito:
    o Balanço olhava o Passivo, a DRE olhava o sinal das naturezas, e num
    balancete com lucro as duas discordavam — o ramo de resultado classificado
    DESPESA soma negativo (o lucro está lá dentro), a DRE concluía "natureza
    implícita" e a referência saía invertida, acusando 17,27 milhões de erro
    numa entrega correta.
    """
    totais = getattr(hierarquia, "totais_por_classe", {}) or {}
    return totais.get("PASSIVO", 0.0) < 0


def _valor_para_o_template(
    bruto: float,
    escala: float,
    codigo_origem: str,
    codigo_template: str,
    projector: TemplateProjector,
    orientacao: dict[str, int],
) -> float:
    """
    Converte o saldo lido no valor que o template espera — **sem apagar o sinal**.

    A versão anterior fazia ``abs(saldo) * sign_for(linha)``. O ``abs()`` é o
    defeito: ele apaga o sinal da origem e o substitui pelo sinal da *linha do
    template*. Como nenhuma linha do BP_GT é "(-)", toda conta redutora do
    Balanço entrava **positiva**.

    Medido no balancete que expôs isso: depreciação acumulada de -155.617,00 e
    amortização de -5.609,52 entraram positivas, inflando o Ativo em
    ``2 x (155.617,00 + 5.609,52) = 322.453,04`` — o Ativo entregue dava
    2.683.506,57 contra 2.361.053,53 declarados na origem.

    Balanço e DRE seguem regras diferentes, e é por isso que existem duas:

    - **Balanço** — o sinal é o da ORIGEM. A conta redutora é negativa lá e
      tem de continuar negativa aqui, senão o total do grupo não fecha. Só a
      *classe* é orientada, para que Ativo e Passivo saiam ambos positivos
      (ver ``_orientacao_por_classe``).
    - **DRE** — depende da convenção da origem, e são duas:

      * **natureza implícita** (receita e despesa ambas positivas): a origem
        não carrega sinal utilizável e quem decide é o rótulo da linha do
        template (``(-) Despesas com pessoal``), porque as fórmulas da DRE
        **somam**. Aí sim ``abs() * sign_for()``.
      * **origem com sinal** (receita credora negativa, despesa devedora
        positiva): o sinal É utilizável, e apagá-lo repete o defeito do
        Balanço. Foi o que aconteceu com "CREDITO DE PIS E COFINS", saldo
        -191.565,72 dentro do ramo de despesas financeiras: um crédito, que
        REDUZ a despesa. Classificado DESPESA pelo ramo e passado por
        ``abs() * (-1)``, entregou -191.565,72 quando devia entregar
        +191.565,72 — a DRE inteira errou por 2 x 191.565,72 = 383.131,44.
    """
    valor = _escalar(bruto, escala)
    classe = classe_from_codigo(codigo_origem)
    if classe in _CLASSES_DO_BALANCO:
        return valor * orientacao.get(classe, 1)
    if orientacao.get(_RESULTADO):
        # Origem com sinal: preserva-o, como no Balanço.
        return valor * orientacao[_RESULTADO]
    return abs(valor) * projector.sign_for(codigo_template)


def _escalar(saldo: Any, escala: float) -> float:
    try:
        return float(saldo) / escala
    except (TypeError, ValueError):
        return 0.0


def _validar(dados: list[dict], result: BuildResult) -> list[str]:
    """Validações da seção 8 do TEMPLATE_GT_BP.md — avisam, não bloqueiam."""
    avisos: list[str] = []
    if not dados:
        avisos.append("Nenhuma linha escrita — o parser provavelmente falhou.")
    invalidos = [d["codigo_padronizado"] for d in dados
                 if not CODE_RE.fullmatch(str(d["codigo_padronizado"]))]
    if invalidos:
        avisos.append(f"{len(invalidos)} código(s) fora do formato ECF: {invalidos[:3]}")
    nao_num = [d for d in dados if not isinstance(d["valor"], (int, float))]
    if nao_num:
        avisos.append(f"{len(nao_num)} valor(es) não numéricos.")
    if result.saldos_ilegiveis:
        avisos.append(
            f"{result.saldos_ilegiveis} conta(s) com saldo ilegível na origem — "
            "o valor não pôde ser convertido e entrou como zero. Confira o "
            "formato numérico do balancete antes de olhar o resto."
        )
    # A conferência que importa: a origem é aritmeticamente consistente?
    if result.hierarquia is not None and result.hierarquia.tem_hierarquia:
        if not result.hierarquia.rollup_integro:
            pior = result.hierarquia.divergencias[0]
            avisos.append(
                f"O BALANCETE DE ORIGEM não fecha em "
                f"{result.hierarquia.pais_divergentes} agrupador(es): a soma dos "
                f"filhos não bate com o total declarado. Pior caso — {pior}. "
                f"O problema é anterior à padronização; confira o arquivo do "
                f"cliente antes de usar esta saída."
            )
        if not result.hierarquia.equacao_fecha:
            avisos.append(
                f"O BALANCETE DE ORIGEM não fecha: Ativo + Passivo + Resultado = "
                f"{result.hierarquia.desequilibrio:,.2f}, deveria ser zero."
            )

    if not result.cobertura_completa:
        avisos.append(
            "FALHA GRAVE: o total do que foi emitido mais o do que ficou de "
            "fora NÃO reproduz o total da origem. Há conta contada duas vezes "
            "ou perdida no caminho. Não use esta saída."
        )

    reconc = result.reconciliacao
    if not reconc.fecha:
        # Só a frase. As contas em si já saem tabuladas na aba "Sumário"
        # (CONTAS QUE EXPLICAM A DIFERENÇA), em "Contas Não Identificadas", e
        # na lista própria da interface — repetir aqui só polui os três.
        avisos.append(reconc.mensagem())
    return avisos


# ---------------------------------------------------------------------------
# Escrita no workbook
# ---------------------------------------------------------------------------


def _conferir_captura(
    dados: list[dict], projector: TemplateProjector, result: BuildResult
) -> None:
    """
    Confere que cada linha escrita é somada por **exatamente uma** linha do
    template — o elo que faltava entre "o dado foi escrito" e "o dado aparece".

    O ponto cego que isto fecha
    ---------------------------

    Escrever em ``_dados_padronizados`` não põe número nenhum na entrega:
    quem soma são as fórmulas ``SUMIFS(..., $C9&"*")`` do template. Entre as
    duas coisas havia um vão que ninguém media, e ele erra dos dois lados:

    - **nenhuma linha captura** o código → o valor está na aba de dados e não
      aparece em lugar nenhum. Pior: as contagens do Sumário o classificam
      como "conta tratada", então o relatório afirma 100% de match enquanto o
      dinheiro sumiu.
    - **duas linhas capturam** → o mesmo valor entra duas vezes. É o efeito
      que o curinga ``*`` permite quando os prefixos de duas linhas são
      aninhados: ``1.01*`` também casa ``1.01.02``, de modo que uma linha
      "pai" e uma linha "filha" somariam o mesmo dado.

    Hoje o template não tem prefixos aninhados — mas isso é uma *propriedade
    dele*, não uma garantia do sistema, e uma edição no Excel a quebra em
    silêncio. Esta função transforma a propriedade em conferência.
    """
    aninhados = projector.prefixos_aninhados()
    if aninhados:
        a, b = aninhados[0]
        result.avisos.append(
            f"TEMPLATE INCONSISTENTE: {len(aninhados)} par(es) de prefixos "
            f"aninhados — {a!r} também captura {b!r}. Todo valor sob o prefixo "
            f"mais longo é somado em DUAS linhas. Corrija o template."
        )

    for linha in dados:
        codigo = str(linha["codigo_padronizado"])
        valor = linha["valor"] if isinstance(linha["valor"], (int, float)) else 0.0
        quantas = len(projector.linhas_que_capturam(codigo))
        if quantas == 0:
            result.linhas_sem_captura.append((codigo, valor))
        elif quantas > 1:
            result.linhas_capturadas_duas_vezes.append((codigo, valor, quantas))

    if result.linhas_sem_captura:
        codigos = sorted({c for c, _ in result.linhas_sem_captura})
        result.avisos.append(
            f"VALOR PERDIDO NA ENTREGA: {len(result.linhas_sem_captura)} linha(s) "
            f"somando {result.valor_sem_captura:,.2f} foram escritas em "
            f"_dados_padronizados mas NENHUMA linha do template as soma — o "
            f"valor não aparece em BP_GT nem em DRE_GT. Códigos: {codigos[:5]}"
        )
    if result.linhas_capturadas_duas_vezes:
        codigos = sorted({c for c, _, _ in result.linhas_capturadas_duas_vezes})
        result.avisos.append(
            f"DUPLA CONTAGEM: {len(result.linhas_capturadas_duas_vezes)} linha(s) "
            f"são somadas por mais de uma linha do template, inflando a entrega "
            f"em {result.valor_contado_duas_vezes:,.2f}. Códigos: {codigos[:5]}"
        )


def _escrever_dados_padronizados(wb, dados: list[dict], mapa_ano_coluna: dict[int, int]) -> None:
    """
    Limpa e realimenta a aba oculta consultada pelas fórmulas SUMIFS.

    Uma linha por conta de cada exercício — nunca consolidar antes de escrever,
    o wildcard do SUMIFS já agrega (docs/TEMPLATE_GT_BP.md §4.2).
    """
    ws = wb["_dados_padronizados"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    for i, linha in enumerate(dados, start=2):
        col_ano = mapa_ano_coluna.get(linha["ano"])
        if col_ano is None:
            continue
        ws.cell(row=i, column=1, value=linha["codigo_padronizado"])
        ws.cell(row=i, column=2, value=linha["descricao_original"])
        ws.cell(row=i, column=col_ano, value=linha["valor"])


def _criar_aba_sumario(wb, nome_cliente, data_base, anos: tuple[int, ...], result: BuildResult) -> None:
    from openpyxl.styles import Font

    if "Sumário" in wb.sheetnames:
        del wb["Sumário"]
    ws = wb.create_sheet("Sumário", 0)

    ultimo = anos[-1] if anos else "—"
    linhas = [
        ("RESUMO DO PROCESSAMENTO", ""),
        ("Uso interno — a entrega ao cliente são as abas BP_GT e DRE_GT.", ""),
        ("", ""),
        ("Cliente:", nome_cliente),
        ("Data-base:", data_base),
        ("Exercícios preenchidos:", ", ".join(str(a) for a in anos)),
        ("Processado em:", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("", ""),
        ("Contas lidas (todos os exercícios):", result.contas_lidas),
        ("Contas com match:", result.contas_tratadas),
        ("Contas sem match:", result.contas_nao_identificadas),
        ("Match rate:", f"{result.match_rate:.1%}"),
        ("Linhas em _dados_padronizados:", result.linhas_escritas),
        ("", ""),
        (f"Total do Ativo em {ultimo} (milhares):", round(result.total_ativo, 2)),
        (f"Total do Passivo + PL em {ultimo} (milhares):", round(result.total_passivo, 2)),
        ("Balanço confere:", "OK" if result.balanco_confere else "NÃO"),
    ]
    for i, (label, value) in enumerate(linhas, 1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.cell(row=2, column=1).font = Font(italic=True, size=9)

    linha = len(linhas) + 2

    # CONFERÊNCIA DOS TOTAIS — a única medida que lê o número que o cliente lê.
    # Vem antes de tudo porque é a que decide se a entrega presta.
    if result.entrega.conferivel:
        ws.cell(row=linha, column=1, value="CONFERÊNCIA DOS TOTAIS").font = Font(bold=True)
        linha += 1
        for cab, col in (("Total", 1), ("No balancete", 2), ("Na entrega", 3),
                         ("Diferença", 4)):
            ws.cell(row=linha, column=col, value=cab).font = Font(bold=True)
        linha += 1
        for conferencia in result.entrega.conferencias:
            for value, col in (
                (conferencia.nome, 1),
                (round(conferencia.origem, 2), 2),
                (round(conferencia.entrega, 2), 3),
                (round(conferencia.diferenca, 2), 4),
            ):
                celula = ws.cell(row=linha, column=col, value=value)
                if not conferencia.confere:
                    celula.font = Font(bold=True, color="C00000")
            linha += 1
        # A DRE entra na mesma tabela: o Balanço pode fechar com ela errada.
        if result.dre.conferivel:
            for conferencia in result.dre.conferencias:
                for value, col in (
                    (conferencia.nome, 1),
                    (round(conferencia.origem, 2), 2),
                    (round(conferencia.entrega, 2), 3),
                    (round(conferencia.diferenca, 2), 4),
                ):
                    celula = ws.cell(row=linha, column=col, value=value)
                    if not conferencia.confere:
                        celula.font = Font(bold=True, color="C00000")
                linha += 1

        fecha = result.entrega.confere and (
            not result.dre.conferivel or result.dre.confere
        )
        ws.cell(row=linha, column=1, value="Conclusão:")
        ws.cell(
            row=linha,
            column=2,
            value=(
                "os totais entregues são os do balancete"
                if fecha
                else "ATENÇÃO: o número que o cliente lê NÃO é o do balancete"
            ),
        ).font = Font(bold=True)
        linha += 2
    elif result.entrega.motivo_nao_conferido:
        ws.cell(row=linha, column=1, value="CONFERÊNCIA DOS TOTAIS")
        ws.cell(
            row=linha,
            column=2,
            value=f"não conferida — {result.entrega.motivo_nao_conferido}",
        )
        linha += 2

    # ORIGEM DOS DADOS — qual arquivo gerou esta entrega, provado por hash.
    # O conteúdo transcrito está nas abas "Original …"; aqui fica o índice.
    if result.origens:
        ws.cell(row=linha, column=1, value="ORIGEM DOS DADOS").font = Font(bold=True)
        linha += 1
        for cab, col in (("Arquivo", 1), ("SHA-256", 2), ("Linhas copiadas", 3),
                         ("Observação", 4)):
            ws.cell(row=linha, column=col, value=cab).font = Font(bold=True)
        linha += 1
        for origem in result.origens:
            for value, col in (
                (origem.path.name, 1),
                (origem.sha256, 2),
                (len(origem.linhas), 3),
                (origem.erro or origem.procedencia, 4),
            ):
                ws.cell(row=linha, column=col, value=value)
            linha += 1
        linha += 1

    if len(result.por_ano) > 1:
        ws.cell(row=linha, column=1, value="POR EXERCÍCIO").font = Font(bold=True)
        linha += 1
        for cab, col in (("Exercício", 1), ("Lidas", 2), ("Match", 3), ("Sem match", 4)):
            ws.cell(row=linha, column=col, value=cab).font = Font(bold=True)
        linha += 1
        for ano in sorted(result.por_ano):
            lidas, ok, nok = result.por_ano[ano]
            for value, col in ((ano, 1), (lidas, 2), (ok, 3), (nok, 4)):
                ws.cell(row=linha, column=col, value=value)
            linha += 1
        linha += 1

    # RECONCILIAÇÃO — a parte que responde "por que não fecha?".
    # Dizer só "não fecha" não serve para nada: o analista precisa ver que a
    # diferença É exatamente a soma de N contas nomeadas. Quando bate, ele
    # sabe que não há nada escondido e decide com segurança.
    reconc = result.reconciliacao
    if not reconc.fecha:
        ws.cell(row=linha, column=1, value="POR QUE NÃO FECHA").font = Font(bold=True)
        linha += 1
        for label, valor in (
            ("Diferença do balanço:", round(reconc.desequilibrio, 2)),
            ("Soma das contas sem destino:", round(reconc.soma_sem_destino, 2)),
            ("Resíduo sem explicação:", round(reconc.residuo, 2)),
        ):
            ws.cell(row=linha, column=1, value=label)
            ws.cell(row=linha, column=2, value=valor)
            linha += 1
        ws.cell(row=linha, column=1, value="Conclusão:")
        ws.cell(
            row=linha,
            column=2,
            value=(
                f"a diferença é 100% explicada por {len(reconc.contas)} conta(s) "
                f"sem destino — nada mais está faltando"
                if reconc.explicada
                else "ATENÇÃO: sobra diferença sem explicação — há conta "
                "contada duas vezes ou perdida"
            ),
        ).font = Font(bold=True)
        linha += 2

        ws.cell(row=linha, column=1, value="CONTAS QUE EXPLICAM A DIFERENÇA").font = Font(
            bold=True
        )
        linha += 1
        for cab, col in (("Código", 1), ("Descrição", 2), ("Valor", 3), ("Motivo", 4)):
            ws.cell(row=linha, column=col, value=cab).font = Font(bold=True)
        linha += 1
        for conta in reconc.contas:
            for value, col in (
                (conta.codigo, 1),
                (conta.descricao, 2),
                (round(conta.valor, 2), 3),
                (conta.motivo, 4),
            ):
                ws.cell(row=linha, column=col, value=value)
            linha += 1
        ws.cell(row=linha, column=2, value="TOTAL").font = Font(bold=True)
        ws.cell(row=linha, column=3, value=round(reconc.soma_sem_destino, 2)).font = Font(
            bold=True
        )
        linha += 2

    if result.avisos:
        ws.cell(row=linha, column=1, value="AVISOS").font = Font(bold=True)
        for j, aviso in enumerate(result.avisos, start=linha + 1):
            ws.cell(row=j, column=1, value=aviso)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 52


def _criar_aba_tabular(wb, titulo: str, headers: list[str], linhas: list[dict], nota: str) -> None:
    from openpyxl.styles import Font

    if titulo in wb.sheetnames:
        del wb[titulo]
    ws = wb.create_sheet(titulo)

    ws.append([nota])
    ws.cell(row=1, column=1).font = Font(italic=True, size=9)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=2, column=c).font = Font(bold=True)
    for item in linhas:
        ws.append([item.get(h) for h in headers])

    for c, h in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=2, column=c).column_letter].width = max(14, len(h) + 4)
    ws.freeze_panes = "A3"


def _ordenar_abas(wb) -> None:
    """Reordena conforme docs/TEMPLATE_GT_BP.md; ignora abas ausentes.

    As cópias do original entram logo antes das abas técnicas
    (``_instrucoes``/``_dados_padronizados``): são consulta ocasional, não
    devem competir com a entrega, mas também não podem ficar depois de uma
    aba oculta onde ninguém as encontra.
    """
    def _e_origem(nome: str) -> bool:
        return nome.startswith(_PREFIXOS_ORIGEM)

    origens = [wb[n] for n in wb.sheetnames if _e_origem(n)]
    ordem: list = []
    for nome in _ORDEM_ABAS:
        if nome.startswith("_") and origens:
            ordem += origens
            origens = []
        if nome in wb.sheetnames:
            ordem.append(wb[nome])
    ordem += origens
    ordem += [ws for ws in wb.worksheets if ws not in ordem]
    # openpyxl não expõe API pública para reordenar abas; _sheets é o caminho usual.
    wb._sheets = ordem
