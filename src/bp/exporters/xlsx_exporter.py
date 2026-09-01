"""XLSX Exporter — Fase 5

Gera workbook Excel estruturado a partir de um balancete (CSV/XLS/XLSX/PDF).

Features:
 - Inclui contas sintéticas e analíticas (sem filtragem)
 - Matching via ContaMatcher (código/descrição, score, needs_review)
 - Roll-up de saldos: soma dos filhos diretos e comparação com saldo original
 - Validação: lista discrepâncias acima de tolerância
 - Sheets:
     Summary, Accounts, Hierarchy, Unmatched, Variations, Synonyms, Validation
 - Ignora descrições presentes em training_ignore.json (marca como ignored)

Uso programático:
    from src.bp.exporters.xlsx_exporter import export_balance_sheet_to_xlsx
    export_balance_sheet_to_xlsx(Path("arquivo.xlsx"), Path("output/balancete.xlsx"))

CLI (ver auxil/export_xlsx.py)
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:  # pragma: no cover
    Workbook = None  # type: ignore

from ..generators.plano_contas import PlanodeContas
from ..matchers import ContaMatcher
from ..parsers.dispatcher import ParseyCaller  # usar dispatcher estado da arte
from ..utils.normalizer import normalize
from ..utils.numero import parse_saldo
from ..validators.hierarquia import agrupar_por_codigo, mapear_filhos

TOLERANCIA_RELATIVA = 0.0005  # 0.05% diferença relativa aceitável
TOLERANCIA_ABSOLUTA = 0.01  # diferença monetária mínima aceitável


def _load_ignore(training_dir: Path) -> set:
    ignore_file = training_dir / "training_ignore.json"
    if ignore_file.exists():
        try:
            data = json.load(open(ignore_file, encoding="utf-8"))
            return set(data.get("ignored", []))
        except Exception:
            return set()
    return set()


def _build_hierarchy(accounts: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    """
    Índice código -> conta, com ``parent_id`` preenchido.

    Duas correções sobre a versão anterior, ambas medidas no balancete RBM:

    1. ``index[codigo] = c`` era **last-write-wins**. Código repetido é normal
       em balancete real (``2.1.1.01.0010`` cobre duas contas distintas); nove
       códigos se repetem no RBM, e onze contas eram descartadas em silêncio.
       Agora as homônimas são fundidas num nó só, somando os saldos — que é o
       que o rollup precisa: o nó vale a soma do que está sob ele.

    2. O pai era só ``".".join(parts[:-1])``. Se o nível intermediário não
       existe no balancete (``1.1`` e ``1.1.1.02`` presentes, ``1.1.1`` não),
       a subárvore inteira ficava órfã e fora da conferência. ``mapear_filhos``
       sobe até o ancestral mais próximo que existe.
    """
    grupos = agrupar_por_codigo(accounts)

    index: dict[str, dict[str, Any]] = {}
    for codigo, contas_do_codigo in grupos.items():
        principal = dict(contas_do_codigo[0])
        if len(contas_do_codigo) > 1:
            principal["saldo"] = sum(
                parse_saldo(c.get("saldo")) or 0.0 for c in contas_do_codigo
            )
            principal["codigos_homonimos"] = len(contas_do_codigo)
            for campo in ("saldo_anterior", "credito", "debito", "saldo_atual"):
                if any(campo in c for c in contas_do_codigo):
                    principal[campo] = sum(
                        parse_saldo(c.get(campo)) or 0.0 for c in contas_do_codigo
                    )
        principal["parent_id"] = None
        index[codigo] = principal

    for pai, filhos in mapear_filhos(grupos).items():
        for filho in filhos:
            index[filho]["parent_id"] = pai
    return index


#: Campos de saldo consultados por `_primary_saldo`, em ordem de precedência.
_CAMPOS_SALDO = ("saldo_atual", "saldo_atual_ctrl", "saldo_atual_controlada", "saldo")


def _primary_saldo(conta: dict[str, Any]) -> float:
    """
    Saldo principal da conta conforme o Contrato V2.

    A conversão fica em ``utils.numero.parse_saldo`` (fonte única). A closure
    ``_to_float_safe`` que vivia aqui inflava decimais com ponto em 100x,
    devolvia 0.0 para negativo entre parênteses e deixava ``NaN`` passar
    intacto para dentro do rollup.
    """
    for campo in _CAMPOS_SALDO:
        if campo in conta:
            valor = parse_saldo(conta.get(campo))
            if valor is not None:
                return valor
    return 0.0


def _saldo_ilegivel(conta: dict[str, Any]) -> bool:
    """
    A conta declara um saldo que não foi possível ler?

    Distingue "conta zerada" (saldo ausente ou 0) de "saldo corrompido" (campo
    presente, valor ilegível). Sem essa distinção o rollup somava ``NaN`` e
    concluía ``rollup_ok=True``, porque ``abs(nan) > tolerancia`` é ``False``
    em Python — a validação aprovava exatamente os dados em que não se pode
    confiar. Ver REVISAO_QUALIDADE.md §2b.
    """
    presentes = [c for c in _CAMPOS_SALDO if c in conta]
    if not presentes:
        return False
    return all(parse_saldo(conta[c]) is None and conta[c] is not None for c in presentes)


def _compute_rollups(index: dict[str, dict[str, Any]]) -> None:
    """Soma os filhos diretos e compara com o saldo declarado do pai."""
    children_map: dict[str, list[str]] = {}
    for codigo, conta in index.items():
        parent = conta.get("parent_id")
        if parent:
            children_map.setdefault(parent, []).append(codigo)

    for codigo, conta in index.items():
        filhos = children_map.get(codigo, [])
        # Um saldo ilegível — próprio ou de qualquer filho — invalida a
        # conferência: a soma deixa de significar alguma coisa.
        corrompido = _saldo_ilegivel(conta) or any(
            _saldo_ilegivel(index[f]) for f in filhos
        )

        if filhos:
            conta["saldo_calculado"] = sum(_primary_saldo(index[f]) for f in filhos)
        else:
            conta["saldo_calculado"] = _primary_saldo(conta)

        saldo_original = _primary_saldo(conta)
        diff = saldo_original - float(conta["saldo_calculado"])
        conta["rollup_diff"] = diff

        rel = (
            abs(diff) / abs(saldo_original)
            if saldo_original != 0
            else (abs(diff) if abs(diff) > 0 else 0)
        )
        dentro_da_tolerancia = not (
            abs(diff) > TOLERANCIA_ABSOLUTA and rel > TOLERANCIA_RELATIVA
        )
        conta["rollup_ok"] = dentro_da_tolerancia and not corrompido
        if corrompido:
            conta["rollup_motivo"] = "saldo ilegível na conta ou em um filho"


def _detect_analytical_level(accounts: list[dict[str, Any]]) -> int | None:
    """Detecta dinamicamente qual é o nível analítico do balancete.

    Lógica:
    - Agrupa contas por nível
    - Para cada nível, verifica se as contas têm filhos
    - O nível analítico é aquele onde NENHUMA conta tem filhos
    - Se não houver nível puramente analítico, retorna None (todas são sintéticas)

    Args:
        accounts: Lista de contas

    Returns:
        Nível analítico (int) ou None se não houver
    """
    if not accounts:
        return None

    # Agrupa por nível
    from collections import defaultdict

    niveis = defaultdict(list)
    for conta in accounts:
        nivel = conta.get("nivel", 0)
        niveis[nivel].append(conta)

    # Verifica cada nível do mais profundo para o mais raso
    for nivel in sorted(niveis.keys(), reverse=True):
        contas_nivel = niveis[nivel]

        # Verifica se TODAS as contas deste nível não têm filhos
        todas_sem_filhos = True
        for conta in contas_nivel:
            codigo = conta.get("codigo", "")
            tem_filhos = any(
                c.get("codigo", "").startswith(codigo + ".")
                and c.get("codigo") != codigo
                and len(c.get("codigo", "")) > len(codigo)
                for c in accounts
            )
            if tem_filhos:
                todas_sem_filhos = False
                break

        # Se todas as contas deste nível não têm filhos, este é o nível analítico
        if (
            todas_sem_filhos and len(contas_nivel) > 5
        ):  # Pelo menos 5 contas para ser válido
            return nivel

    return None  # Não há nível puramente analítico


def _is_analytical(
    conta: dict[str, Any],
    all_accounts: list[dict[str, Any]],
    analytical_level: int | None = None,
) -> bool:
    """Detecta se conta é analítica (não deve ser mapeada).

    Contas analíticas são detalhes específicos da empresa:
    - Nível hierarquico >= nível analítico detectado dinamicamente
    - Razão social de fornecedor/cliente
    - Número de conta corrente e agência
    - CNPJ/CPF específico

    Args:
        conta: Dicionário da conta
        all_accounts: Lista completa de contas (para detectar filhos)
        analytical_level: Nível analítico detectado dinamicamente

    Returns:
        True se é analítica (não mapear), False se é sintética (mapear)
    """
    import re

    # 1. Verifica se tem filhos (se tem, é sintético)
    codigo = conta.get("codigo", "")
    has_children = any(
        c.get("codigo", "").startswith(codigo + ".")
        and c.get("codigo") != codigo
        and len(c.get("codigo", "")) > len(codigo)
        for c in all_accounts
    )

    if has_children:
        return False  # Tem filhos, é sintético

    # 2. Se foi detectado nível analítico, usar esse critério
    if analytical_level is not None:
        nivel = conta.get("nivel", 0)
        if nivel >= analytical_level:
            return True

    # 3. Padrões de descrição analítica
    descricao = conta.get("descricao", "").upper()

    analytical_patterns = [
        r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}",  # CNPJ
        r"LTDA\.?|S/A|S\.A\.|EIRELI|ME\b|EPP\b",  # Tipo societário
        r"AG[ÊE]NCIA\s*\d+",  # Agência bancária
        r"C/C\s*\d+",  # Conta corrente
        r"CONTA\s*CORRENTE\s*\d+",  # Conta corrente expandido
        r"CPF\s*\d{3}\.\d{3}\.\d{3}-\d{2}",  # CPF
        r"BANCO\s+\d{3}",  # Código de banco (ex: BANCO 001)
    ]

    for pattern in analytical_patterns:
        if re.search(pattern, descricao):
            return True

    # 4. Código muito específico (> 6 partes)
    if codigo and len(codigo.split(".")) > 6:
        return True

    return False


def _match_accounts(
    accounts: list[dict[str, Any]], matcher: ContaMatcher, ignored_set: set
) -> None:
    # Detecta o nível analítico dinamicamente
    analytical_level = _detect_analytical_level(accounts)

    for conta in accounts:
        desc = conta.get("descricao", "")
        norm = normalize(desc)

        # Marca contas analíticas (não devem ser mapeadas)
        conta["is_analytical"] = _is_analytical(conta, accounts, analytical_level)

        if norm in ignored_set:
            conta["ignored"] = True
            continue

        # Pula matching para contas analíticas (não precisam revisão)
        if conta["is_analytical"]:
            conta["match_codigo"] = None
            conta["match_descricao"] = None
            conta["match_score"] = 0.0
            conta["needs_review"] = False  # Analíticas não vão para Unmatched
            conta["ignored"] = False
            continue

        # Apenas contas sintéticas fazem matching
        result = matcher.match(desc, codigo_origem=conta.get("codigo"))
        conta["match_codigo"] = result.decision.codigo if result.decision else None
        conta["match_descricao"] = (
            result.decision.descricao if result.decision else None
        )
        conta["match_score"] = result.decision.score if result.decision else 0.0
        conta["needs_review"] = (
            result.needs_review
        )  # Apenas sintéticas podem precisar revisão
        conta["ignored"] = False


def _create_workbook() -> Workbook:
    if Workbook is None:
        raise RuntimeError("openpyxl não instalado.")
    return Workbook()


def _style_header(ws, row: int = 1):  # pragma: no cover (estético)
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
        )
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _write_summary(wb: Workbook, accounts: list[dict[str, Any]]):
    ws = wb.create_sheet("Summary")
    total = len(accounts)
    analytical = sum(1 for a in accounts if a.get("is_analytical"))
    synthetic = total - analytical
    matched = sum(1 for a in accounts if a.get("match_codigo"))
    needs = sum(1 for a in accounts if a.get("needs_review"))
    ignored = sum(1 for a in accounts if a.get("ignored"))
    discrep = sum(1 for a in accounts if not a.get("rollup_ok", True))
    ws.append(["Metric", "Value"])
    ws.append(["Generated At", datetime.now().isoformat()])
    ws.append(["Total Accounts", total])
    ws.append(["Synthetic (Mappable)", synthetic])
    ws.append(["Analytical (Detail)", analytical])
    ws.append(["Matched", matched])
    ws.append(
        ["Match Rate % (Synthetic)", (matched / synthetic * 100) if synthetic else 0]
    )
    ws.append(["Needs Review", needs])
    ws.append(["Ignored", ignored])
    ws.append(["Rollup Discrepancies", discrep])
    _style_header(ws)


def _detect_balance_type(accounts: list[dict[str, Any]]) -> str:
    """Detecta tipo de balancete baseado nas colunas de saldo presentes.

    Returns:
        'controladora': saldo_anterior_ctrl, saldo_atual_ctrl
        'controlada': saldo_anterior_controlada, saldo_atual_controlada
        'geral': saldo_anterior, credito, debito, saldo_atual
        'unico': apenas saldo (sem histórico)
    """
    if not accounts:
        return "unico"

    # Verifica primeiro item
    first = accounts[0]

    # Controladora
    if "saldo_anterior_ctrl" in first or "saldo_atual_ctrl" in first:
        return "controladora"

    # Controlada
    if "saldo_anterior_controlada" in first or "saldo_atual_controlada" in first:
        return "controlada"

    # Geral (com movimentação)
    if "credito" in first or "debito" in first:
        return "geral"

    # Único (apenas saldo)
    return "unico"


def _write_accounts_sheet(wb: Workbook, accounts: list[dict[str, Any]]):
    """Escreve aba Accounts com 19 colunas obrigatórias conforme contrato V2."""
    ws = wb.create_sheet("Accounts")

    balance_type = _detect_balance_type(accounts)

    # Colunas 1-5: Sempre presentes
    headers = [
        "nivel",  # 1
        "codigo_original",  # 2
        "codigo_alocado",  # 3
        "descricao_original",  # 4
        "descricao_plano_contas",  # 5
    ]

    # Colunas 6-9: Variam conforme tipo de balancete
    if balance_type == "controladora":
        headers.extend(
            [
                "saldo_anterior_ctrl",  # 6
                "saldo_atual_ctrl",  # 9 (colunas 7-8 não aplicáveis)
            ]
        )
    elif balance_type == "controlada":
        headers.extend(
            [
                "saldo_anterior_controlada",  # 6
                "saldo_atual_controlada",  # 9
            ]
        )
    elif balance_type == "geral":
        headers.extend(
            [
                "saldo_anterior",  # 6
                "credito",  # 7
                "debito",  # 8
                "saldo_atual",  # 9
            ]
        )
    else:  # unico
        headers.extend(
            [
                "saldo",  # Saldo único (compatibilidade)
            ]
        )

    # Colunas 10-19: Sempre presentes
    headers.extend(
        [
            "parent_id",  # 10
            "is_analytical",  # 11
            "match_codigo",  # 12
            "match_descricao",  # 13
            "match_score",  # 14
            "needs_review",  # 15
            "ignored",  # 16
            "saldo_somado",  # 17
            "rollup_diff",  # 18
            "rollup_ok",  # 19
        ]
    )

    ws.append(headers)

    # Escrever dados
    for a in accounts:
        row = [
            a.get("nivel"),
            a.get("codigo"),  # codigo_original
            a.get("match_codigo"),  # codigo_alocado (redundância intencional)
            a.get("descricao"),  # descricao_original
            a.get("match_descricao"),  # descricao_plano_contas (redundância)
        ]

        # Adicionar colunas de saldo conforme tipo
        if balance_type == "controladora":
            row.extend(
                [
                    a.get("saldo_anterior_ctrl"),
                    a.get("saldo_atual_ctrl"),
                ]
            )
        elif balance_type == "controlada":
            row.extend(
                [
                    a.get("saldo_anterior_controlada"),
                    a.get("saldo_atual_controlada"),
                ]
            )
        elif balance_type == "geral":
            row.extend(
                [
                    a.get("saldo_anterior"),
                    a.get("credito"),
                    a.get("debito"),
                    a.get("saldo_atual"),
                ]
            )
        else:  # unico
            row.extend([a.get("saldo", 0.0)])

        # Colunas 10-19
        row.extend(
            [
                a.get("parent_id"),
                a.get("is_analytical", False),
                a.get("match_codigo"),
                a.get("match_descricao"),
                a.get("match_score", 0.0),
                a.get("needs_review", False),
                a.get("ignored", False),
                a.get("saldo_calculado", 0.0),  # saldo_somado
                a.get("rollup_diff", 0.0),
                a.get("rollup_ok", True),
            ]
        )

        ws.append(row)

    _style_header(ws)


def _write_hierarchy_sheet(wb: Workbook, index: dict[str, dict[str, Any]]):
    ws = wb.create_sheet("Hierarchy")
    ws.append(
        [
            "codigo",
            "descricao",
            "nivel",
            "parent_id",
            "saldo",
            "saldo_calculado",
            "diff",
            "ok",
        ]
    )
    # Ordena por código para hierarquia previsível
    for codigo in sorted(index.keys()):
        c = index[codigo]
        ws.append(
            [
                codigo,
                c.get("descricao"),
                c.get("nivel"),
                c.get("parent_id"),
                c.get("saldo", 0.0),
                c.get("saldo_calculado", 0.0),
                c.get("rollup_diff", 0.0),
                c.get("rollup_ok", True),
            ]
        )
    _style_header(ws)


def _write_unmatched(wb: Workbook, accounts: list[dict[str, Any]]):
    ws = wb.create_sheet("Unmatched")
    ws.append(
        [
            "codigo",
            "descricao",
            "saldo",
            "nivel",
            "is_analytical",
            "needs_review",
            "match_score",
        ]
    )
    for a in accounts:
        # Apenas contas sintéticas (não analíticas) que precisam revisão
        if a.get("needs_review") and not a.get("is_analytical"):
            ws.append(
                [
                    a.get("codigo"),
                    a.get("descricao"),
                    a.get("saldo", 0.0),
                    a.get("nivel"),
                    a.get("is_analytical", False),
                    a.get("needs_review", False),
                    a.get("match_score", 0.0),
                ]
            )
    _style_header(ws)


def _write_variations(wb: Workbook, training_dir: Path):  # pragma: no cover (I/O)
    var_file = training_dir / "account_variations.json"
    ws = wb.create_sheet("Variations")
    ws.append(["codigo", "frequency", "variations"])
    if var_file.exists():
        data = json.load(open(var_file, encoding="utf-8"))
        for codigo, meta in data.items():
            vars_list = ", ".join(meta.get("variations", [])[:10])
            ws.append([codigo, meta.get("frequency", 0), vars_list])
    _style_header(ws)


def _write_synonyms(wb: Workbook, training_dir: Path):  # pragma: no cover (I/O)
    syn_file = training_dir / "learned_patterns.json"
    ws = wb.create_sheet("Synonyms")
    ws.append(["term", "mapped_terms"])
    if syn_file.exists():
        data = json.load(open(syn_file, encoding="utf-8"))
        for term, mapped in data.get("synonyms", {}).items():
            ws.append([term, ", ".join(mapped[:15])])
    _style_header(ws)


def _write_validation(wb: Workbook, accounts: list[dict[str, Any]]):
    ws = wb.create_sheet("Validation")
    ws.append(
        ["codigo", "descricao", "saldo", "saldo_calculado", "diff", "rel_diff_%", "ok"]
    )
    for a in accounts:
        # Usar saldo principal (detectar qual campo usar)
        saldo = _primary_saldo(a)
        calc = float(a.get("saldo_calculado", 0.0))
        diff = saldo - calc
        rel = (
            (abs(diff) / saldo * 100)
            if saldo
            else (abs(diff) * 100 if abs(diff) > 0 else 0)
        )
        ws.append(
            [
                a.get("codigo"),
                a.get("descricao"),
                saldo,
                calc,
                diff,
                rel,
                a.get("rollup_ok", True),
            ]
        )
    _style_header(ws)


def _write_original(wb: Workbook, original_df: pd.DataFrame | None):
    """Escreve aba Original com dados fonte preservados."""
    ws = wb.create_sheet("Original")

    if original_df is None or original_df.empty:
        ws.append(["No original data available"])
        return

    # Headers: ensure no pandas NA types
    headers = [
        "" if (c is None or (isinstance(c, float) and pd.isna(c))) else str(c)
        for c in list(original_df.columns)
    ]
    ws.append(headers)

    # Data rows: replace pd.NA/NaN with None and convert numpy types to Python scalars
    for _, row in original_df.iterrows():
        values = []
        for v in row.tolist():
            if pd.isna(v):
                values.append(None)
            else:
                try:
                    values.append(v.item() if hasattr(v, "item") else v)
                except Exception:
                    values.append(v)
        ws.append(values)

    _style_header(ws)


def export_balance_sheet_to_xlsx(
    input_path: Path,
    output_path: Path,
    plano_path: Path | None = None,
    training_dir: Path = Path("src/bp/training"),
    auto_match_threshold: float = 0.85,
    requery_threshold: float = 0.60,
    original_data: pd.DataFrame | None = None,
) -> Path:
    """Exporta um balancete único para XLSX estruturado (CONTRATO V2).

    Args:
        input_path: Arquivo de entrada (CSV/XLS/XLSX/PDF)
        output_path: Caminho do XLSX de saída
        plano_path: Caminho do plano de contas (default data/plano_contas.json)
        training_dir: Diretório de treinamento para variações/ignore
        auto_match_threshold: Limite de auto aceitação
        requery_threshold: Limite mínimo para considerar candidatos
        original_data: DataFrame com dados fonte originais (para aba Original)

    Returns:
        Path do arquivo XLSX gerado
    """
    if plano_path is None:
        plano_path = Path("data/plano_contas.json")

    plano = PlanodeContas(plano_path)
    matcher = ContaMatcher(
        plano,
        cache_path=str(training_dir / "training_cache.json"),
        auto_accept_threshold=auto_match_threshold,
        requery_threshold=requery_threshold,
    )

    # Parse via dispatcher, preferindo também obter DataFrame original
    accounts: list[dict[str, Any]]
    original_df_local = original_data
    try:
        if original_df_local is None and hasattr(ParseyCaller, "parse_with_original"):
            accounts, original_df_local = ParseyCaller(input_path).parse_with_original()
        else:
            accounts = ParseyCaller(input_path).parse()
    except Exception:
        accounts = ParseyCaller(input_path).parse()

    # Matching (inclui analíticas)
    ignored_set = _load_ignore(training_dir)
    _match_accounts(accounts, matcher, ignored_set)

    # Hierarquia e rollups
    index = _build_hierarchy(accounts)
    _compute_rollups(index)

    # Workbook
    wb = _create_workbook()
    # Remove sheet default
    default = wb.active
    wb.remove(default)

    _write_summary(wb, accounts)
    _write_accounts_sheet(wb, accounts)
    _write_hierarchy_sheet(wb, index)
    _write_unmatched(wb, accounts)
    _write_variations(wb, training_dir)
    _write_synonyms(wb, training_dir)
    _write_validation(wb, accounts)
    _write_original(wb, original_df_local)  # Nova aba Original

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


__all__ = ["export_balance_sheet_to_xlsx"]
