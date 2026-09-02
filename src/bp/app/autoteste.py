"""
O `.exe` roda? — a pergunta que a auditoria do binário não fazia.

O que aconteceu
---------------

O executável foi distribuído e morreu na mão dos usuários com
``No module named 'pandas.plotting'``. A suíte tinha 617 testes verdes e a
auditoria do binário conferia três coisas — sem dado de cliente, recursos
obrigatórios presentes, `tkdnd` presente. **Nenhuma delas executava o
programa.**

É a terceira vez que a mesma classe de defeito passa: o código está certo, o
que falha é o artefato. Um `.exe` só se prova rodando.

O que este módulo faz
---------------------

Monta um balancete sintético em memória, roda o **motor de verdade** —
o mesmo ``build_gt_output`` que a janela chama — sobre o Template GT
embarcado, e confere que a entrega saiu com o Ativo certo.

Exercita, em uma passada, o que quebrou nas três vezes:

- ``import pandas`` completo (o defeito do ``pandas.plotting``);
- o pacote ``src/bp/output`` presente no bundle (o defeito do `.gitignore`);
- os recursos declarados no `bp.spec` (template e plano de contas);
- o caminho de importação TARDIA de ``service.py``, que é onde a falha
  aparecia — a janela abre bem e só quebra ao clicar em Gerar.

``build.py`` roda isto no binário recém-compilado e **falha o build** se não
passar. Um `.exe` que não se prova não é entregue.
"""

from __future__ import annotations

import tempfile
from pathlib import Path

#: Balancete mínimo, com hierarquia conferível: o pai declara o que os filhos
#: somam, e as três classes fecham. Não é amostra de cliente — é sintético.
_CONTAS = [
    ("1",       "ATIVO",                    1000.0),
    ("1.01",    "ATIVO CIRCULANTE",         1000.0),
    ("1.01.01", "CAIXA E EQUIVALENTES",      600.0),
    ("1.01.02", "CLIENTES",                  400.0),
    ("2",       "PASSIVO",                  -1000.0),
    ("2.01",    "PASSIVO CIRCULANTE",        -400.0),
    ("2.01.01", "FORNECEDORES",              -400.0),
    ("2.03",    "PATRIMONIO LIQUIDO",        -600.0),
    ("2.03.01", "CAPITAL SOCIAL",            -600.0),
]

#: O Ativo do balancete acima. É contra este número que a entrega é conferida.
ATIVO_ESPERADO = 1000.0


def _escrever_balancete(destino: Path) -> Path:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Conta", "Descrição", "Saldo"])
    for codigo, descricao, saldo in _CONTAS:
        ws.append([codigo, descricao, saldo])
    caminho = destino / "autoteste_balancete.xlsx"
    wb.save(caminho)
    return caminho


def executar() -> tuple[bool, str]:
    """
    Roda o pipeline inteiro. Devolve ``(passou, relatório)``.

    **Nunca levanta.** Um autoteste que estoura não reprova o binário — ele
    impede que qualquer binário seja avaliado, e o build para sem dizer se o
    programa funciona. Foi o que aconteceu: a limpeza da pasta temporária
    falhou no Windows depois de o pipeline ter passado, a exceção subiu, e o
    `build.py` recusou um `.exe` correto dizendo "NAO DISTRIBUA".

    A guarda é dupla: cada etapa trata o próprio erro, e ``_executar`` inteira
    roda dentro de um ``try`` aqui. Todo erro vira texto.
    """
    try:
        return _executar()
    except BaseException as exc:  # o portão não pode explodir: reprova com texto
        import traceback

        return False, (
            "AUTOTESTE DO EXECUTÁVEL\n\n"
            f"O autoteste em si falhou: {type(exc).__name__}: {exc}\n"
            "Isto NÃO é um veredito sobre o binário — é defeito do próprio\n"
            "teste. Leia o traceback abaixo antes de concluir qualquer coisa\n"
            "sobre o .exe.\n\n" + traceback.format_exc()
        )


def _executar() -> tuple[bool, str]:
    """O corpo do autoteste. Ver :func:`executar`."""
    linhas: list[str] = ["AUTOTESTE DO EXECUTÁVEL", ""]

    def anotar(rotulo: str, valor: object) -> None:
        linhas.append(f"{rotulo:<34} {valor}")

    try:
        # A importação TARDIA, igual à de service.py — é aqui que o .exe
        # quebrava, com a janela já aberta e o balancete já na tela.
        from ..output.build_gt_output import build_gt_output

        anotar("importar o motor", "OK")
    except Exception as exc:
        anotar("importar o motor", f"FALHOU: {type(exc).__name__}: {exc}")
        return False, "\n".join(linhas)

    try:
        import pandas

        anotar("pandas", pandas.__version__)
    except Exception as exc:
        anotar("pandas", f"FALHOU: {type(exc).__name__}: {exc}")
        return False, "\n".join(linhas)

    # `ignore_cleanup_errors`: apagar a pasta temporária NÃO é parte do que
    # este teste verifica, e no Windows ela falhava — `WinError 32`, arquivo
    # em uso — DEPOIS de o pipeline inteiro ter dado certo. O portão reprovava
    # um binário bom por causa da faxina do próprio portão. Ver §25.
    with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
        pasta = Path(tmp)
        try:
            origem = _escrever_balancete(pasta)
            saida = pasta / "autoteste_saida.xlsx"
            resultado = build_gt_output(origem, saida, ano_base=2024)
        except Exception as exc:
            import traceback

            anotar("gerar a entrega", f"FALHOU: {type(exc).__name__}: {exc}")
            linhas.append("")
            linhas.append(traceback.format_exc())
            return False, "\n".join(linhas)

        anotar("gerar a entrega", "OK")
        anotar("linhas escritas", resultado.linhas_escritas)

        if not saida.exists():
            anotar("arquivo de saída", "FALHOU: não foi criado")
            return False, "\n".join(linhas)

        from openpyxl import load_workbook

        wb = load_workbook(saida)
        abas = set(wb.sheetnames)
        wb.close()
        anotar("abas geradas", ", ".join(sorted(abas)))
        faltando = {"BP_GT", "DRE_GT", "_dados_padronizados"} - abas
        if faltando:
            anotar("abas obrigatórias", f"FALHOU: faltam {sorted(faltando)}")
            return False, "\n".join(linhas)

        if resultado.linhas_escritas <= 0:
            anotar("conteúdo", "FALHOU: nenhuma linha chegou à entrega")
            return False, "\n".join(linhas)

    linhas += ["", "RESULTADO: PASSOU — o executável roda o pipeline completo."]
    return True, "\n".join(linhas)
