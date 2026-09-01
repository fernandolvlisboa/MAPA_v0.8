"""
Ponte entre a janela e o núcleo.

Tudo que a interface precisa saber fazer **fora** de desenhar widget mora aqui:
adivinhar o exercício pelo nome do arquivo, sugerir o nome do cliente, validar a
seleção, escolher o nome do arquivo de saída e chamar ``build_gt_output``.

Duas razões para existir como módulo separado da ``ui.py``:

1. **É testável.** Nada aqui precisa de tela — a suíte cobre os palpites e as
   validações sem abrir janela nenhuma.
2. **É a fronteira.** A janela não importa nada de ``bp.output`` /
   ``bp.matchers`` diretamente. Se o núcleo mudar de assinatura, muda um arquivo.
"""

from __future__ import annotations

import re
import unicodedata
from collections.abc import Callable, Iterable, Sequence
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

#: O que o dispatcher sabe ler (ver src/bp/parsers/dispatcher.py).
EXTENSOES_ACEITAS: tuple[str, ...] = (".xlsx", ".xls", ".csv", ".txt", ".pdf")

#: O Template GT tem cinco colunas de exercício. Mais que isso o núcleo recusa
#: com ValueError; a interface prefere avisar antes de processar.
MAX_EXERCICIOS = 5

ANO_MIN = 1990


def ano_maximo() -> int:
    """Limite superior aceito para um exercício (o ano que vem, no máximo)."""
    return date.today().year + 1


# ---------------------------------------------------------------------------
# Palpites a partir do nome do arquivo
#
# O usuário não deveria digitar o que já está escrito no nome do arquivo que ele
# acabou de arrastar. Todo palpite aqui é *sugestão*: aparece preenchido e ele
# corrige se estiver errado.
# ---------------------------------------------------------------------------

_ANO_RE = re.compile(r"(?:19|20)\d{2}")

_MESES = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}
#: "Dez24", "dez 25", "dez-24" — muito comum em nome de balancete.
_MES_ANO_RE = re.compile(
    r"\b(" + "|".join(_MESES) + r")[a-z]*[\s._-]?(\d{2})\b", re.IGNORECASE
)
#: "3T25", "1T2024" — nomenclatura de release trimestral.
_TRIMESTRE_RE = re.compile(r"\b[1-4]T(\d{2})\b", re.IGNORECASE)

#: Palavras que descrevem o *documento*, não o cliente. Saem do palpite de nome.
_RUIDO_NOME = {
    "balancete", "balancetes", "balanco", "balancos", "balan", "bp", "dre", "df",
    "dfs", "demonstracao", "demonstracoes", "financeiras", "contabil", "razao",
    "consolidado", "consolidada", "controladora", "parecer", "auditado",
    "excel", "planilha", "arquivo", "copia", "final", "ultimo", "ultima",
    "novo", "nova", "revisado", "ajustado", "em", "de", "do", "da", "e", "-",
    "balanc", "patrimonial", "patrimoniais", "resultado", "exercicio", "gt",
    *_MESES,
}


def _sem_acento(texto: str) -> str:
    forma = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in forma if not unicodedata.combining(c))


def ano_do_nome(nome: str) -> int | None:
    """
    Exercício provável a partir do nome do arquivo. ``None`` quando não dá.

    Cobre os formatos que aparecem no corpus real: ``Balancete 2024.xlsx``,
    ``202404_2024.xls``, ``1222024.csv`` (MMAAAA colado), ``2012-12.TXT``,
    ``Balanc dez 25.xls``, ``... Dez24 - Parecer.pdf``.
    """
    texto = _sem_acento(Path(nome).stem)
    limite = ano_maximo()

    achados = [int(m) for m in _ANO_RE.findall(texto)]
    achados = [a for a in achados if ANO_MIN <= a <= limite]
    if achados:
        # Repetição é sinal ("072022 122022" é o mesmo exercício duas vezes);
        # empate fica com o primeiro, que é como as pessoas escrevem a data.
        return max(achados, key=lambda a: (achados.count(a), -achados.index(a)))

    for padrao in (_MES_ANO_RE, _TRIMESTRE_RE):
        achado = padrao.search(texto)
        if achado:
            ano = 2000 + int(achado.groups()[-1])
            if ANO_MIN <= ano <= limite:
                return ano
    return None


def cliente_do_nome(nomes: str | Path | Iterable[str | Path]) -> str:
    """
    Nome de cliente provável a partir do(s) nome(s) de arquivo.

    Tira datas, números soltos e as palavras que descrevem o documento; o que
    sobra costuma ser o cliente (``Balancete 072022 122022 - RBM`` -> ``RBM``).

    Devolve ``""`` quando não sobra nada — de propósito. Chutar "Balancete
    042025 em excel" como nome de cliente é pior do que não chutar: sai
    impresso na capa da entrega. Campo vazio faz a tela pedir que ele digite.
    """
    if isinstance(nomes, (str, Path)):
        nomes = [nomes]
    palpites = [_cliente_de_um(Path(n).stem) for n in nomes]
    validos = [p for p in palpites if p]
    if not validos:
        return ""
    # Série histórica costuma ter o mesmo cliente em todos os arquivos; o
    # palpite que mais se repete é o mais confiável.
    return max(validos, key=validos.count)


def _cliente_de_um(base: str) -> str:
    tokens = [t for t in re.split(r"[\s_\-.,;()\[\]]+", base) if t]
    mantidos = [
        t
        for t in tokens
        if not t.isdigit()
        and not _ANO_RE.fullmatch(t)
        and not _MES_ANO_RE.fullmatch(t)
        and not _TRIMESTRE_RE.fullmatch(t)
        and _sem_acento(t).lower() not in _RUIDO_NOME
        # Fragmento de acento estropiado ("Balanços" -> "Balan_os") vira token
        # curto e minúsculo. Sigla curta em caixa alta (BR, GT) é legítima.
        and (len(t) > 2 or t.isupper())
    ]
    return " ".join(mantidos)


# ---------------------------------------------------------------------------
# Nome e caminho do arquivo entregue
# ---------------------------------------------------------------------------

#: Caracteres que o Windows recusa em nome de arquivo.
_INVALIDOS_RE = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def sanitizar_nome(texto: str) -> str:
    """Transforma um nome digitado em algo que o sistema de arquivos aceita."""
    limpo = _INVALIDOS_RE.sub("", texto).strip(" .")
    limpo = re.sub(r"\s+", " ", limpo)
    return limpo or "Cliente"


def nome_de_saida(cliente: str, anos: Sequence[int]) -> str:
    """``Cliente_2024.xlsx`` ou ``Cliente_2022-2024.xlsx``."""
    base = sanitizar_nome(cliente).replace(" ", "_")
    ordenados = sorted(anos)
    if not ordenados:
        periodo = "sem_ano"
    elif len(ordenados) == 1:
        periodo = str(ordenados[0])
    else:
        periodo = f"{ordenados[0]}-{ordenados[-1]}"
    return f"{base}_{periodo}.xlsx"


def caminho_sem_colisao(pasta: Path, nome: str) -> Path:
    """
    Caminho livre dentro de ``pasta``, acrescentando ``(2)``, ``(3)``...

    Sobrescrever a entrega anterior sem avisar é a forma mais barata de perder
    trabalho já revisado. O app nunca faz isso.
    """
    alvo = pasta / nome
    if not alvo.exists():
        return alvo
    tronco, sufixo = alvo.stem, alvo.suffix
    for n in range(2, 1000):
        candidato = pasta / f"{tronco} ({n}){sufixo}"
        if not candidato.exists():
            return candidato
    raise FileExistsError(f"não achei nome livre para {nome} em {pasta}")


# ---------------------------------------------------------------------------
# Seleção de arquivos e validação
# ---------------------------------------------------------------------------


@dataclass
class Entrada:
    """Um balancete escolhido pelo usuário e o exercício a que ele se refere.

    Uma pasta de trabalho traz vários exercícios em abas ("Balancetes 2020" …
    "Balancetes 2026"). Cada aba marcada vira uma ``Entrada`` própria, com o
    mesmo ``path`` e ``aba`` diferente — daí o mesmo arquivo poder aparecer
    mais de uma vez na lista, um exercício por linha.
    """

    path: Path
    ano: int | None = None
    #: Aba da planilha. ``None`` quando o arquivo traz um exercício só.
    aba: str | None = None

    def __post_init__(self) -> None:
        self.path = Path(self.path)
        if self.ano is None:
            self.ano = ano_do_nome(self.aba or self.path.name)

    @property
    def nome(self) -> str:
        return f"{self.path.name} · {self.aba}" if self.aba else self.path.name


def diagnosticar_arquivo(caminho: str | Path):
    """
    O que o arquivo é, e o que precisa ser perguntado sobre ele.

    Duas perguntas diferentes saem daqui, e a interface as distingue pelo
    diagnóstico:

    - **balancete com vários exercícios** ("Balancetes 2020" … "2026") →
      *quais exercícios usar?*, com o teto de cinco do template;
    - **arquivo que não parece balancete puro** (a empresa já consolidou, numa
      aba "Consolidado" ou "Output Modelo") → *em qual aba está o balanço?*,
      listando **todas** as abas, inclusive as de 36 linhas que o filtro normal
      descartaria.

    Arquivo de aba única com árvore conferível não gera pergunta nenhuma.
    """
    from ..parsers.abas import diagnosticar

    return diagnosticar(caminho)


def selecionar(caminhos: Iterable[str | Path]) -> tuple[list[Entrada], list[Path]]:
    """
    Separa o que dá para ler do que não dá.

    Devolve ``(aceitos, recusados)``. Pasta arrastada vira os arquivos aceitos
    de dentro dela — quem arrasta a pasta quer o conteúdo, não um erro.
    """
    aceitos: list[Entrada] = []
    recusados: list[Path] = []
    vistos: set[Path] = set()

    def _considerar(p: Path) -> None:
        if p in vistos:
            return
        vistos.add(p)
        if p.suffix.lower() in EXTENSOES_ACEITAS:
            aceitos.append(Entrada(p))
        else:
            recusados.append(p)

    for bruto in caminhos:
        caminho = Path(bruto).expanduser()
        if caminho.is_dir():
            for filho in sorted(caminho.iterdir()):
                if filho.is_file():
                    _considerar(filho)
        else:
            _considerar(caminho)
    return aceitos, recusados


def validar(entradas: Sequence[Entrada], cliente: str | None = None) -> list[str]:
    """Problemas que impedem a geração, em português de gente."""
    problemas: list[str] = []
    if not entradas:
        return ["Escolha pelo menos um balancete."]

    # Compara com o texto cru: sanitizar_nome() tem "Cliente" como fallback e
    # transformaria um campo vazio num nome aparentemente válido.
    if cliente is not None and not _INVALIDOS_RE.sub("", cliente).strip(" ."):
        problemas.append("Escreva o nome do cliente — ele vai impresso na entrega.")

    if len(entradas) > MAX_EXERCICIOS:
        problemas.append(
            f"O template comporta {MAX_EXERCICIOS} exercícios e você escolheu "
            f"{len(entradas)} arquivos. Tire {len(entradas) - MAX_EXERCICIOS}."
        )

    sem_ano = [e.nome for e in entradas if not e.ano]
    if sem_ano:
        problemas.append("Informe o exercício de: " + ", ".join(sem_ano))

    duplicadas = [
        e.nome
        for e in entradas
        if sum(1 for o in entradas if o.path == e.path and o.aba == e.aba) > 1
    ]
    if duplicadas:
        problemas.append(
            "A mesma aba foi escolhida duas vezes: " + ", ".join(sorted(set(duplicadas)))
        )

    faltando = [e.nome for e in entradas if not e.path.exists()]
    if faltando:
        problemas.append("Não encontrei mais o arquivo: " + ", ".join(faltando))

    anos = [e.ano for e in entradas if e.ano]
    repetidos = sorted({a for a in anos if anos.count(a) > 1})
    if repetidos:
        problemas.append(
            "Há mais de um arquivo para "
            + ", ".join(str(a) for a in repetidos)
            + ". Cada exercício entra uma vez só."
        )
    return problemas


# ---------------------------------------------------------------------------
# Execução
# ---------------------------------------------------------------------------


@dataclass
class ContaPendente:
    """Uma conta que o sistema não soube classificar — a fila do analista."""

    ano: int | None
    codigo: str
    descricao: str
    motivo: str
    valor: float | None


@dataclass
class Resultado:
    """O que a tela de resultado precisa mostrar, já mastigado."""

    ok: bool
    saida: Path | None = None
    contas_lidas: int = 0
    contas_tratadas: int = 0
    contas_nao_identificadas: int = 0
    saldos_ilegiveis: int = 0
    match_rate: float = 0.0
    balanco_confere: bool = False
    anos: tuple[int, ...] = ()
    alertas: list[str] = field(default_factory=list)
    pendentes: list[ContaPendente] = field(default_factory=list)
    erro: str | None = None

    @property
    def precisa_atencao(self) -> bool:
        """Verde ou amarelo? Amarelo quando a entrega não pode sair como está."""
        return bool(self.alertas) or not self.balanco_confere


def gerar(
    entradas: Sequence[Entrada],
    pasta_saida: Path,
    cliente: str,
    em_milhares: bool = False,
    progresso: Callable[[str], None] | None = None,
) -> Resultado:
    """
    Padroniza os balancetes e devolve o Template GT preenchido.

    Roda em thread de trabalho: **não** toca em widget e nunca levanta exceção
    para cima — falha vira ``Resultado(ok=False, erro=...)``, que a tela sabe
    mostrar sem stacktrace.

    Args:
        entradas: um arquivo por exercício, já validados por :func:`validar`.
        pasta_saida: pasta onde gravar (criada se não existir).
        cliente: nome que vai na célula B4 de BP_GT/DRE_GT.
        em_milhares: True quando o balancete **já** está em milhares — aí o
            valor entra como está. False (o comum) divide por 1.000, porque o
            template diz "Em milhares de reais".
        progresso: callback de texto para a barra de andamento.
    """

    def _passo(texto: str) -> None:
        if progresso:
            progresso(texto)

    try:
        from ..output.build_gt_output import FonteBalancete, build_gt_output
    except Exception as exc:  # instalação quebrada: mensagem, não stacktrace
        return Resultado(ok=False, erro=f"Não consegui carregar o motor do BP: {exc}")

    escala = 1.0 if em_milhares else 1000.0
    fontes = [
        FonteBalancete(e.path, int(e.ano), escala, aba=e.aba)
        for e in entradas
        if e.ano
    ]
    anos = tuple(sorted(f.ano for f in fontes))

    pasta_saida = Path(pasta_saida)
    try:
        pasta_saida.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        return Resultado(ok=False, erro=f"Não consigo gravar em {pasta_saida}: {exc}")

    destino = caminho_sem_colisao(pasta_saida, nome_de_saida(cliente, anos))

    _passo("Lendo os balancetes...")
    try:
        bruto = build_gt_output(
            fontes,
            destino,
            nome_cliente=sanitizar_nome(cliente),
            on_progress=progresso,
        )
    except FileNotFoundError as exc:
        return Resultado(ok=False, erro=f"Arquivo não encontrado: {exc}")
    except PermissionError:
        return Resultado(
            ok=False,
            erro=(
                f"Não consegui gravar {destino.name}. Se ele estiver aberto no "
                "Excel, feche e tente de novo."
            ),
        )
    except ValueError as exc:
        return Resultado(ok=False, erro=str(exc))
    except Exception as exc:  # rede caiu, planilha corrompida, o que for
        return Resultado(ok=False, erro=f"{type(exc).__name__}: {exc}")

    _passo("Conferindo o resultado...")
    return Resultado(
        ok=True,
        saida=bruto.output_path,
        contas_lidas=bruto.contas_lidas,
        contas_tratadas=bruto.contas_tratadas,
        contas_nao_identificadas=bruto.contas_nao_identificadas,
        saldos_ilegiveis=bruto.saldos_ilegiveis,
        match_rate=bruto.match_rate,
        balanco_confere=bruto.balanco_confere,
        anos=bruto.anos,
        alertas=_alertas(bruto.avisos),
        pendentes=ler_pendentes(bruto.output_path),
    )


#: Traduções de aviso do núcleo para a linguagem da tela. O núcleo fala com o
#: analista; a tela fala com quem só quer entregar a planilha.
#: A ordem importa, por dois motivos:
#: 1. A mensagem do caso PERIGOSO ("sobram X sem explicação") contém a do caso
#:    normal ("Não fecha por X ... N conta(s) sem destino"). Se a geral vier
#:    primeiro, ela casa antes e a tela tranquiliza quando devia gritar.
#: 2. O aviso de desequilíbrio *cita* o saldo ilegível como causa provável,
#:    então precisa ser testado antes — senão os dois viram a mesma frase.
def _explica_desequilibrio(m: re.Match[str]) -> str:
    """
    O núcleo prova que a diferença é a soma de N contas. A tela precisa dizer
    isso com os números, não com um "não fechou" genérico: saber que faltam
    R$ X e que são exatamente N contas somando R$ X é o que permite ao usuário
    decidir se entrega ou volta ao balancete.
    """
    return (
        f"O balanço não fechou por {m.group('valor')}. "
        f"São {m.group('n')} conta(s) que o BP não soube classificar, e a soma "
        f"delas dá exatamente esse valor — a lista está abaixo e na aba "
        f"'Sumário' da planilha."
    )


#: Traduções de aviso do núcleo para a linguagem da tela. O núcleo fala com o
#: analista; a tela fala com quem só quer entregar a planilha.
#: A ordem importa: o aviso de desequilíbrio do núcleo *cita* o saldo ilegível
#: como causa provável, então o padrão do desequilíbrio precisa ser testado
#: antes — senão os dois avisos viram a mesma frase e a tela repete.
_TRADUCOES: tuple[tuple[re.Pattern[str], str | Callable[[re.Match[str]], str]], ...] = (
    (
        re.compile(r"sobram .* sem explicação", re.IGNORECASE),
        "ATENÇÃO: o balanço não fechou e a diferença NÃO é explicada pelas "
        "contas sem classificação. Há valor sendo contado duas vezes ou "
        "perdido. Não entregue esta planilha.",
    ),
    (
        re.compile(
            r"Não fecha por (?P<valor>[\d.,]+)\..*?(?P<n>\d+) conta\(s\) sem destino",
            re.IGNORECASE | re.DOTALL,
        ),
        _explica_desequilibrio,
    ),
    (
        re.compile(r"Ativo .* !=\s*Passivo", re.IGNORECASE),
        "O balanço não fechou: Ativo e Passivo+PL ficaram diferentes. "
        "Não entregue antes de resolver.",
    ),
    (
        re.compile(r"saldo ileg[ií]vel", re.IGNORECASE),
        "Algumas contas tinham saldo que não deu para ler e entraram como zero. "
        "Confira o formato dos números no balancete de origem.",
    ),
)


def _alertas(avisos: Sequence[str]) -> list[str]:
    """Traduz e tira repetição — dois avisos técnicos podem virar a mesma frase."""
    saida: list[str] = []
    for aviso in avisos:
        frase = _humanizar(aviso)
        if frase not in saida:
            saida.append(frase)
    return saida


def _humanizar(aviso: str) -> str:
    """Mantém o prefixo do ano e troca o miolo técnico por linguagem de tela."""
    prefixo = ""
    corpo = aviso
    marca = re.match(r"^\[(\d{4})\]\s*", aviso)
    if marca:
        prefixo = f"{marca.group(1)}: "
        corpo = aviso[marca.end() :]
    for padrao, texto in _TRADUCOES:
        achado = padrao.search(corpo)
        if achado:
            return prefixo + (texto(achado) if callable(texto) else texto)
    return prefixo + corpo


def ler_pendentes(planilha: Path, limite: int = 200) -> list[ContaPendente]:
    """
    Lê a aba "Contas Não Identificadas" da planilha recém-gerada.

    A fila de revisão já é escrita pelo núcleo; a tela só a lê de volta em vez
    de pedir ao núcleo um canal novo. Falha aqui não estraga a entrega — o
    arquivo está no disco de qualquer jeito.
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(planilha, read_only=True, data_only=True)
        if "Contas Não Identificadas" not in wb.sheetnames:
            return []
        ws = wb["Contas Não Identificadas"]
        pendentes: list[ContaPendente] = []
        for linha in ws.iter_rows(min_row=3, max_row=2 + limite, values_only=True):
            if linha is None or all(v is None for v in linha):
                continue
            ano, codigo, descricao, motivo, valor = (list(linha) + [None] * 5)[:5]
            pendentes.append(
                ContaPendente(
                    ano=int(ano) if isinstance(ano, (int, float)) else None,
                    codigo=str(codigo or ""),
                    descricao=str(descricao or ""),
                    motivo=str(motivo or ""),
                    valor=float(valor) if isinstance(valor, (int, float)) else None,
                )
            )
        wb.close()
        return pendentes
    except Exception:
        return []


def abrir_no_sistema(caminho: Path) -> bool:
    """Abre arquivo ou pasta no programa padrão do sistema. False se não deu."""
    import subprocess
    import sys

    try:
        if sys.platform == "win32":
            import os

            os.startfile(str(caminho))  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(caminho)])
        else:
            subprocess.Popen(["xdg-open", str(caminho)])
        return True
    except Exception:
        return False


def resumo_selecao(entradas: Sequence[Entrada]) -> dict[str, Any]:
    """Dados que o rodapé da tela de entrada mostra, sem formatar nada."""
    anos = sorted({e.ano for e in entradas if e.ano})
    return {"arquivos": len(entradas), "anos": anos}
